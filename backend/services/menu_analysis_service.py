"""
Menu compliance analysis engine.
Parses uploaded menu files and runs compliance rules against them.
Handles Hebrew rule types: item_frequency_weekly, item_frequency_monthly,
count_min, count_max, item_present, no_repeat_weekly, no_repeat_daily,
no_consecutive, frequency, mandatory.
"""
import re
import csv
import io
import json
import logging
from datetime import date, timedelta
from typing import Optional
from collections import Counter

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, or_

from backend.models.menu_compliance import MenuCheck, MenuDay, CheckResult, ComplianceRule
from backend.models.dish_catalog import DishCatalog
from backend.services.claude_service import claude_service

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Month name → number mapping (English + Hebrew)
# ---------------------------------------------------------------------------
MONTH_NAME_MAP = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
    # Hebrew month names
    "ינואר": 1, "פברואר": 2, "מרץ": 3, "אפריל": 4,
    "מאי": 5, "יוני": 6, "יולי": 7, "אוגוסט": 8,
    "ספטמבר": 9, "אוקטובר": 10, "נובמבר": 11, "דצמבר": 12,
}

# Hebrew plural → singular stems for category matching
HEBREW_PLURAL_MAP = {
    "סלטים": "סלט",
    "מרקים": "מרק",
    "ירקות": "ירק",
    "קינוחים": "קינוח",
    "תוספות": "תוספת",
    "פירות": "פרי",
    "עוגות": "עוגה",
    "לחמים": "לחם",
    "מנות": "מנה",
    "שתיות": "שתיה",
    "דגים": "דג",
    "עופות": "עוף",
    "בשרים": "בשר",
    "קטניות": "קטנית",
    "פחמימות": "פחמימה",
}


def _parse_month_number(month: str) -> int:
    """Convert month string (name or number, possibly with prefix) to int.

    Examples:
        "3" → 3, "03" → 3, "March" → 3, "march" → 3,
        "2025-03" → 3, "מרץ" → 3
    """
    if not month:
        return 1

    # Try "YYYY-MM" format
    if "-" in month:
        part = month.split("-")[-1].strip()
        try:
            return int(part)
        except ValueError:
            pass

    # Try direct int
    try:
        return int(month)
    except ValueError:
        pass

    # Try month name lookup
    cleaned = month.strip().lower()
    if cleaned in MONTH_NAME_MAP:
        return MONTH_NAME_MAP[cleaned]

    # Try Hebrew (no lowering needed — already in map)
    if month.strip() in MONTH_NAME_MAP:
        return MONTH_NAME_MAP[month.strip()]

    logger.warning(f"Could not parse month '{month}', defaulting to 1")
    return 1


AI_MENU_PARSE_PROMPT = """You are a menu parser for Israeli institutional catering.
Parse the following raw menu text into structured JSON.

The menu covers a month of daily meals at a corporate cafeteria in Israel.
Each day should have categories like: עיקרית (main), תוספות (sides), סלטים (salads),
מרק (soup), קינוח (dessert), לחם (bread), שתיה (drinks).

Return ONLY a JSON array of day objects:
[
  {
    "date": "YYYY-MM-DD",
    "day_of_week": "Sunday|Monday|Tuesday|Wednesday|Thursday",
    "items": {
      "category_name": ["item1", "item2"]
    }
  }
]

If dates are not clear, generate weekdays (Sun-Thu) for the given month/year.
If the text is unreadable or empty, return an empty array [].
Do not include Fridays/Saturdays (Shabbat).
"""


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

def _read_file_raw(file_path: str) -> str:
    """Read raw text from a file (CSV, Excel, PDF, or text)."""
    raw_text = ""
    try:
        if file_path.endswith(".csv"):
            for enc in ("utf-8-sig", "cp1255", "latin-1"):
                try:
                    with open(file_path, "r", encoding=enc) as f:
                        raw_text = f.read()
                    if raw_text.strip():
                        break
                except (UnicodeDecodeError, UnicodeError):
                    continue
        elif file_path.endswith((".xlsx", ".xls")):
            raw_text = _read_excel_structured(file_path)
        elif file_path.endswith(".pdf"):
            try:
                import pdfplumber
                with pdfplumber.open(file_path) as pdf:
                    pages = [p.extract_text() or "" for p in pdf.pages]
                    raw_text = "\n".join(pages)
            except ImportError:
                logger.warning("pdfplumber not installed — cannot parse PDF")
        else:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                raw_text = f.read()
    except Exception as e:
        logger.error(f"Failed to read file {file_path}: {e}")

    return raw_text


def _read_excel_structured(file_path: str) -> str:
    """Read Excel file preserving column/row structure for AI parsing."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — cannot parse Excel")
        return ""

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue
            all_text_parts.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                cells = []
                for c in row:
                    if c is not None:
                        val = str(c).strip()
                        if val:
                            cells.append(val)
                if cells:
                    all_text_parts.append(" | ".join(cells))

        wb.close()
        return "\n".join(all_text_parts)
    except Exception as e:
        logger.error(f"Excel parsing error: {e}")
        return ""


def extract_all_dishes_from_file(file_path: str) -> list[str]:
    """Extract all unique dish-like text values directly from a menu file.
    Returns a flat list of unique Hebrew dish names — no AI needed."""
    dishes: set[str] = set()

    try:
        if file_path.endswith((".xlsx", ".xls")):
            dishes = _extract_dishes_from_excel(file_path)
        elif file_path.endswith(".csv"):
            dishes = _extract_dishes_from_csv(file_path)
        else:
            # Plain text — extract Hebrew words/phrases
            raw = _read_file_raw(file_path)
            for line in raw.split("\n"):
                line = line.strip()
                if _is_dish_name(line):
                    dishes.add(line)
    except Exception as e:
        logger.error(f"Dish extraction error: {e}")

    return sorted(dishes)


def _extract_dishes_from_excel(file_path: str) -> set[str]:
    """Extract dish names from Excel cells directly."""
    import openpyxl

    dishes: set[str] = set()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                val = str(cell).strip()
                if _is_dish_name(val):
                    dishes.add(val)

    wb.close()
    return dishes


def _extract_dishes_from_csv(file_path: str) -> set[str]:
    """Extract dish names from CSV cells."""
    dishes: set[str] = set()
    for enc in ("utf-8-sig", "cp1255", "latin-1"):
        try:
            with open(file_path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                for row in reader:
                    for cell in row:
                        val = cell.strip()
                        if _is_dish_name(val):
                            dishes.add(val)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    return dishes


# Hebrew character range check
_HEBREW_RE = re.compile(r'[\u0590-\u05FF]')
# Skip patterns: dates, day names, headers, numbers-only, single chars
_SKIP_PATTERNS = re.compile(
    r'^(?:\d{1,2}[./]\d{1,2}[./]?\d{0,4}|'  # dates
    r'יום\s+[א-ת]|'                           # day names like יום ראשון
    r'None|'
    r'\d+$|'                                    # numbers only
    r'.{0,2}$)'                                 # too short
)
_HEADER_KEYWORDS = {
    "תפריט", "חודש", "שבוע", "תאריך", "הערות", "עמדת שף",
    "sheet", "page", "menu", "date", "week",
}


def _is_dish_name(val: str) -> bool:
    """Heuristic: is this cell value likely a dish name?"""
    if not val or len(val) < 3 or len(val) > 80:
        return False
    if not _HEBREW_RE.search(val):
        return False
    if _SKIP_PATTERNS.match(val):
        return False
    val_lower = val.lower()
    if any(kw in val_lower for kw in _HEADER_KEYWORDS):
        return False
    # Must have at least 2 Hebrew characters
    hebrew_chars = sum(1 for c in val if '\u0590' <= c <= '\u05FF')
    if hebrew_chars < 2:
        return False
    return True


async def parse_menu_file(file_path: str, month: str, year: int) -> list[dict]:
    """Parse a menu file (CSV, Excel, or text) into structured day data."""

    # For Excel files, try direct column extraction first (free, accurate)
    if file_path.endswith((".xlsx", ".xls")):
        structured = _extract_days_from_excel_columns(file_path, month, year)
        if structured and len(structured) >= 3:
            items_count = sum(
                len(items)
                for day in structured
                for items in day.get("items", {}).values()
                if isinstance(items, list)
            )
            if items_count > 10:
                logger.info(f"Column extraction succeeded: {len(structured)} days, {items_count} items")
                return structured
            logger.warning(f"Column extraction found {len(structured)} days but only {items_count} items — trying AI")

    raw_text = _read_file_raw(file_path)

    if not raw_text.strip():
        logger.warning(f"Empty raw text from file: {file_path}")
        return _generate_placeholder_days(month, year)

    logger.info(f"Menu raw text: {len(raw_text)} chars from {file_path}")

    # Send full text to Claude (up to 50K chars — Claude handles large context)
    try:
        parse_prompt = f"""Month: {month}, Year: {year}

Raw menu text:
{raw_text[:50000]}"""

        result = await claude_service.generate_structured_response(
            prompt=parse_prompt,
            system_prompt=AI_MENU_PARSE_PROMPT,
            response_format={"type": "array", "items": {"type": "object"}}
        )

        if isinstance(result, list) and len(result) > 0:
            # Verify the parsed data actually has items
            items_count = sum(
                len(items)
                for day in result
                for items in (day.get("items", {}).values())
                if isinstance(items, list)
            )
            logger.info(f"Claude parsed {len(result)} days with {items_count} total items")
            if items_count > 0:
                return result
            logger.warning("Claude returned days but with 0 items — falling back")
    except Exception as e:
        logger.error(f"Claude menu parsing failed: {e}")

    # Fallback: build days from direct extraction
    logger.info("Using direct extraction fallback")
    return _build_days_from_direct_extraction(file_path, month, year)


def _build_days_from_direct_extraction(
    file_path: str, month: str, year: int
) -> list[dict]:
    """Build day structures by extracting dishes directly from the file.
    Distributes all found dishes across weekdays in the month."""
    all_dishes = extract_all_dishes_from_file(file_path)
    logger.info(f"Direct extraction found {len(all_dishes)} unique dishes")

    if not all_dishes:
        return _generate_placeholder_days(month, year)

    # Generate weekdays for the month
    days = _generate_placeholder_days(month, year)

    # Try to detect column-based day structure from Excel
    if file_path.endswith((".xlsx", ".xls")):
        structured = _extract_days_from_excel_columns(file_path, month, year)
        if structured and any(d.get("items") for d in structured):
            return structured

    # Fallback: put all dishes under "עיקרית" for each day
    # (compliance engine uses substring matching across all items)
    for day in days:
        day["items"] = {"עיקרית": all_dishes}

    return days


def _extract_days_from_excel_columns(
    file_path: str, month: str, year: int
) -> list[dict]:
    """Extract day-structured data from Excel where columns = days.

    Handles multi-sheet weekly menus (common Israeli catering format):
    - Multiple sheets, one per week (שבוע 1, שבוע 2, etc.)
    - Row 1-2: dates and/or day names (ראשון-חמישי)
    - Column A: category labels (מרק, גריל, ציפסר, etc.)
    - Columns B-F: dishes for each weekday
    """
    from datetime import datetime as dt

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Excel open error: {e}")
        return []

    hebrew_days = {"ראשון", "שני", "שלישי", "רביעי", "חמישי"}
    month_num = _parse_month_number(month)
    day_names_map = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    days_data: list[dict] = []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if not hasattr(ws, "iter_rows"):
                continue

            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) < 3:
                continue

            # --- Find date row and day column indices ---
            date_row_idx = -1
            day_col_indices: list[int] = []
            day_dates: list[date | None] = []

            for row_idx in range(min(5, len(all_rows))):
                row = all_rows[row_idx]
                found_dates: list[tuple[int, date | None]] = []
                for col_idx, cell in enumerate(row):
                    if col_idx == 0:
                        continue  # Skip column A (categories)
                    if cell is None:
                        continue
                    # Check for datetime objects (Excel dates)
                    if isinstance(cell, dt):
                        found_dates.append((col_idx, cell.date() if hasattr(cell, 'date') else None))
                        continue
                    cell_str = str(cell).strip()
                    # Check for date strings
                    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', cell_str)
                    if date_match:
                        try:
                            d = date(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
                            found_dates.append((col_idx, d))
                        except ValueError:
                            pass
                        continue
                    date_match = re.search(r'(\d{1,2})[./](\d{1,2})', cell_str)
                    if date_match:
                        try:
                            d = date(year, int(date_match.group(2)), int(date_match.group(1)))
                            found_dates.append((col_idx, d))
                        except ValueError:
                            pass
                        continue

                if len(found_dates) >= 1:
                    # Accept single-date rows too — handles weekly sheets that
                    # contain only one populated day (e.g. a Sunday-only sheet
                    # with May 31 in column B). Without this, the entire sheet
                    # is silently dropped from the compliance check.
                    date_row_idx = row_idx
                    day_col_indices = [f[0] for f in found_dates]
                    day_dates = [f[1] for f in found_dates]
                    break

            # Fallback: look for Hebrew day name row
            if not day_col_indices:
                for row_idx in range(min(5, len(all_rows))):
                    row = all_rows[row_idx]
                    cols = []
                    for col_idx, cell in enumerate(row):
                        if col_idx == 0:
                            continue
                        if cell and any(d in str(cell) for d in hebrew_days):
                            cols.append(col_idx)
                    if len(cols) >= 3:
                        day_col_indices = cols
                        day_dates = [None] * len(cols)
                        break

            if not day_col_indices:
                continue

            # --- Determine data start row (after headers) ---
            data_start = max(date_row_idx + 1, 2)
            # Skip the day-names row if it's right after dates
            if data_start < len(all_rows):
                check_row = all_rows[data_start]
                if any(
                    check_row[c] and any(d in str(check_row[c]) for d in hebrew_days)
                    for c in day_col_indices
                    if c < len(check_row)
                ):
                    data_start += 1

            # --- Extract dishes per day column ---
            for i, col_idx in enumerate(day_col_indices):
                day_items: dict[str, list[str]] = {}

                for row_idx in range(data_start, len(all_rows)):
                    row = all_rows[row_idx]
                    # Get category from column A
                    category = str(row[0]).strip() if row[0] else "עיקרית"
                    category = category if category and category != "None" else "עיקרית"

                    if col_idx >= len(row) or row[col_idx] is None:
                        continue
                    val = str(row[col_idx]).strip()
                    if not val or val == "None" or len(val) < 2:
                        continue
                    # Skip non-dish values
                    if not _HEBREW_RE.search(val):
                        continue

                    if category not in day_items:
                        day_items[category] = []
                    day_items[category].append(val)

                if not any(day_items.values()):
                    continue

                # Determine date
                day_date = day_dates[i] if i < len(day_dates) and day_dates[i] else None
                if not day_date:
                    # Generate date from position
                    day_date = date(year, month_num, min(len(days_data) + 1, 28))

                days_data.append({
                    "date": day_date.isoformat(),
                    "day_of_week": day_names_map[day_date.weekday()],
                    "items": day_items,
                })

        sheet_count = len(wb.sheetnames)
        wb.close()
        logger.info(f"Multi-sheet column extraction: {len(days_data)} days from {sheet_count} sheets")
        return days_data

    except Exception as e:
        logger.error(f"Excel column extraction error: {e}")
        try:
            wb.close()
        except Exception:
            pass
        return []


def _generate_placeholder_days(month: str, year: int) -> list[dict]:
    """Generate placeholder weekday entries when parsing fails."""
    try:
        month_num = _parse_month_number(month)
    except (ValueError, TypeError):
        month_num = 1

    start = date(year, month_num, 1)
    if month_num == 12:
        end = date(year + 1, 1, 1)
    else:
        end = date(year, month_num + 1, 1)

    days = []
    current = start
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    while current < end:
        weekday = current.weekday()
        if weekday < 5:
            days.append({
                "date": current.isoformat(),
                "day_of_week": day_names[weekday],
                "items": {}
            })
        current += timedelta(days=1)

    return days


# ---------------------------------------------------------------------------
# Hebrew rule name parsing
# ---------------------------------------------------------------------------

def _extract_item_and_freq(name: str) -> tuple:
    """Extract (item_keyword, frequency, is_max) from a Hebrew rule name.

    Examples:
        "אסאדו 3 פעמים בחודש"      → ("אסאדו", 3.0, False)
        "המבורגר פעם בשבוע"         → ("המבורגר", 1.0, False)
        "נקניקיות מקסימום פעם בשבוע" → ("נקניקיות", 1.0, True)
        "חזה עוף בגריל 5 פעמים בשבוע" → ("חזה עוף בגריל", 5.0, False)
        "כנאפה אסאדו"              → ("כנאפה", 1.0, False)  # split compound
    """
    is_max = "מקסימום" in name
    clean = name.replace("מקסימום", "").strip()

    # Pattern: "ITEM X.Y פעמ..."
    match = re.match(r'^(.+?)\s+(\d+(?:\.\d+)?)\s+פעמ', clean)
    if match:
        return _clean_keyword(match.group(1).strip()), float(match.group(2)), is_max

    # Pattern: "ITEM פעם ב..."
    match = re.match(r'^(.+?)\s+פעם\s+ב', clean)
    if match:
        return _clean_keyword(match.group(1).strip()), 1.0, is_max

    # No frequency pattern — extract meaningful keyword from compound name
    keyword = _extract_short_keyword(clean)
    return keyword, 1.0, is_max


def _clean_keyword(keyword: str) -> str:
    """Clean up extracted keyword — strip trailing noise words."""
    # Strip trailing "ביום" / "בשבוע" / "בחודש"
    keyword = re.sub(r'\s+(ביום|בשבוע|בחודש|בחודשי|בשבועי)$', '', keyword)
    # Strip trailing "סוגי" (e.g. "11 סוגי סלטים" → already handled by count)
    keyword = re.sub(r'\s+סוגי$', '', keyword)
    return keyword.strip()


def _extract_short_keyword(name: str) -> str:
    """Extract a meaningful short keyword from a compound rule name.

    When there's no frequency pattern, the rule name itself is compound.
    Split on common delimiters and extract the first meaningful Hebrew part.

    Examples:
        "כנאפה אסאדו"               → "כנאפה"  (first word)
        "אין חזרה על מרק"            → "מרק"    (last noun)
        "סלט חומוס יומי"            → "חומוס"   (after prefix strip)
        "לחם שום"                   → "לחם שום" (short enough, keep)
    """
    # Strip "ביום" suffix
    clean = re.sub(r'\s+(ביום|בשבוע|בחודש)$', '', name).strip()

    # Strip common rule prefixes
    prefix_patterns = [
        r'^אין\s+(לחזור|חזרה)\s+(על\s+)?',  # "אין לחזור על..."
        r'^אין\s+אותו\s+',                    # "אין אותו..."
        r'^אין\s+',                            # "אין..."
    ]
    for pattern in prefix_patterns:
        clean = re.sub(pattern, '', clean).strip()

    # If "או" (or) present, split and take first part
    if " או " in clean:
        clean = clean.split(" או ")[0].strip()

    # If "/" present, take first part
    if "/" in clean:
        clean = clean.split("/")[0].strip()

    # If still longer than 3 words, take first 2 meaningful words
    words = clean.split()
    if len(words) > 3:
        clean = " ".join(words[:2])

    return clean.strip() if clean.strip() else name


def _extract_daily_count(name: str) -> tuple:
    """Extract (count, category_keyword) from count_min/count_max names.

    Examples:
        "מינימום 11 סוגי סלטים ביום" → (11, "סלטים")
        "מינימום 2 סוגי מרק ביום"    → (2, "מרק")
        "מקסימום 2 מנות בשר טחון ביום" → (2, "בשר טחון")
    """
    # "מינימום X סוגי CATEGORY ביום"
    match = re.match(r'^מינימום\s+(\d+)\s+סוגי\s+(.+?)(?:\s+ביום)?$', name)
    if match:
        return int(match.group(1)), match.group(2).strip()

    # "מקסימום X מנות ITEM ביום"
    match = re.match(r'^מקסימום\s+(\d+)\s+מנות\s+(.+?)(?:\s+ביום)?$', name)
    if match:
        return int(match.group(1)), match.group(2).strip()

    # "מינימום X CATEGORY"
    match = re.match(r'^מינימום\s+(\d+)\s+(.+)', name)
    if match:
        return int(match.group(1)), match.group(2).strip()

    return 1, name


def _extract_item_present_keyword(name: str) -> str:
    """Extract item keyword from item_present / count_min daily-item rules.

    Examples:
        "סלט חומוס יומי"      → "חומוס"
        "מרק צח/ירקות יומי"   → "מרק צח"
        "גריל יומי"           → "גריל"
        "מנת בקר יומית"       → "בקר"
        "קרוטונים יומיים"     → "קרוטונים"
        "סלט ירקות ביום"      → "ירקות"
    """
    # Strip trailing "ביום", "יומי", "יומית", etc.
    clean = re.sub(r'\s+(יומי|יומית|יומיים|יומיות|ביום)$', '', name).strip()
    # Remove common prefixes like "סלט", "מנת", "מנה"
    for prefix in ["סלט ", "מנת ", "מנה "]:
        if clean.startswith(prefix):
            clean = clean[len(prefix):]
            break
    # Take first part before "/" if compound
    if "/" in clean:
        clean = clean.split("/")[0].strip()
    return clean


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _derive_comparison(actual: int, expected: int) -> str:
    """Derive comparison label: above, under, or even."""
    if actual > expected:
        return "above"
    if actual < expected:
        return "under"
    return "even"


# Mapping from Hebrew keywords in rules → dish catalog categories
KEYWORD_TO_CATALOG_CATEGORY = {
    "בקר": ["protein_beef"],
    "עוף": ["protein_chicken", "chicken_breast", "schnitzel"],
    "שניצל": ["schnitzel"],
    "חזה עוף": ["chicken_breast"],
    "דג": ["fish"],
    "מרק": ["soup"],
    "סלט": ["salads"],
    "סלטים": ["salads"],
    "קינוח": ["desserts"],
    "קינוחים": ["desserts"],
    "טבעוני": ["vegan"],
    "קטניות": ["legumes"],
    "פחמימ": ["carbs"],
    "תוספ": ["side_dish"],
}


def _count_catalog_matches(
    keyword: str,
    item_counter: Counter,
    catalog_map: dict[str, str],
) -> int:
    """Count items matching keyword via catalog category lookup.

    If keyword maps to known catalog categories, also count items
    whose catalog category matches — even if the keyword doesn't
    appear as a substring in the dish name."""
    # Find matching catalog categories
    matched_categories: list[str] = []
    for kw, cats in KEYWORD_TO_CATALOG_CATEGORY.items():
        if kw in keyword.lower() or keyword.lower() in kw:
            matched_categories.extend(cats)

    if not matched_categories:
        return 0

    # Count items whose catalog entry matches these categories
    extra = 0
    for item_name, count in item_counter.items():
        if item_name in catalog_map:
            if catalog_map[item_name] in matched_categories:
                extra += count

    return extra


def _strip_hebrew_prefixes(word: str) -> list[str]:
    """Strip common Hebrew prefix letters and return all possible stems.

    Hebrew has single-letter prefixes that attach directly to words:
    ב (in/with), ה (the), ו (and), ל (to/for), מ (from), כ (like), ש (that)
    And two-letter combos: וב, ול, וה, מה, שב, של, כש, etc.

    Examples:
        "בחריימה"  → ["בחריימה", "חריימה"]
        "והחריימה" → ["והחריימה", "החריימה", "חריימה"]
        "שניצל"    → ["שניצל", "ניצל"]  (but "שניצל" IS a word, so we keep both)
    """
    stems = [word]
    if len(word) <= 2:
        return stems

    # Two-letter prefix combos (check first)
    two_prefixes = ["וב", "ול", "וה", "ומ", "וכ", "וש", "מה", "שב", "של", "כש", "לב"]
    for prefix in two_prefixes:
        if word.startswith(prefix) and len(word) > len(prefix) + 1:
            stem = word[len(prefix):]
            if stem not in stems:
                stems.append(stem)

    # Single-letter prefixes
    single_prefixes = "בהולמכש"
    if word[0] in single_prefixes and len(word) > 2:
        stem = word[1:]
        if stem not in stems:
            stems.append(stem)

    return stems


def _normalize_hebrew(keyword: str) -> list[str]:
    """Return keyword + its singular/plural variants + prefix-stripped forms.

    Examples:
        "סלטים" → ["סלטים", "סלט"]
        "סלט"   → ["סלט", "סלטים"]
        "בשר"   → ["בשר", "בשרים"]
        "מרק"   → ["מרק", "מרקים"]
        "חריימה" → ["חריימה"]  (no plural form)
    """
    variants = [keyword]

    # Check if keyword IS a plural form → add singular
    if keyword in HEBREW_PLURAL_MAP:
        variants.append(HEBREW_PLURAL_MAP[keyword])

    # Check if keyword IS a singular form → add plural
    for plural, singular in HEBREW_PLURAL_MAP.items():
        if keyword == singular and plural not in variants:
            variants.append(plural)

    # Common Hebrew plural suffixes heuristic
    if keyword.endswith("ים") and len(keyword) > 3:
        stem = keyword[:-2]
        if stem not in variants:
            variants.append(stem)
    elif keyword.endswith("ות") and len(keyword) > 3:
        stem = keyword[:-2]
        if stem not in variants:
            variants.append(stem)

    return variants


def _count_item_occurrences(
    item_keyword: str,
    item_counter: Counter,
    catalog_map: dict[str, str] | None = None,
) -> int:
    """Count how many times an item appears using substring match with plural/singular awareness.

    Also checks catalog categories for enhanced matching."""
    variants = _normalize_hebrew(item_keyword.lower())
    matched_items: set[str] = set()
    total = 0

    # 1. Substring match with variants
    for k, v in item_counter.items():
        if any(var in k for var in variants):
            total += v
            matched_items.add(k)

    # 2. Catalog-enhanced matching (only for items not already matched)
    if catalog_map:
        catalog_extra = 0
        for item_name, count in item_counter.items():
            if item_name in matched_items:
                continue  # already counted
            if item_name in catalog_map:
                cat = catalog_map[item_name]
                # Check if rule keyword maps to this catalog category
                for kw, cats in KEYWORD_TO_CATALOG_CATEGORY.items():
                    if (kw in item_keyword.lower() or item_keyword.lower() in kw) and cat in cats:
                        catalog_extra += count
                        break
        total += catalog_extra

    return total


def _matches_any_variant(variants: list[str], text: str) -> bool:
    """Check if any keyword variant appears in the text.

    Uses substring matching AND Hebrew prefix stripping for robust matching.
    E.g., keyword "חריימה" matches text "דג בחריימה" because:
    1. Substring: "חריימה" is in "דג בחריימה" → True
    2. Prefix strip: words ["דג", "בחריימה"] → stems include "חריימה" → True

    IMPORTANT: We do NOT check `stem in var` (whether a text stem appears
    inside the keyword) because it causes massive false positives.
    E.g., "בצל" (onion) → strip prefix → "צל" → "צל" in "צלי כתף" = True!
    """
    text_lower = text.lower()
    # 1. Direct substring match (handles most cases)
    if any(var in text_lower for var in variants):
        return True

    # 2. Word-level prefix stripping — for cases where substring fails
    #    Only check: exact stem match OR variant found inside stem.
    #    Never check stem-inside-variant (causes false positives with
    #    short stems like "צל" matching long keywords like "צלי כתף").
    words = text_lower.split()
    for word in words:
        stems = _strip_hebrew_prefixes(word)
        for stem in stems:
            if any(var == stem or var in stem for var in variants):
                return True

    return False


def _count_daily_item_presence(item_keyword: str, daily_categories: list) -> int:
    """Count how many days an item appears on (with plural/singular awareness)."""
    variants = _normalize_hebrew(item_keyword.lower())
    return sum(
        1 for dc in daily_categories
        if any(_matches_any_variant(variants, str(item)) for _, item in dc["items"])
    )


def _find_item_days(item_keyword: str, daily_categories: list) -> list[str]:
    """Return list of dates where the item appears."""
    variants = _normalize_hebrew(item_keyword.lower())
    return [
        dc["date"] for dc in daily_categories
        if any(_matches_any_variant(variants, str(item)) for _, item in dc["items"])
    ]


def _find_missing_days(item_keyword: str, daily_categories: list) -> list[str]:
    """Return list of dates where the item does NOT appear."""
    variants = _normalize_hebrew(item_keyword.lower())
    return [
        dc["date"] for dc in daily_categories
        if not any(_matches_any_variant(variants, str(item)) for _, item in dc["items"])
    ]


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

async def run_compliance_check(
    check_id: int,
    db: AsyncSession
) -> dict:
    """Run all active compliance rules against a menu check."""
    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()
    if not check:
        raise ValueError(f"MenuCheck {check_id} not found")

    # Load active rules for this site (site-specific + global)
    rules_result = await db.execute(
        select(ComplianceRule).where(
            ComplianceRule.is_active == True,
            or_(ComplianceRule.site_id == None, ComplianceRule.site_id == check.site_id),
        )
    )
    rules = rules_result.scalars().all()

    # Parse the menu file (or re-use existing days if file unavailable)
    days_data = []
    if check.file_path:
        try:
            days_data = await parse_menu_file(check.file_path, check.month, check.year)
        except Exception:
            days_data = []

    if days_data:
        # Fresh parse succeeded — delete old days and store new ones
        from sqlalchemy import delete as sql_delete
        await db.execute(
            sql_delete(MenuDay).where(MenuDay.menu_check_id == check.id)
        )

        week_num = 1
        prev_day = None
        for day_info in days_data:
            day_date_str = day_info.get("date", "")
            try:
                day_date = date.fromisoformat(day_date_str)
            except (ValueError, TypeError):
                continue

            if prev_day and day_date.isocalendar()[1] != prev_day.isocalendar()[1]:
                week_num += 1
            prev_day = day_date

            menu_day = MenuDay(
                menu_check_id=check.id,
                date=day_date,
                day_of_week=day_info.get("day_of_week", ""),
                week_number=week_num,
                is_holiday=False,
                is_theme_day=False,
                menu_items=day_info.get("items", {})
            )
            db.add(menu_day)
    else:
        # File unavailable — re-use existing stored MenuDay data
        existing_days = await db.execute(
            select(MenuDay).where(MenuDay.menu_check_id == check.id)
        )
        stored_days = existing_days.scalars().all()
        for md in stored_days:
            items = md.menu_items or {}
            day_info = {
                "date": md.date.isoformat() if md.date else "",
                "day_of_week": md.day_of_week or "",
                "items": items,
            }
            days_data.append(day_info)

    # Flush parsed days to DB before AI check reads them
    await db.flush()

    # --- AI-powered check (default) ---
    try:
        logger.info(f"Running AI compliance check for check {check_id}")
        ai_results = await run_ai_compliance_check(check_id, db)
        logger.info(f"AI compliance check returned {len(ai_results)} items")

        # Auto-extract dishes to catalog
        extracted = await _auto_extract_dishes_to_catalog(
            days_data, check.id, [], db, file_path=check.file_path
        )

        await db.commit()

        # Re-read updated totals from check object
        await db.refresh(check)
        return {
            "total_findings": check.total_findings or 0,
            "critical_findings": check.critical_findings or 0,
            "warnings": check.warnings or 0,
            "passed_rules": check.passed_rules or 0,
            "days_parsed": len(days_data),
            "rules_checked": len(rules),
            "dishes_above": check.dishes_above or 0,
            "dishes_under": check.dishes_under or 0,
            "dishes_even": check.dishes_even or 0,
            "dishes_extracted": extracted,
            "ai_powered": True,
        }
    except Exception as e:
        logger.warning(f"AI compliance check failed, falling back to rule-based: {e}")

    # --- Fallback: rule-based check (if AI unavailable) ---
    catalog_result = await db.execute(
        select(DishCatalog).where(DishCatalog.category.isnot(None))
    )
    catalog_entries = catalog_result.scalars().all()
    catalog_map = {
        entry.dish_name.lower(): entry.category
        for entry in catalog_entries
        if entry.category
    }
    logger.info(f"Fallback: loaded {len(catalog_map)} categorized dishes from catalog")

    check_results = _evaluate_rules(rules, days_data, catalog_map=catalog_map)

    critical_count = 0
    warning_count = 0
    passed_count = 0
    above_count = 0
    under_count = 0
    even_count = 0

    for cr in check_results:
        evidence = cr.get("evidence") or {}
        if cr.get("rule_id"):
            evidence = {**evidence, "rule_id": cr["rule_id"]}
        result_obj = CheckResult(
            menu_check_id=check.id,
            rule_name=cr["rule_name"],
            rule_category=cr.get("rule_category"),
            passed=cr["passed"],
            severity=cr["severity"],
            finding_text=cr.get("finding_text"),
            evidence=evidence,
            reviewed=False,
        )
        db.add(result_obj)

        if cr["passed"]:
            passed_count += 1
        elif cr["severity"] == "critical":
            critical_count += 1
        else:
            warning_count += 1

        evidence = cr.get("evidence") or {}
        comparison = evidence.get("comparison", "even")
        if comparison == "above":
            above_count += 1
        elif comparison == "under":
            under_count += 1
        else:
            even_count += 1

    check.total_findings = critical_count + warning_count
    check.critical_findings = critical_count
    check.warnings = warning_count
    check.passed_rules = passed_count
    check.dishes_above = above_count
    check.dishes_under = under_count
    check.dishes_even = even_count

    extracted = await _auto_extract_dishes_to_catalog(
        days_data, check.id, check_results, db, file_path=check.file_path
    )

    await db.commit()

    return {
        "total_findings": check.total_findings,
        "critical_findings": critical_count,
        "warnings": warning_count,
        "passed_rules": passed_count,
        "days_parsed": len(days_data),
        "rules_checked": len(rules),
        "dishes_above": above_count,
        "dishes_under": under_count,
        "dishes_even": even_count,
        "dishes_extracted": extracted,
        "ai_powered": False,
    }


# ---------------------------------------------------------------------------
# Auto-extract dishes to catalog
# ---------------------------------------------------------------------------

async def _auto_extract_dishes_to_catalog(
    days_data: list[dict],
    check_id: int,
    check_results: list[dict],
    db: AsyncSession,
    file_path: str | None = None,
) -> int:
    """Extract all unique dish names and add to catalog.
    Uses direct file extraction first (most reliable), then falls back to parsed days.
    Dishes that passed compliance checks are marked as approved."""
    unique_dishes: set[str] = set()

    # 1. Try direct file extraction first (bypasses AI parsing issues)
    if file_path:
        import os
        if os.path.exists(file_path):
            file_dishes = extract_all_dishes_from_file(file_path)
            unique_dishes.update(file_dishes)
            logger.info(f"Direct file extraction: {len(file_dishes)} dishes from {file_path}")

    # 2. Also collect from parsed days_data (may have additional items from AI parsing)
    for day in days_data:
        items = day.get("items", {})
        for _category, item_list in items.items():
            if isinstance(item_list, list):
                for item in item_list:
                    name = str(item).strip()
                    if name and len(name) >= 3:
                        unique_dishes.add(name)
            elif isinstance(item_list, str):
                name = item_list.strip()
                if name and len(name) >= 3:
                    unique_dishes.add(name)

    if not unique_dishes:
        logger.warning(f"No dishes found for check {check_id}")
        return 0

    logger.info(f"Total unique dishes to catalog: {len(unique_dishes)}")

    # Get existing catalog entries
    existing_result = await db.execute(select(DishCatalog.dish_name))
    existing_names = {row[0] for row in existing_result.all()}

    # Collect items that passed compliance (approved)
    approved_keywords: set[str] = set()
    for cr in check_results:
        if cr.get("passed"):
            evidence = cr.get("evidence") or {}
            item = evidence.get("item_searched") or evidence.get("category_keyword", "")
            if item:
                approved_keywords.add(item.lower())

    new_count = 0
    for dish_name in sorted(unique_dishes):
        if dish_name not in existing_names:
            is_approved = any(
                kw in dish_name.lower() for kw in approved_keywords
            ) if approved_keywords else False

            db.add(DishCatalog(
                dish_name=dish_name,
                approved=is_approved,
                source_check_id=check_id,
            ))
            existing_names.add(dish_name)
            new_count += 1

    return new_count


# ---------------------------------------------------------------------------
# Rule evaluation
# ---------------------------------------------------------------------------

def _detect_fallback_mode(days_data: list[dict]) -> bool:
    """Detect if all days have identical items (fallback mode).

    When the direct extraction fallback is used, every day gets the same
    set of dishes. In this mode, no_repeat / no_consecutive rules should
    be treated as "not evaluable" rather than penalized."""
    if len(days_data) < 2:
        return False

    # Compare items of first two days
    items_0 = sorted(
        str(item)
        for item_list in days_data[0].get("items", {}).values()
        if isinstance(item_list, list)
        for item in item_list
    )
    items_1 = sorted(
        str(item)
        for item_list in days_data[1].get("items", {}).values()
        if isinstance(item_list, list)
        for item in item_list
    )

    return items_0 == items_1 and len(items_0) > 10


def _evaluate_rules(
    rules: list,
    days_data: list[dict],
    catalog_map: dict[str, str] | None = None,
) -> list[dict]:
    """Evaluate compliance rules against parsed menu days."""
    results = []
    is_fallback = _detect_fallback_mode(days_data)
    if catalog_map is None:
        catalog_map = {}

    all_items_flat = []
    daily_categories = []
    for day in days_data:
        items = day.get("items", {})
        day_items = []
        for category, item_list in items.items():
            if isinstance(item_list, list):
                for item in item_list:
                    all_items_flat.append(item.lower() if isinstance(item, str) else "")
                    day_items.append((category, item))
        daily_categories.append({
            "date": day.get("date", ""),
            "day_of_week": day.get("day_of_week", ""),
            "categories": list(items.keys()),
            "items": day_items,
        })

    total_days = len(days_data)
    item_counter = Counter(all_items_flat)

    for rule in rules:
        rule_result = _check_single_rule(
            rule, days_data, daily_categories, item_counter, total_days,
            is_fallback=is_fallback,
            catalog_map=catalog_map,
        )
        results.append(rule_result)

    return results


def _check_single_rule(
    rule,
    days_data: list[dict],
    daily_categories: list[dict],
    item_counter: Counter,
    total_days: int,
    is_fallback: bool = False,
    catalog_map: dict[str, str] | None = None,
) -> dict:
    """Check a single compliance rule — handles all Hebrew rule types."""
    params = rule.parameters or {}
    rule_type = rule.rule_type or "mandatory"
    category = rule.category or ""
    name = rule.name or ""

    rule_id = getattr(rule, "id", None)
    base = {
        "rule_name": name,
        "rule_category": category,
        "severity": "critical" if rule.priority <= 1 else "warning",
        "rule_id": rule_id,
    }

    if total_days == 0:
        return {
            **base,
            "passed": False,
            "finding_text": "No menu days found to check against this rule.",
            "evidence": {
                "days_checked": 0,
                "expected_count": 0,
                "actual_count": 0,
                "comparison": "even",
            },
        }

    weeks_in_month = max(total_days / 5, 1)

    # =================================================================
    # item_frequency_weekly — "אטריות מרק 2 פעמים בשבוע"
    # =================================================================
    if rule_type == "item_frequency_weekly":
        item_keyword, freq, is_max = _extract_item_and_freq(name)
        if params.get("item"):
            item_keyword = params["item"]
        if params.get("min_per_week"):
            freq = params["min_per_week"]
            is_max = False
        if params.get("max_per_week"):
            freq = params["max_per_week"]
            is_max = True

        found_days = _find_item_days(item_keyword, daily_categories)
        actual = len(found_days)
        expected = round(freq * weeks_in_month)
        comparison = _derive_comparison(actual, expected)

        if is_max:
            passed = actual <= expected
        else:
            passed = actual >= expected

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"'{item_keyword}': Expected {'≤' if is_max else '≥'}{expected}, "
                f"Actual {actual}"
            ),
            "evidence": {
                "item_searched": item_keyword,
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
                "weekly_freq": freq,
                "is_max_rule": is_max,
                "found_on_days": found_days,
            },
        }

    # =================================================================
    # item_frequency_monthly — "אסאדו 3 פעמים בחודש"
    # =================================================================
    if rule_type == "item_frequency_monthly":
        item_keyword, freq, is_max = _extract_item_and_freq(name)
        if params.get("item"):
            item_keyword = params["item"]
        if params.get("min_per_month"):
            freq = params["min_per_month"]
            is_max = False
        if params.get("max_per_month"):
            freq = params["max_per_month"]
            is_max = True

        found_days = _find_item_days(item_keyword, daily_categories)
        actual = len(found_days)
        expected = round(freq)
        comparison = _derive_comparison(actual, expected)

        if is_max:
            passed = actual <= expected
        else:
            passed = actual >= expected

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"'{item_keyword}': Expected {'≤' if is_max else '≥'}{expected}/month, "
                f"Actual {actual}"
            ),
            "evidence": {
                "item_searched": item_keyword,
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
                "monthly_freq": freq,
                "is_max_rule": is_max,
                "found_on_days": found_days,
            },
        }

    # =================================================================
    # count_min — "מינימום 11 סוגי סלטים ביום" or "מנת בקר יומית"
    # =================================================================
    if rule_type == "count_min":
        # Check if it's a "מינימום X סוגי..." pattern
        min_count, cat_keyword = _extract_daily_count(name)
        has_min_pattern = name.startswith("מינימום")

        if has_min_pattern:
            # Daily variety check — how many distinct items per day match category
            cat_variants = _normalize_hebrew(cat_keyword.lower())
            daily_counts = []
            met_days = []
            unmet_days = []
            for dc in daily_categories:
                day_count = sum(
                    1 for _, item in dc["items"]
                    if _matches_any_variant(cat_variants, str(item))
                    or any(
                        any(var in c.lower() for var in cat_variants)
                        for c in dc["categories"]
                    )
                )
                daily_counts.append(day_count)
                if day_count >= min_count:
                    met_days.append(dc["date"])
                else:
                    unmet_days.append(dc["date"])

            avg_daily = round(sum(daily_counts) / max(len(daily_counts), 1), 1)
            expected = min_count
            actual = round(avg_daily)
            comparison = _derive_comparison(actual, expected)
            passed = avg_daily >= min_count

            return {
                **base,
                "passed": passed,
                "finding_text": (
                    f"'{cat_keyword}': Min {min_count}/day, "
                    f"Avg {avg_daily}/day"
                ),
                "evidence": {
                    "category_keyword": cat_keyword,
                    "expected_count": expected,
                    "actual_count": actual,
                    "comparison": comparison,
                    "avg_daily": avg_daily,
                    "min_required": min_count,
                    "found_on_days": met_days,
                    "missing_on_days": unmet_days[:10],
                },
            }
        else:
            # Daily item presence check — "גריל יומי", "מנת בקר יומית"
            item_keyword = _extract_item_present_keyword(name)
            if params.get("item"):
                item_keyword = params["item"]

            actual = _count_daily_item_presence(item_keyword, daily_categories)
            expected = total_days
            comparison = _derive_comparison(actual, expected)
            passed = actual >= expected
            found_days = _find_item_days(item_keyword, daily_categories)
            missing_days = _find_missing_days(item_keyword, daily_categories)

            return {
                **base,
                "passed": passed,
                "finding_text": (
                    f"'{item_keyword}': Present {actual}/{total_days} days"
                ),
                "evidence": {
                    "item_searched": item_keyword,
                    "expected_count": expected,
                    "actual_count": actual,
                    "comparison": comparison,
                    "found_on_days": found_days,
                    "missing_on_days": missing_days[:10],
                },
            }

    # =================================================================
    # count_max — "מקסימום 2 מנות בשר טחון ביום"
    # =================================================================
    if rule_type == "count_max":
        max_count, item_keyword = _extract_daily_count(name)
        if params.get("item"):
            item_keyword = params["item"]

        daily_counts = []
        exceeded_days = []
        for dc in daily_categories:
            day_count = sum(
                1 for _, item in dc["items"]
                if item_keyword.lower() in str(item).lower()
            )
            daily_counts.append(day_count)
            if day_count > max_count:
                exceeded_days.append(dc["date"])

        max_found = max(daily_counts) if daily_counts else 0
        avg_daily = round(sum(daily_counts) / max(len(daily_counts), 1), 1)
        expected = max_count
        actual = round(avg_daily)
        comparison = _derive_comparison(actual, expected)
        passed = max_found <= max_count

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"'{item_keyword}': Max {max_count}/day, "
                f"Peak {max_found}/day, Avg {avg_daily}/day"
            ),
            "evidence": {
                "item_searched": item_keyword,
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
                "max_daily_found": max_found,
                "avg_daily": avg_daily,
                "found_on_days": exceeded_days,
            },
        }

    # =================================================================
    # item_present — "סלט חומוס יומי", "מרק צח/ירקות יומי"
    # =================================================================
    if rule_type == "item_present":
        item_keyword = _extract_item_present_keyword(name)
        if params.get("item"):
            item_keyword = params["item"]

        actual = _count_daily_item_presence(item_keyword, daily_categories)
        expected = total_days
        comparison = _derive_comparison(actual, expected)
        passed = actual >= expected
        found_days = _find_item_days(item_keyword, daily_categories)
        missing_days = _find_missing_days(item_keyword, daily_categories)

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"'{item_keyword}': Present {actual}/{total_days} days"
            ),
            "evidence": {
                "item_searched": item_keyword,
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
                "found_on_days": found_days,
                "missing_on_days": missing_days[:10],
            },
        }

    # =================================================================
    # no_repeat_weekly — "אין אותו מרק פעמיים בשבוע"
    # =================================================================
    if rule_type == "no_repeat_weekly":
        if is_fallback:
            return {
                **base,
                "passed": True,
                "finding_text": "Cannot evaluate — menu parsed in fallback mode (same items on all days). Re-upload with structured data.",
                "evidence": {
                    "expected_count": 0,
                    "actual_count": 0,
                    "comparison": "even",
                    "note": "Skipped: fallback mode",
                },
            }
        repeats = 0
        weeks: dict[int, list] = {}
        for dc in daily_categories:
            d = dc.get("date", "")
            try:
                week = date.fromisoformat(d).isocalendar()[1]
            except (ValueError, TypeError):
                continue
            if week not in weeks:
                weeks[week] = []
            for _, item in dc["items"]:
                weeks[week].append(str(item).lower())

        for week_items in weeks.values():
            counter = Counter(week_items)
            repeats += sum(v - 1 for v in counter.values() if v > 1)

        expected = 0
        comparison = _derive_comparison(repeats, expected)
        passed = repeats == 0

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"Weekly repeats found: {repeats}"
                if repeats > 0 else "No weekly repeats"
            ),
            "evidence": {
                "expected_count": expected,
                "actual_count": repeats,
                "comparison": comparison,
            },
        }

    # =================================================================
    # no_repeat_daily — "אין לחזור על אותו סלט פעמיים ביום"
    # =================================================================
    if rule_type == "no_repeat_daily":
        if is_fallback:
            return {
                **base,
                "passed": True,
                "finding_text": "Cannot evaluate — menu parsed in fallback mode. Re-upload with structured data.",
                "evidence": {
                    "expected_count": 0,
                    "actual_count": 0,
                    "comparison": "even",
                    "note": "Skipped: fallback mode",
                },
            }
        repeats = 0
        for dc in daily_categories:
            items_lower = [str(item).lower() for _, item in dc["items"]]
            counter = Counter(items_lower)
            repeats += sum(v - 1 for v in counter.values() if v > 1)

        expected = 0
        comparison = _derive_comparison(repeats, expected)
        passed = repeats == 0

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"Daily repeats found: {repeats}"
                if repeats > 0 else "No daily repeats"
            ),
            "evidence": {
                "expected_count": expected,
                "actual_count": repeats,
                "comparison": comparison,
            },
        }

    # =================================================================
    # no_consecutive — "אין לחזור על אותו פריט יום אחר יום"
    # =================================================================
    if rule_type == "no_consecutive":
        if is_fallback:
            return {
                **base,
                "passed": True,
                "finding_text": "Cannot evaluate — menu parsed in fallback mode. Re-upload with structured data.",
                "evidence": {
                    "expected_count": 0,
                    "actual_count": 0,
                    "comparison": "even",
                    "note": "Skipped: fallback mode",
                },
            }
        consecutive_repeats = 0
        prev_items: set = set()
        for dc in daily_categories:
            current_items = {str(item).lower() for _, item in dc["items"]}
            overlap = prev_items & current_items
            consecutive_repeats += len(overlap)
            prev_items = current_items

        expected = 0
        comparison = _derive_comparison(consecutive_repeats, expected)
        passed = consecutive_repeats == 0

        return {
            **base,
            "passed": passed,
            "finding_text": (
                f"Consecutive-day repeats: {consecutive_repeats}"
                if consecutive_repeats > 0 else "No consecutive repeats"
            ),
            "evidence": {
                "expected_count": expected,
                "actual_count": consecutive_repeats,
                "comparison": comparison,
            },
        }

    # =================================================================
    # Legacy: frequency (old seeded rules with parameters)
    # =================================================================
    if rule_type == "frequency":
        max_per_week = params.get("max_per_week")
        min_per_week = params.get("min_per_week")
        target_item = params.get("item", "").lower()

        if target_item:
            actual = _count_item_occurrences(target_item, item_counter, catalog_map)
            weekly_avg = actual / weeks_in_month

            if min_per_week is not None:
                expected = round(min_per_week * weeks_in_month)
            elif max_per_week is not None:
                expected = round(max_per_week * weeks_in_month)
            else:
                expected = 0

            comparison = _derive_comparison(actual, expected)

            if max_per_week and weekly_avg > max_per_week:
                return {
                    **base, "passed": False,
                    "finding_text": (
                        f"'{target_item}': ~{weekly_avg:.1f}/week "
                        f"(max {max_per_week}). Expected ≤{expected}, Actual {actual}"
                    ),
                    "evidence": {
                        "expected_count": expected, "actual_count": actual,
                        "comparison": comparison,
                    },
                }
            if min_per_week and weekly_avg < min_per_week:
                return {
                    **base, "passed": False,
                    "finding_text": (
                        f"'{target_item}': ~{weekly_avg:.1f}/week "
                        f"(min {min_per_week}). Expected ≥{expected}, Actual {actual}"
                    ),
                    "evidence": {
                        "expected_count": expected, "actual_count": actual,
                        "comparison": comparison,
                    },
                }

            return {
                **base, "passed": True,
                "finding_text": f"Expected: {expected}, Actual: {actual}",
                "evidence": {
                    "expected_count": expected, "actual_count": actual,
                    "comparison": comparison,
                },
            }

        return {
            **base, "passed": True, "finding_text": None,
            "evidence": {"expected_count": 0, "actual_count": 0, "comparison": "even"},
        }

    # =================================================================
    # Legacy: mandatory (old seeded rules with parameters)
    # =================================================================
    if rule_type == "mandatory":
        required_category = params.get("required_category", "").lower()
        required_item = params.get("required_item", "").lower()

        if required_category:
            missing_days = []
            for dc in daily_categories:
                cats_lower = [c.lower() for c in dc["categories"]]
                if not any(required_category in c for c in cats_lower):
                    missing_days.append(dc["date"])

            expected = total_days
            actual = total_days - len(missing_days)
            comparison = _derive_comparison(actual, expected)

            if missing_days:
                return {
                    **base, "passed": False,
                    "finding_text": (
                        f"Category '{required_category}' missing on "
                        f"{len(missing_days)}/{total_days} days"
                    ),
                    "evidence": {
                        "expected_count": expected, "actual_count": actual,
                        "comparison": comparison, "missing_days": missing_days[:5],
                    },
                }
            return {
                **base, "passed": True,
                "finding_text": f"Present all {total_days} days",
                "evidence": {
                    "expected_count": expected, "actual_count": actual,
                    "comparison": comparison,
                },
            }

        if required_item:
            actual = _count_daily_item_presence(required_item, daily_categories)
            expected = total_days
            comparison = _derive_comparison(actual, expected)

            if actual == 0:
                return {
                    **base, "passed": False,
                    "finding_text": (
                        f"Required item '{required_item}' not found in any day"
                    ),
                    "evidence": {
                        "expected_count": expected, "actual_count": 0,
                        "comparison": comparison,
                    },
                }
            return {
                **base, "passed": True,
                "finding_text": f"Found on {actual}/{total_days} days",
                "evidence": {
                    "expected_count": expected, "actual_count": actual,
                    "comparison": comparison,
                },
            }

    # =================================================================
    # Unknown rule type — try to treat as item frequency using name parsing
    # =================================================================
    item_keyword, freq, is_max = _extract_item_and_freq(name)
    if freq > 0 and item_keyword:
        actual = _count_item_occurrences(item_keyword, item_counter, catalog_map)
        expected = round(freq)
        comparison = _derive_comparison(actual, expected)
        passed = actual <= expected if is_max else actual >= expected

        return {
            **base,
            "passed": passed,
            "finding_text": f"'{item_keyword}': Expected {expected}, Actual {actual}",
            "evidence": {
                "item_searched": item_keyword,
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
            },
        }

    return {
        **base,
        "passed": True,
        "finding_text": None,
        "evidence": {
            "note": "Rule type not evaluable",
            "expected_count": 0,
            "actual_count": 0,
            "comparison": "even",
        },
    }


# ---------------------------------------------------------------------------
# AI-powered compliance check (replaces keyword matching)
# ---------------------------------------------------------------------------

AI_COMPLIANCE_CHECK_PROMPT = """You are a menu compliance checker for HP Israel corporate catering.

Your job: compare CONTRACT REQUIREMENTS against the ACTUAL MENU and count how many times each required dish appears.

CRITICAL RULES:
1. Count unique DAYS a dish appears (not portions per day)
2. Use intelligent Hebrew matching — the menu often uses different names for the same dish:
   • "בשר צלי מס 5" or "צלי כתף" or "בשר צלי מספר 6" = "צלי כתף מספר 5/6"
   • "נסיכה בחריימה" or "נסיכה ברוטב חריימה" or "דג חריימה" = "חריימה של נסיכה"
   • "שניצל מקסיקני" or "שניצל פריך" = counts as "שניצל בהכנה מקומית"
   • IMPORTANT: "שניצלוני הבית" / "שניצלונים" / "שניצלון" do NOT count toward "שניצל בהכנה מקומית" — they are unapproved (see anomalies)
   • "פרגית על הגריל" or "פרגית ממולאה" or "פרגית במרינדה" = "פרגית צרפתית"
   • "סטייק פרגית" = separate from "פרגית צרפתית" — count each independently
   • "רבעי עוף" or "עופות טבית" or "עוף בגריל מסתובב" = "עופות שלמים"
   • "חזה עוף בגריל במרינדה" or "חזה עוף על הגריל" = "חזה עוף בגריל"
   • "סלט פטריות אסייאתי" = "סלט פטריות אסייתי" (typo tolerance)
   • "עגבניות בצל סגול וזיתי קלמטה" = "סלט עגבניות איטלקי"
   • "עגבניות חריפות" or "עגבניות פיקנטי" = "סלט עגבניות חריף"
   • "ממולאים" includes "זוקיני ממולא" or "מפרום" or "מוסקה"
   • "אסאדו" or "שורט ריבס" or "משפונדרה" = "אסאדו צלוי"
   • "כבד עוף" = "כבד עוף בסילאן עם פירה"
   • "כרע עוף ממולא" or "כרעיים ממולאות" = "כרע עוף ממולא"
   • "בקלאוות בשר" or "בקלווה בשר" = "בקלאוות בשר"
   • "בשר ראש" in any preparation = "בשר ראש"
   • "לשון ברוטב" = "לשון ברוטב פטריות"
   • "שווארמה הודו" or "ירך הודו" = "שווארמה הודו"
   • "שווארמה פרגית" = normally separate from "שווארמה הודו" BUT for NZ site: שווארמה פרגית is an APPROVED substitute for "שווארמה הודו ירך נקבה" — count it and note "הוגשה שווארמה פרגית כתחליף להודו"
   • "סלט קיסר" or "קיסר סלט" = "סלט קיסר"
   • "פיצוחים וירוקים" = look for items with nuts/seeds and greens
   • "סלט אבוקדו" — count any item containing "אבוקדו" appearing in a [סלט מורכב N] / [בר בריא*] row (the menu often abbreviates to just "אבוקדו"). Same principle for ALL salad rules: if a [סלט מורכב N] row contains the keyword (e.g. "ארטישוק", "קיסר", "קינואה", "טבולה", "פטריות אסייתי", "וואקמה", "עגבניות איטלקי"), it counts. Short single-word items in salad rows are still salads.
   • "טבולה" or "בורגול" = "סלט טבולה בורגול"
   • "פיש אנד צ'יפס" or "פיש & צ'יפס" or "דג מטוגן עם צ'יפס" = "פיש & צ'יפס"
   • "קציצות דגים" or "קציצות דג" = "קציצות דגים ביתיות"
   • "סלמון" in any form = "פילה סלמון נורווגי". BUT "שווארמה דג" is NOT salmon/amnon/lavrak
   • "אמנון" or "מושט" = "פילה אמנון". BUT "שווארמה דג" or "מנת דג" is NOT amnon — flag vague names
   • "לברק" = "פילה לברק"
   • "בריסקט" or "חזה בקר" = "בריסקט" (same cut, different names)
   • "סלט קינואה" is NOT "אטריות מרק" — these are completely different dishes
   • "פרגית צרפתית" counts ONLY as פרגית צרפתית — NOT as סטייק פרגית (they are separate dishes)
   • "שווארמה" counts ONLY as שווארמה — "שווארמה בבאגט" or "שווארמה בלאפה" = still שווארמה
3. Frequency conversion (IMPORTANT — use {total_days} working days for this month):
   - "פעם בשבוע" = {total_days} ÷ 5 (rounded) per month (NOT always 4!)
   - "פעמיים בשבוע" = ({total_days} ÷ 5) × 2 per month
   - "שלוש פעמים בשבוע" = ({total_days} ÷ 5) × 3 per month
   - "פעמיים בחודש" = 2/month
   - "שלוש פעמים בחודש" = 3/month
   - "פעם בחודש" = 1/month
   - "פעם ברבעון" = 0 (skip for monthly check)
   - "כל יום" = {total_days}/month
   - "לפחות פעם בשבוע" = at least {total_days} ÷ 5 per month
4. If expected=0, the item is OPTIONAL — skip it (do not include in results)
5. shortage = expected - actual (positive = missing/חוסר, negative = surplus/עודף)
6. OUTPUT FORMAT — קבוצה (column A) MUST be a HEBREW value, never an English DB code. Map the rule's DB category to one of: מיוחדים, סלטים, עוף, בקר, מנות גריל, דגים, קינוחים, חריגים. Mapping table:
   - daily_structure → מיוחדים
   - daily_mandatory → choose by rule name: דג*→דגים, עוף/שניצל/חזה עוף/פרגית→עוף, בקר/בריסקט/אסאדו→בקר, סלט→סלטים, עוגת/קינוח/פירות→קינוחים, otherwise→מיוחדים
   - weekly_frequency / monthly_frequency → choose by rule name (same logic)
   - prohibition → חריגים
   The Hebrew מיוחדים group (column A=מיוחדים) is also where structural rules live (min 2 soups/day, 11 salads/day, etc.).
6a. FREQUENCY TEXT (column C) MUST always be populated. If the rule's parameters.frequency_text is empty, EXTRACT it from the rule name. Examples:
   - rule name "שניצל 5 פעמים בשבוע" → frequency_text "5 פעמים בשבוע"
   - rule name "פירה 4 פעמים בחודש" → frequency_text "4 פעמים בחודש"
   - rule name "מנת דג יומית" or any "*יומי*" → frequency_text "כל יום"
   - rule name "מינימום 2 סוגי מרק ביום" → frequency_text "כל יום (2/יום)"
   - rule name "אין אותו מרק פעמיים בשבוע" → frequency_text "מקסימום פעם בשבוע"
6b. POSITION-BASED COUNTING — CRITICAL RULE. The menu is given to you with row category labels in [brackets]. For EVERY rule below, you MUST count ONLY items in the listed [row labels]. Do NOT keyword-search across the whole day. Do NOT count items embedded inside protein dish names (e.g. "פרגית צלויה בגריל לצד סלט עגבניות חריף" is ONE protein dish in [עמדת גריל] — the "סלט עגבניות חריף" inside it does NOT count toward the salad rule). Do NOT count vegetables embedded inside beef dishes (e.g. "בשר גולש מס 10 עם ירקות" is ONE beef dish, not a vegetable serving).

   Daily rules — strict row-only matching:
   - "פחמימה מלאה ביום" → ONLY count from [פחמימה מלאה] row. expected = total_days, count = number of days where this row is non-empty
   - "קטנייה ביום" → ONLY count from [קיטניה] / [קטניות] row
   - "מנת דג יומית" → ONLY count from [בריאות מנת דג] / [מנת דג*] rows. Every working day this row will have a fish item — expected = total_days, count = total_days
   - "מנה טבעונית יומית" → ONLY count from [בריאות מנה טבעונית] / [מנה טבעונית*] rows. Vegan dishes include שווארמה טבעונית, המבורגר טבעוני, טופו, סטייק טופו, מוקפץ טופו, קציצות סלק, חומוס בליווי טחון סויה, צ'ילי קונקרנה טבעוני. expected = total_days
   - "מנת בקר יומית" → ANY row containing a beef dish (usually [עמדת אוכל רחוב] / [עמדת שף] / [עמדה משתנה]). Beef = בקר/בריסקט/חזה בקר/אסאדו/בשר/קבב/קציצות בקר/בורקס בשר/פילו בשר/מפרום
   - "שניצל יומי" / "שניצל 5 פעמים בשבוע" → ONLY [ציפסר] row. EXCLUDE שניצלוני הבית/שניצלונים/שניצלון
   - "חזה עוף בגריל 5 פעמים בשבוע" / "גריל יומי" → ONLY [עמדת גריל 2] row containing חזה עוף. Every day will match — expected = total_days
   - "קרוטונים יומיים" / "שקדי מרק יומיים" / "אטריות מרק 2 פעמים בשבוע" → ONLY [תוספות למרק] row
   - "מרק צח/ירקות יומי" → ONLY [מרק היום] / [מרק*] rows. Count days where at least one soup is מרק צח or מרק ירקות
   - "מינימום 2 סוגי מרק ביום" → ONLY [מרק היום] / [מרק*] rows — count days with ≥2 distinct soup names
   - "מינימום 11 סוגי סלטים ביום" → ONLY [סלט מורכב 1] through [סלט מורכב 11] + [בר בריא*] rows — count days with ≥11 distinct salads. NEVER count salad words appearing inside protein dish names.
   - "סלט פירות / פירות חתוכים" → ONLY [סלט פירות*] / [סלט פירות או פרי*] row
   - "עמדת אוכל רחוב 5 פעמים בשבוע" → ONLY [עמדת אוכל רחוב] row — count days where it's non-empty
   - "כבד עוף עם בצל ופטריות על פירה" → ONLY [עמדת אוכל רחוב] / [עמדה משתנה] containing כבד עוף
   - "סלמון פעם בשבוע" / "פילה סלמון" → [בריאות מנת דג] / [עמדת שף] / [עמדה משתנה] containing סלמון
   - "פילה אמנון" → [בריאות מנת דג] containing אמנון or מושט (synonyms)
   - "פילה לברק" → [בריאות מנת דג] containing לברק
   - "חריימה של נסיכה" → [בריאות מנת דג] containing חריימה or נסיכה
   - "קציצות דגים ביתיות" → ONLY [בריאות מנת דג] containing קציצות דג. NEVER count "קציצות בקר" / "קציצות פרגית" / "קציצות סלק" toward fish dish rule.
   - "פיש & ציפס" → [בריאות מנת דג] containing פיש
   - Salad-frequency rules (סלט אבוקדו, סלט עגבניות איטלקי, סלט עגבניות חריף, סלט קיסר, סלט ארטישוק, סלט קינואה, טאבולה בורגול, סלט פטריות, סלט עשבי תיבול ופיצוחים, סלט וואקמה) → ONLY [סלט מורכב N] / [בר בריא*] rows. The salad must be a STANDALONE row item, NOT mentioned as a side served alongside a protein. Example: "פרגית צלויה בגריל לצד סלט עגבניות חריף" appearing in [עמדת גריל] row is a פרגית dish — does NOT count toward "סלט עגבניות חריף" requirement.
   - "אסאדו 3 פעמים בחודש" → count any [row] containing אסאדו, INCLUDING shredded versions: אסאדו מפורק / בשר אסאדו / רצועות אסאדו / סינייה אסאדו / פילו ממולא בשר אסאדו all count toward אסאדו (target 3-3.5/month)
   - "בריסקט 2 פעמים בחודש" → count any [row] containing בריסקט or חזה בקר (same cut)
   - "צלי כתף" → count any [row] containing "בשר צלי" with cut number מס' 5 or מס' 6
   - "בשר ראש" → count any [row] containing בשר ראש with cut number מס' 4 or מס' 10. NOTE: "בשר גולש" is a separate cut, NOT בשר ראש unless explicitly named
   - "המבורגר ביתי" → ONLY [עמדת אוכל רחוב] / [עמדת שף] containing המבורגר. Do NOT count "המבורגר טבעוני" (that's vegan).
   - "שווארמה פעם בשבוע" → count days with שווארמה הודו / שווארמה פרגית. Do NOT count שווארמה דג / שווארמה טבעוני (different protein) toward שווארמה הודו requirement.
   - "סטייק פרגית פעם בשבוע" → ONLY [עמדת אוכל רחוב] / [עמדת גריל*] containing סטייק פרגית / שיפודי פרגית / קציצות פרגית / מסאחן פרגית / מוקפץ פרגית / פרגית צלויה
   - "עוגת שמרים/עוגת קראנץ' 4 פעמים בחודש" → ONLY count items containing "עוגת שמרים" or "קראנץ'" (typically every Thursday). Do NOT count עוגת סולת / עוגת גזר / etc.
   - "עוגה מסוגים שונים" / variety check → count distinct cake names in [קינוח עוגה] row across the month — must be ≥ 5 distinct varieties
6c. SITE FILTERING — when checking site_name="קרית גת" (KG), SKIP these NZ-only rules entirely (do NOT include them in output): מינימום 4 סוגי ירקות אנטיפסטי, מסאחן פרגית, קציצות עוף ברוטב חמוסטה, מאפה בקר וחציל שרוף, סינייה אסאדו, פילו במילוי בקר מפורק, כנאפה אסאדו. The dish דפי פילו ממולא בשר מפורק IS the substitute for כנאפה אסאדו (count it for NZ only).
6d. CUT NUMBER ENFORCEMENT — for these rules, count ONLY items with explicit "מספר X" / "מס X" / "מס' X" cut number:
   - "עוף שלם" / "עופות שלמים" → must say מס' 2 / מספר 2 (not just "רבעי עוף")
   - "בריסקט" → must say מס' 3 or מס' 10. NOTE: חזה בקר = בריסקט (same cut, count both)
   - "צלי כתף" → must say מס' 5 or מס' 6 (both accepted)
   - "בשר ראש" → must say מס' 4 or מס' 10
6e. EQUIVALENCE — count interchangeably:
   - מושט = אמנון (same fish — count both toward "פילה אמנון")
   - חזה בקר = בריסקט (same cut)
   - דפי פילו ממולא בשר מפורק = כנאפה אסאדו substitute (NZ only)
7. ANOMALY DETECTION — also flag these issues as separate חריגים rows (group="חריגים", expected=0, actual=0):
   - Vague dish names like "מנת דג", "מנת בשר", "מנת עוף" — the supplier MUST specify the exact dish. Flag: "שם מנה לא מדויק - הספק צריך לפרט"
   - Same dish appearing on CONSECUTIVE days — ONLY flag if the dish name is clearly the SAME dish (identical or near-identical name, same protein/ingredient). E.g. "פילה לברק" on Apr 1 AND Apr 2 = flag. "פילה לברק" on Apr 1 and "מסאחן פרגית" on Apr 2 = NOT the same dish, do NOT flag. Each consecutive-day anomaly must list only the matched items for THAT specific dish. Create a SEPARATE anomaly row for each dish with consecutive repeats. Flag: "מנה חוזרת ימים רצופים"
   - Same dish appearing TWICE on the same day. Flag: "מנה כפולה באותו יום"
   - More than 1 ground-meat dish on the same day. בשר טחון includes: קציצות, המבורגר, קבב, בורקס בשר, פסטל בשר, סמבוסק בשר, פילו ממולא בשר, פילו בשר מפורק, מפרום בשר, קובה בשר, ראוויולי בשר. Flag: "יותר ממנת בשר טחון אחת ביום". Include all problematic items for that day in matched_items, set found_dates to that single date.
   - More than 1 פרגית dish on the same day (any of: פרגית צרפתית, פרגית במילוי, פרגית צלויה/בגריל, סטייק פרגית, מסאחן פרגית, מוקפץ פרגית, שיפודי פרגית, קציצות פרגית, שווארמה פרגית). Flag: "שתי מנות פרגית באותו יום — לא מאושר"
   - More than 1 dish from the SAME primary protein on the same day (general rule). Proteins to track: סלמון, אמנון, לברק, אסאדו, בריסקט, חזה בקר, חזה עוף, בשר ראש. Flag: "שתי מנות מאותו חומ\"ג באותו יום"
   - בקר/בשר dish WITHOUT a cut number. Beef cuts (בקר, בריסקט, חזה בקר, אסאדו, בשר צלי, בשר גולש, בשר ראש, לשון, בשר מפורק, שווארמה בקר, קציצות בקר, בורקס בשר, פילו בשר, סמבוסק בשר, פסטל בשר) MUST contain "מספר X" or "מס X" or "מס' X" (where X is a digit). Examples that PASS: "בשר צלי מספר 6", "חזה בקר מס 3", "בשר ראש מספר 10". Examples that FAIL and must be flagged: "בשר מפורק", "בורקס בשר", "קציצות בקר ברוטב עגבניות", "אסאדו בפריסה" (without מספר). Flag: "מנת בקר ללא ציון מספר נתח — חובה לציין מספר נתח". Create one anomaly row per offending dish, with that dish's date(s) in found_dates and the dish text in matched_items.
   - Any appearance of שניצלוני הבית / שניצלונים / שניצלון. Flag: "מנה לא מאושרת — שניצלונים אינם תחליף לשניצל". List all offending dates and items. CRITICAL: NEVER count שניצלוני/שניצלון items toward ANY שניצל rule (שניצל יומי, שניצל 5 פעמים בשבוע, שניצל בהכנה מקומית) — they are unapproved across all schnitzel requirements.
   - Ground meat anomaly הערות column MUST list the violation date(s) (e.g. "14.5: המבורגר + בורקס בשר"). Same for פרגית/protein/duplicate anomalies — always include the date in הערות so the user can locate the violation.
   - Same salad appearing in TWO different [סלט מורכב N] rows on the same day. Flag: "סלט כפול באותו יום". Include date.
   - Same exact item appearing on consecutive days (any row). Flag: "מנה חוזרת ימים רצופים". This applies to non-protein items too: a salad named identically on Mon and Tue, a soup named identically on Wed and Thu, etc. Include both dates and the item.
   - Same soup type (excluding מרק צח) appearing TWICE in one week. Flag: "אותו מרק פעמיים בשבוע". Include both dates and the soup name.
8. SUBSTITUTION & COUNTING RULES (from compliance report notes):
   • "שווארמה פרגית" is an APPROVED substitute for "שווארמה הודו ירך נקבה" (NZ) — count it as fulfilling that requirement and add note "הוגשה שווארמה פרגית כתחליף להודו"
   • "אנטריקוט" is an APPROVED substitute for "סטייק סינטה" — count it and note "הוגש אנטריקוט כתחליף לסינטה"
   • פרגית dishes requirement (NZ): these dishes ALL count toward the weekly פרגית requirement: קציצות פרגית, מסאחן פרגית, מוקפץ פרגית, שיפודי פרגית, סטייק פרגית. They are ALSO counted individually for their own rules.
   • "פרגית ממולאת" and "כרע עוף ממולא" do NOT count toward the פרגית requirement — count them separately under עוף
   • "שוברי שגרה" = a special THEME DAY (יום עיראקי, יום מרוקאי, etc.) — count days where menu features a special ethnic/regional cuisine theme
   • "ימים מיוחדים" = a special OCCASION DAY (יום העצמאות, פורים, חנוכה, etc.) — count days that coincide with Israeli national/religious holidays

CONTRACT REQUIREMENTS:
{rules_table}

ACTUAL MENU ({month_name} {year}, {site_name}, {total_days} working days):
{menu_text}

Return ONLY a JSON array with TWO sections:

SECTION 1 — Contract compliance items (one per required dish):
{{
  "group": "one of: מיוחדים/סלטים/עוף/בקר/מנות גריל/דגים/קינוחים",
  "dish": "dish name from contract in Hebrew",
  "frequency_text": "frequency in Hebrew (e.g. פעמיים בחודש)",
  "expected": number,
  "actual": number,
  "shortage": number,
  "found_dates": ["YYYY-MM-DD", ...],
  "matched_items": ["EXACT menu item text as it appeared in the menu — REQUIRED for every day where actual > 0", ...],
  "notes": "optional — e.g. הוגש בריסקט במקום סינטה"
}}

CRITICAL: matched_items MUST contain the exact text from the menu for every day counted in "actual". Never leave matched_items empty when actual > 0.

SECTION 2 — Anomaly items (if any found):
{{
  "group": "חריגים",
  "dish": "description of the anomaly",
  "frequency_text": "",
  "expected": 0,
  "actual": 0,
  "shortage": 0,
  "found_dates": ["dates where anomaly found"],
  "matched_items": ["the problematic menu items"],
  "notes": "explanation of the issue"
}}

Include ALL contract requirements with expected > 0.
Order by group, then by shortage descending (biggest shortages first within each group).
Anomalies (חריגים) go at the end.
"""

HEBREW_MONTHS = {
    1: "ינואר", 2: "פברואר", 3: "מרץ", 4: "אפריל",
    5: "מאי", 6: "יוני", 7: "יולי", 8: "אוגוסט",
    9: "ספטמבר", 10: "אוקטובר", 11: "נובמבר", 12: "דצמבר",
}


async def run_ai_compliance_check(
    check_id: int,
    db: AsyncSession,
) -> list[dict]:
    """Run AI-powered compliance check using Claude.

    Sends full menu + contract rules to Claude for intelligent matching.
    Returns results in the same format as the manual check spreadsheet.
    """
    # Load check
    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()
    if not check:
        raise ValueError(f"MenuCheck {check_id} not found")

    # Load site name
    from backend.models.site import Site
    site_result = await db.execute(select(Site).where(Site.id == check.site_id))
    site = site_result.scalar_one_or_none()
    site_name = site.name if site else "Unknown"

    # Load parsed menu days
    days_result = await db.execute(
        select(MenuDay)
        .where(MenuDay.menu_check_id == check_id)
        .order_by(MenuDay.date)
    )
    days = days_result.scalars().all()
    if not days:
        raise ValueError(f"No parsed menu days for check {check_id}")

    # Load active compliance rules for this site (site-specific + global)
    rules_result = await db.execute(
        select(ComplianceRule).where(
            ComplianceRule.is_active == True,
            or_(ComplianceRule.site_id == None, ComplianceRule.site_id == check.site_id),
        )
    )
    rules = rules_result.scalars().all()

    # ---- Translate DB category codes → Hebrew קבוצה BEFORE sending to Claude ----
    # The DB stores English category codes (daily_structure, daily_mandatory,
    # weekly_frequency, monthly_frequency, prohibition). The output report
    # MUST use Hebrew קבוצה values matching the V24/V25 format. We translate
    # here so Claude only ever sees Hebrew categories and echoes them back.
    HEB_GROUP_BY_KEYWORD = [
        # Order matters — first match wins. Keywords searched in rule name.
        ("דג", "דגים"),
        ("סלמון", "דגים"), ("אמנון", "דגים"), ("מושט", "דגים"),
        ("לברק", "דגים"), ("חריימה", "דגים"), ("פיש", "דגים"),
        ("שניצל", "עוף"), ("עוף", "עוף"), ("פרגית", "עוף"),
        ("הודו", "עוף"), ("חזה עוף", "עוף"), ("גריל יומי", "עוף"),
        ("בקר", "בקר"), ("בריסקט", "בקר"), ("אסאדו", "בקר"),
        ("צלי", "בקר"), ("בשר ראש", "בקר"), ("המבורגר", "בקר"),
        ("שווארמה", "מנות גריל"), ("סינטה", "מנות גריל"), ("אנטריקוט", "מנות גריל"),
        ("נקניקיות", "מנות גריל"),
        ("סלט", "סלטים"), ("טאבולה", "סלטים"), ("אבוקדו", "סלטים"),
        ("קיסר", "סלטים"), ("ארטישוק", "סלטים"), ("ירקות חמים", "סלטים"),
        ("פטריות", "סלטים"),
        ("עוגה", "קינוחים"), ("עוגת", "קינוחים"), ("פירות", "קינוחים"),
        ("קינוח", "קינוחים"),
    ]

    def _hebrew_group(rule_name: str, db_category: str) -> str:
        if db_category == "prohibition":
            return "חריגים"
        for kw, heb in HEB_GROUP_BY_KEYWORD:
            if kw in rule_name:
                return heb
        return "מיוחדים"  # default for daily_structure/daily_mandatory/etc.

    def _extract_freq_text(rule_name: str, db_freq_text: str) -> str:
        """Extract Hebrew frequency text from rule name when DB freq_text is empty."""
        if db_freq_text:
            return db_freq_text
        import re as _re
        # Common patterns in rule names
        m = _re.search(r"(\d+\s*פעמים?\s*ב(שבוע|חודש))", rule_name)
        if m:
            return m.group(1)
        m = _re.search(r"(פעם\s*ב(שבוע|חודש|רבעון))", rule_name)
        if m:
            return m.group(1)
        if "יומי" in rule_name or "יומית" in rule_name or "יומיים" in rule_name or "ביום" in rule_name:
            return "כל יום"
        if "מקסימום" in rule_name:
            return rule_name.split("מקסימום", 1)[1].strip() if "מקסימום" in rule_name else ""
        return ""

    # Site filter: NZ-only rules excluded from KG
    NZ_ONLY_KEYWORDS = [
        "אנטיפסטי", "מסאחן", "חמוסטה", "מאפה בקר",
        "סינייה אסאדו", "כנאפה אסאדו", "פילו במילוי בקר",
        "ויאטנמי", "ווקאמה", "סלק ותפוח",
    ]

    # Build compact rules table with TRANSLATED Hebrew categories
    rules_lines = []
    skipped_nz_only = 0
    for r in rules:
        params = r.parameters or {}
        freq_text = _extract_freq_text(r.name, params.get("frequency_text", ""))
        expected = params.get("count") or params.get("min_count") or params.get("expected", "")
        desc = r.description or ""

        # Skip NZ-only rules when checking KG
        if site_name and ("קרית גת" in site_name or site_name.upper() == "KG"):
            if any(nz_kw in r.name for nz_kw in NZ_ONLY_KEYWORDS):
                skipped_nz_only += 1
                continue

        heb_cat = _hebrew_group(r.name, r.category or "")
        line = f"- [{heb_cat}] {r.name}"
        if freq_text:
            line += f" | {freq_text}"
        if expected:
            line += f" | expected: {expected}"
        if desc and desc != r.name and desc != f"{params.get('frequency_text', '')} — {r.name}":
            line += f" | {desc}"
        rules_lines.append(line)
    rules_table = "\n".join(rules_lines)
    if skipped_nz_only:
        logger.info(f"AI compliance: skipped {skipped_nz_only} NZ-only rules for {site_name}")

    # Build daily menu text — PRESERVE row category labels (פחמימה מלאה,
    # ציפסר, בריאות מנת דג, סלט מורכב, etc.) so position-based rules can
    # work. Without this, Claude can't tell whether "קציצות בקר" is in the
    # protein row vs. a salad row.
    menu_lines = []
    for day in days:
        items = day.menu_items or {}
        date_str = day.date.isoformat() if day.date else ""
        dow = day.day_of_week or ""
        menu_lines.append(f"=== {date_str} ({dow}) ===")
        for category, item_list in items.items():
            cat_label = str(category).strip() if category else "אחר"
            if isinstance(item_list, list):
                clean = [str(it).strip() for it in item_list if str(it).strip() and len(str(it).strip()) >= 3]
                if clean:
                    menu_lines.append(f"  [{cat_label}] {' | '.join(clean)}")
            elif isinstance(item_list, str) and item_list.strip():
                menu_lines.append(f"  [{cat_label}] {item_list.strip()}")
    menu_text = "\n".join(menu_lines)

    month_name = HEBREW_MONTHS.get(check.month, str(check.month))

    # Build prompt
    prompt = AI_COMPLIANCE_CHECK_PROMPT.format(
        rules_table=rules_table,
        menu_text=menu_text,
        month_name=month_name,
        year=check.year,
        site_name=site_name,
        total_days=len(days),
    )

    logger.info(
        f"AI compliance check [v3-row-strict]: {len(rules)} rules "
        f"(skipped {skipped_nz_only} NZ-only for {site_name}), "
        f"{len(days)} days, prompt ~{len(prompt)} chars"
    )

    # Call Claude — use generate_response directly so our detailed schema
    # instructions in the prompt are not overridden by a second schema append
    raw_response = await claude_service.generate_response(
        prompt=prompt,
        system_prompt=(
            "You are a precise menu compliance auditor. "
            "Return ONLY a valid JSON array — no markdown, no code blocks, no extra text. "
            "ALWAYS populate matched_items with the exact menu text for every item you counted."
        ),
        max_tokens=16384,
    )

    # Extract the outermost JSON array from the response.
    # Claude sometimes wraps the JSON in markdown fences or adds preamble text —
    # this approach finds the array regardless of surrounding content.
    raw_response = raw_response.strip()
    json_text = raw_response

    # 1. Strip markdown code fences first
    if raw_response.startswith("```"):
        first_newline = raw_response.find("\n")
        if first_newline != -1:
            json_text = raw_response[first_newline + 1:]
        else:
            json_text = raw_response[3:]
        if json_text.rstrip().endswith("```"):
            json_text = json_text.rstrip()[:-3].rstrip()

    # 2. Find the outermost JSON array by bracket matching
    #    (handles any preamble/postamble text Claude might add)
    start = json_text.find("[")
    if start != -1:
        depth = 0
        end = -1
        for i, ch in enumerate(json_text[start:], start):
            if ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        if end != -1:
            json_text = json_text[start:end + 1]

    try:
        ai_results = json.loads(json_text)
    except json.JSONDecodeError as e:
        logger.error(
            f"AI compliance: JSON parse error: {e}\n"
            f"Raw response (first 800 chars): {raw_response[:800]}"
        )
        raise ValueError(
            f"Claude returned invalid JSON: {str(e)}. "
            f"Raw (first 200 chars): {raw_response[:200]}"
        )

    if not isinstance(ai_results, list):
        ai_results = ai_results.get("results", []) if isinstance(ai_results, dict) else []

    logger.info(f"AI compliance check returned {len(ai_results)} items")

    # ---- Guarantee Hebrew קבוצה + frequency_text on every result ----
    # Defensive: even if Claude returned an English code, force Hebrew
    # mapping using the same lookup used to build the rules table.
    HEB_GROUPS = {"מיוחדים", "סלטים", "עוף", "בקר", "מנות גריל", "דגים", "קינוחים", "חריגים"}
    for _r in ai_results:
        _grp = _r.get("group", "") or ""
        if _grp not in HEB_GROUPS:
            _r["group"] = _hebrew_group(_r.get("dish", ""), _grp)
        if not _r.get("frequency_text"):
            _r["frequency_text"] = _extract_freq_text(_r.get("dish", ""), "")

    # ---------------------------------------------------------------------------
    # Fallback: populate matched_items from MenuDay records when Claude omits them
    # For each result where actual > 0 but matched_items is empty, find menu items
    # from the found_dates that contain keywords from the dish name.
    # ---------------------------------------------------------------------------
    # Build date → flat items map once
    _date_items: dict[str, list[str]] = {}
    for _day in days:
        _key = _day.date.isoformat() if _day.date else ""
        _flat: list[str] = []
        for _cat_vals in (_day.menu_items or {}).values():
            if isinstance(_cat_vals, list):
                _flat.extend(str(x).strip() for x in _cat_vals if str(x).strip())
        _date_items[_key] = _flat

    # Stop-words and generic catering category words — excluded from keyword matching
    _SKIP = {"של", "עם", "על", "את", "בו", "לו", "כן", "לא", "יש"}
    _GENERIC = {
        "סלט", "סלטי", "מנת", "מנה", "מנות", "פילה", "קציצות",
        "ביתי", "ביתית", "ביתיות", "מקומי", "מקומית", "ברוטב",
        "מבושל", "מבושלת", "צלוי", "צלויה", "טרי", "טרייה",
        "ממולא", "ממולאת", "בגריל", "בתנור", "מטוגן", "מטוגנת",
    }

    for _item in ai_results:
        if _item.get("matched_items"):
            continue  # Claude already provided them
        if not (_item.get("actual", 0) > 0):
            continue  # Nothing was found, nothing to reconstruct

        dish = _item.get("dish", "")
        found_dates = _item.get("found_dates", [])

        # Specific keywords only: >= 4 chars, not generic/stop-word
        _specific = [w for w in dish.split() if len(w) >= 4 and w not in _SKIP and w not in _GENERIC]
        keywords = _specific or [w for w in dish.split() if len(w) >= 4 and w not in _SKIP]
        if not keywords:
            continue

        # If Claude didn't return found_dates, search all menu days
        dates_to_search = found_dates if found_dates else sorted(_date_items.keys())

        matched: list[str] = []
        seen: set[str] = set()
        for d_str in dates_to_search:
            for menu_item in _date_items.get(d_str, []):
                if menu_item in seen:
                    continue
                if any(kw in menu_item for kw in keywords):
                    matched.append(menu_item)
                    seen.add(menu_item)

        if matched:
            _item["matched_items"] = matched

    # Clear old check results and store new ones
    from sqlalchemy import delete as sql_delete
    await db.execute(
        sql_delete(CheckResult).where(CheckResult.menu_check_id == check_id)
    )

    critical_count = 0
    warning_count = 0
    passed_count = 0
    above_count = 0
    under_count = 0
    even_count = 0

    for item in ai_results:
        expected = item.get("expected", 0)
        actual = item.get("actual", 0)
        shortage = item.get("shortage", expected - actual)
        comparison = "under" if shortage > 0 else ("above" if shortage < 0 else "even")
        passed = shortage <= 0
        severity = "critical" if shortage > 0 else "info"

        result_obj = CheckResult(
            menu_check_id=check_id,
            rule_name=item.get("dish", ""),
            rule_category=item.get("group", ""),
            passed=passed,
            severity=severity,
            finding_text=f"{item.get('dish', '')}: Expected {expected}, Actual {actual}",
            evidence={
                "item_searched": item.get("dish", ""),
                "expected_count": expected,
                "actual_count": actual,
                "comparison": comparison,
                "found_on_days": item.get("found_dates", []),
                "matched_items": item.get("matched_items", []),
                "frequency_text": item.get("frequency_text", ""),
                "shortage": shortage,
                "notes": item.get("notes", ""),
                "ai_checked": True,
            },
            reviewed=False,
        )
        db.add(result_obj)

        if passed:
            passed_count += 1
        else:
            critical_count += 1

        if comparison == "above":
            above_count += 1
        elif comparison == "under":
            under_count += 1
        else:
            even_count += 1

    # Update check totals
    check.total_findings = critical_count + warning_count
    check.critical_findings = critical_count
    check.warnings = warning_count
    check.passed_rules = passed_count
    check.dishes_above = above_count
    check.dishes_under = under_count
    check.dishes_even = even_count

    await db.commit()

    return ai_results
