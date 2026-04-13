"""
Automated IMAP email poller for daily meal reports.

Connects to an IMAP mailbox (Gmail, Outlook, etc.), searches for
unread meal-report emails, extracts CSV/Excel attachments or inline tables,
and upserts them into the daily_meal_counts table.

Supports:
- Cibus Pluxee salary reports (noreply@notifications.pluxee.co.il)
- FoodHouse daily meal reports
- Forwarded emails (searches body for original sender)

Config (in .env):
    IMAP_HOST=imap.gmail.com
    IMAP_EMAIL=your@gmail.com
    IMAP_PASSWORD=abcd efgh ijkl mnop   # Gmail app password
    MEAL_EMAIL_SENDER=pluxee             # partial match on From or body
    MEAL_EMAIL_SUBJECT=Salary Report     # partial match on Subject
    MEAL_POLL_INTERVAL_MIN=60
"""
import asyncio
import imaplib
import email as email_lib
from email.header import decode_header
from datetime import date, datetime, timedelta
from typing import Optional
import csv
import io
import logging
import re
from html.parser import HTMLParser

from backend.config import get_settings
from backend.database import AsyncSessionLocal
from backend.api.webhooks import _upsert_daily_meals

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────
#  IMAP helpers (sync — run in executor for async usage)
# ──────────────────────────────────────────────────────

def _decode_header_value(raw: str) -> str:
    """Decode a possibly-encoded MIME header."""
    parts = decode_header(raw)
    decoded = []
    for data, charset in parts:
        if isinstance(data, bytes):
            decoded.append(data.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(data)
    return " ".join(decoded)


def _extract_csv_from_attachment(msg: email_lib.message.Message) -> Optional[str]:
    """Walk MIME parts looking for a .csv attachment."""
    for part in msg.walk():
        content_disp = str(part.get("Content-Disposition") or "")
        filename = part.get_filename()
        if filename:
            filename = _decode_header_value(filename)

        if "attachment" in content_disp and filename and filename.lower().endswith(".csv"):
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            # Try multiple Hebrew encodings
            for enc in ["cp1255", "windows-1255", "utf-8-sig", "utf-8", "iso-8859-8"]:
                try:
                    return payload.decode(enc)
                except (UnicodeDecodeError, LookupError):
                    continue
    return None


def _extract_excel_from_attachment(msg: email_lib.message.Message) -> Optional[list[dict]]:
    """Walk MIME parts looking for an Excel attachment (.xlsx/.xls).
    Returns parsed rows [{name, quantity}] or None.
    Also checks by content type for emails that don't set Content-Disposition."""
    all_rows: list[dict] = []

    for part in msg.walk():
        content_disp = str(part.get("Content-Disposition") or "")
        content_type = part.get_content_type() or ""
        filename = part.get_filename()
        if filename:
            filename = _decode_header_value(filename)

        # Skip CSV files — handled by CSV extractor
        if filename and filename.lower().endswith(".csv"):
            continue

        # Match by filename extension
        is_excel_by_name = (
            filename and (
                filename.lower().endswith(".xlsx")
                or filename.lower().endswith(".xls")
            )
        )

        # Match by MIME type (handles cases without Content-Disposition)
        excel_mime_types = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        )
        is_excel_by_type = content_type in excel_mime_types

        if is_excel_by_name or (is_excel_by_type and filename):
            payload = part.get_payload(decode=True)
            if not payload:
                continue

            try:
                parsed = _parse_excel_bytes(payload, filename or "unknown.xlsx")
                if parsed:
                    all_rows.extend(parsed)
            except Exception as e:
                logger.error(f"Error parsing Excel attachment {filename}: {e}")
                continue

    return all_rows if all_rows else None


def _parse_excel_bytes(data: bytes, filename: str = "") -> list[dict]:
    """Parse Excel bytes into [{name, quantity}] rows."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=True
        )
    except ImportError:
        logger.warning("openpyxl not installed — cannot parse Excel attachment")
        return []
    except Exception as e:
        logger.error(f"Failed to open Excel {filename}: {e}")
        return []

    rows: list[dict] = []
    sheet_names = list(wb.sheetnames)
    logger.info(f"Excel {filename}: sheets={sheet_names}, size={len(data)} bytes")
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue

        header_row = None
        name_col = None
        qty_col = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            str_cells = [str(c).strip().lower() if c else "" for c in row]

            # Detect header row
            if header_row is None:
                for i, cell in enumerate(str_cells):
                    if any(kw in cell for kw in [
                        "שם", "מסעדה", "name", "restaurant", "description",
                        "תיאור", "שם מסעדה", "שם עסק", "vendor", "supplier",
                    ]):
                        name_col = i
                    if any(kw in cell for kw in [
                        "כמות", "quantity", "count", "סכום", "total", "מספר",
                        "עסקאות", "transactions", "סה\"כ", "סהכ",
                    ]):
                        qty_col = i
                if name_col is not None and qty_col is not None:
                    header_row = row_idx
                    logger.info(f"Excel sheet '{sheet_name}': header at row {row_idx}, name_col={name_col}, qty_col={qty_col}, headers={str_cells[:5]}")
                    continue

                # Fallback: if first row has text + number pattern
                if row_idx == 0 and len(row) >= 2:
                    for i, c in enumerate(row):
                        if c and isinstance(c, str) and len(c) > 2:
                            name_col = i
                            break
                    for i, c in enumerate(row):
                        if isinstance(c, (int, float)) and c > 0:
                            qty_col = i
                            break
                    if name_col is not None and qty_col is not None:
                        # First row is data, not header
                        header_row = -1
                        try:
                            name_val = str(row[name_col]).strip()
                            qty_val = float(row[qty_col])
                            if name_val and qty_val >= 0:
                                rows.append({"name": name_val, "quantity": qty_val})
                        except (ValueError, TypeError, IndexError):
                            pass
                        continue

                continue

            # Data rows
            if name_col is not None and qty_col is not None:
                try:
                    raw_cells = list(row) if not isinstance(row, (list, tuple)) else row
                    name_val = str(raw_cells[name_col]).strip() if raw_cells[name_col] else ""
                    qty_raw = raw_cells[qty_col]

                    if not name_val or name_val.lower() == "none":
                        continue

                    qty_val = float(str(qty_raw).replace(",", "")) if qty_raw else 0
                    # Include 0-quantity rows (represent closed sites — show as zero)
                    if qty_val >= 0:
                        rows.append({"name": name_val, "quantity": qty_val})
                except (ValueError, TypeError, IndexError):
                    continue

    wb.close()
    # Log detailed parsing results
    if rows:
        logger.info(f"Excel parsed {len(rows)} rows from {filename}: {[r['name'][:40] + '=' + str(r['quantity']) for r in rows]}")
    else:
        logger.warning(f"Excel parsed 0 rows from {filename} (sheets: {sheet_names})")
    return rows


def _extract_csv_from_body(msg: email_lib.message.Message) -> Optional[str]:
    """Fallback: look for CSV-like data in the email body (plain text)."""
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype == "text/plain":
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            for enc in ["cp1255", "windows-1255", "utf-8-sig", "utf-8", "iso-8859-8"]:
                try:
                    text = payload.decode(enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            else:
                continue
            # Check if it looks like CSV with Hebrew restaurant names
            if "מסעדה" in text or "עסקאות" in text:
                return text
    return None


class _HTMLTableParser(HTMLParser):
    """Minimal HTML table parser that extracts rows as lists of cell text."""

    def __init__(self):
        super().__init__()
        self._tables: list[list[list[str]]] = []
        self._current_table: list[list[str]] = []
        self._current_row: list[str] = []
        self._current_cell: str = ""
        self._in_cell = False

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self._current_table = []
        elif tag == "tr":
            self._current_row = []
        elif tag in ("td", "th"):
            self._current_cell = ""
            self._in_cell = True

    def handle_endtag(self, tag):
        if tag in ("td", "th"):
            self._in_cell = False
            self._current_row.append(self._current_cell.strip())
        elif tag == "tr":
            if self._current_row:
                self._current_table.append(self._current_row)
        elif tag == "table":
            if self._current_table:
                self._tables.append(self._current_table)

    def handle_data(self, data):
        if self._in_cell:
            self._current_cell += data


def _extract_html_tables(msg: email_lib.message.Message) -> list[dict]:
    """Parse HTML email body for table data with meal counts."""
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype != "text/html":
            continue
        payload = part.get_payload(decode=True)
        if not payload:
            continue

        for enc in ["utf-8", "cp1255", "windows-1255", "utf-8-sig", "iso-8859-8"]:
            try:
                html = payload.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            continue

        parser = _HTMLTableParser()
        try:
            parser.feed(html)
        except Exception as e:
            logger.warning(f"HTML table parse error: {e}")
            continue

        rows: list[dict] = []
        for table in parser._tables:
            if len(table) < 2:
                continue

            # Auto-detect name/quantity columns from header row
            header = [c.lower().strip() for c in table[0]]
            name_col = None
            qty_col = None

            name_keywords = ["שם", "מסעדה", "תיאור", "name", "restaurant", "description"]
            qty_keywords = ["כמות", "quantity", "count", "סכום", "total", "מספר", "עסקאות"]

            for i, cell in enumerate(header):
                if any(kw in cell for kw in name_keywords):
                    name_col = i
                if any(kw in cell for kw in qty_keywords):
                    qty_col = i

            if name_col is None or qty_col is None:
                # Fallback: first text col + first numeric col
                for data_row in table[1:]:
                    if len(data_row) < 2:
                        continue
                    for i, c in enumerate(data_row):
                        if c and not c.replace(",", "").replace(".", "").replace("-", "").isdigit():
                            name_col = i
                            break
                    for i, c in enumerate(data_row):
                        try:
                            float(c.replace(",", ""))
                            qty_col = i
                            break
                        except (ValueError, AttributeError):
                            continue
                    break

            if name_col is None or qty_col is None:
                continue

            for data_row in table[1:]:
                if len(data_row) <= max(name_col, qty_col):
                    continue
                name_val = data_row[name_col].strip()
                if not name_val:
                    continue
                try:
                    qty_val = float(data_row[qty_col].replace(",", ""))
                    if qty_val >= 0:
                        rows.append({"name": name_val, "quantity": qty_val})
                except (ValueError, TypeError, IndexError):
                    continue

        if rows:
            return rows

    return []


def _parse_csv_text(text: str) -> list[dict]:
    """Parse CSV text into rows [{name, quantity}]."""
    rows: list[dict] = []
    reader = csv.reader(io.StringIO(text))
    header = next(reader, None)
    for csv_row in reader:
        if len(csv_row) >= 2 and csv_row[0].strip():
            try:
                qty = float(csv_row[1].strip().replace(",", ""))
                rows.append({"name": csv_row[0].strip(), "quantity": qty})
            except ValueError:
                continue
    return rows


def _extract_date_from_email(msg: email_lib.message.Message) -> date:
    """Try to extract the meal date from subject line first, then Date header.

    Subject examples:
        "FW: Salary Report 2026-03-02 3185" → 2026-03-02
        "Hadri_Ochel_2026-03-01"            → 2026-03-01
    """
    # 1. Try subject line (most reliable for meal date)
    subject = _decode_header_value(msg.get("Subject", ""))
    date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', subject)
    if date_match:
        try:
            return date(
                int(date_match.group(1)),
                int(date_match.group(2)),
                int(date_match.group(3)),
            )
        except (ValueError, TypeError):
            pass

    # 2. Try DD/MM/YYYY or DD.MM.YYYY in subject
    date_match = re.search(r'(\d{1,2})[./](\d{1,2})[./](\d{4})', subject)
    if date_match:
        try:
            return date(
                int(date_match.group(3)),
                int(date_match.group(2)),
                int(date_match.group(1)),
            )
        except (ValueError, TypeError):
            pass

    # 3. Fall back to email Date header
    date_str = msg.get("Date", "")
    if date_str:
        try:
            parsed = email_lib.utils.parsedate_to_datetime(date_str)
            return parsed.date()
        except Exception:
            pass

    return date.today()


def _fetch_meal_emails(
    imap_host: str,
    imap_email: str,
    imap_password: str,
    sender_filter: str = "",
    subject_filter: str = "",
    reprocess: bool = False,
    since_date: Optional[date] = None,
) -> list[dict]:
    """
    Connect to IMAP, find meal report emails, extract CSV/Excel data.
    Returns list of {date, rows, message_uid} dicts.

    Args:
        reprocess: If True, search ALL emails (not just UNSEEN) since since_date.
                   Does NOT mark emails as read when reprocessing.
        since_date: Only used when reprocess=True. Defaults to 30 days ago.
    """
    results: list[dict] = []

    try:
        mail = imaplib.IMAP4_SSL(imap_host)
        mail.login(imap_email, imap_password)
        mail.select("INBOX")

        # Build IMAP search criteria
        criteria_parts = []
        if not reprocess:
            criteria_parts.append("UNSEEN")
        else:
            # When reprocessing, search all emails since a date
            effective_since = since_date or (date.today() - timedelta(days=30))
            imap_date_str = effective_since.strftime("%d-%b-%Y")
            criteria_parts.append(f'SINCE {imap_date_str}')

        if subject_filter:
            criteria_parts.append(f'SUBJECT "{subject_filter}"')
        if sender_filter:
            criteria_parts.append(f'FROM "{sender_filter}"')

        search_criteria = "(" + " ".join(criteria_parts) + ")"
        logger.info(f"IMAP search: {search_criteria} (reprocess={reprocess})")
        status, data = mail.uid("search", None, search_criteria)
        logger.info(f"IMAP primary search: status={status}, count={len(data[0].split()) if data and data[0] else 0}")

        # Also run a diagnostic ALL search to compare
        try:
            diag_status, diag_data = mail.uid("search", None, f'(SUBJECT "{subject_filter}")')
            diag_count = len(diag_data[0].split()) if diag_data and diag_data[0] else 0
            logger.info(f"IMAP diag (SUBJECT only): {diag_count} emails total with subject '{subject_filter}'")
            diag_status2, diag_data2 = mail.uid("search", None, "(UNSEEN)")
            diag_count2 = len(diag_data2[0].split()) if diag_data2 and diag_data2[0] else 0
            logger.info(f"IMAP diag (UNSEEN only): {diag_count2} unread emails total")
        except Exception as de:
            logger.warning(f"Diagnostic search failed: {de}")

        # If no results with both filters, try subject-only (handles forwarded emails)
        if (status != "OK" or not data[0]) and sender_filter and subject_filter:
            logger.info("No results with sender+subject filter, trying subject-only")
            fallback_parts = []
            if not reprocess:
                fallback_parts.append("UNSEEN")
            elif since_date or reprocess:
                effective_since = since_date or (date.today() - timedelta(days=30))
                imap_date_str = effective_since.strftime("%d-%b-%Y")
                fallback_parts.append(f'SINCE {imap_date_str}')
            fallback_parts.append(f'SUBJECT "{subject_filter}"')
            fallback_criteria = "(" + " ".join(fallback_parts) + ")"
            status, data = mail.uid("search", None, fallback_criteria)
            logger.info(f"IMAP fallback search: status={status}, count={len(data[0].split()) if data and data[0] else 0}")

        if status != "OK" or not data[0]:
            logger.warning(f"No emails found. status={status}, data={data}")
            mail.logout()
            return results

        uids = data[0].split()
        logger.info(f"Found {len(uids)} meal email(s) (reprocess={reprocess})")

        for uid in uids:
            try:
                # Use BODY.PEEK[] to avoid auto-marking as \Seen
                status, msg_data = mail.uid("fetch", uid, "(BODY.PEEK[])")
                if status != "OK" or not msg_data[0]:
                    continue

                raw = msg_data[0][1]
                msg = email_lib.message_from_bytes(raw)

                subject = _decode_header_value(msg.get("Subject", ""))
                logger.info(f"Processing meal email: {subject}")

                # Log email parts for debugging
                parts_info = []
                for part in msg.walk():
                    ct = part.get_content_type()
                    fn = part.get_filename()
                    if fn:
                        fn = _decode_header_value(fn)
                    disp = str(part.get("Content-Disposition") or "")
                    size = len(part.get_payload(decode=True) or b"")
                    parts_info.append(f"{ct} fn={fn} disp={disp[:30]} size={size}")
                logger.info(f"Email parts: {parts_info}")

                # Try Excel attachment first, then CSV, then HTML tables, then body
                rows = None

                # 1. Try Excel attachment
                excel_rows = _extract_excel_from_attachment(msg)
                if excel_rows:
                    rows = excel_rows
                    logger.info(f"Parsed {len(rows)} rows from Excel attachment")

                # 2. Try CSV attachment
                if not rows:
                    csv_text = _extract_csv_from_attachment(msg)
                    if csv_text:
                        rows = _parse_csv_text(csv_text)
                        if rows:
                            logger.info(f"Parsed {len(rows)} rows from CSV attachment")

                # 3. Try HTML tables in email body
                if not rows:
                    html_rows = _extract_html_tables(msg)
                    if html_rows:
                        rows = html_rows
                        logger.info(f"Parsed {len(rows)} rows from HTML table")

                # 4. Try CSV-like data in body
                if not rows:
                    csv_text = _extract_csv_from_body(msg)
                    if csv_text:
                        rows = _parse_csv_text(csv_text)
                        if rows:
                            logger.info(f"Parsed {len(rows)} rows from body CSV")

                if not rows:
                    logger.warning(f"No parseable data found in email: {subject}")
                    # Don't mark as read — will retry on next poll
                    continue

                meal_date = _extract_date_from_email(msg)

                results.append({
                    "date": meal_date,
                    "rows": rows,
                    "uid": uid,
                    "subject": subject,
                })

                # Only mark as read after successful parse (skip when reprocessing)
                if not reprocess:
                    mail.uid("store", uid, "+FLAGS", "\\Seen")

            except Exception as e:
                logger.error(f"Error processing email uid={uid}: {e}")
                continue

        mail.logout()

    except imaplib.IMAP4.error as e:
        logger.error(f"IMAP connection error: {e}")
    except Exception as e:
        logger.error(f"Email poller error: {e}")

    return results


# ──────────────────────────────────────────────────────
#  Async poll function (called by scheduler)
# ──────────────────────────────────────────────────────

async def poll_meal_emails(
    reprocess: bool = False,
    since_date: Optional[date] = None,
) -> dict:
    """
    Check for new meal report emails and process them.

    Args:
        reprocess: If True, re-scan ALL emails (including already-read) since since_date.
        since_date: Start date for reprocessing (default: 30 days ago).
    """
    settings = get_settings()

    if not settings.IMAP_HOST or not settings.IMAP_EMAIL or not settings.IMAP_PASSWORD:
        return {"status": "skipped", "reason": "IMAP not configured"}

    # Run IMAP fetch directly (blocking but reliable — thread executor
    # was causing Gmail IMAP to return empty results)
    emails = _fetch_meal_emails(
        imap_host=settings.IMAP_HOST,
        imap_email=settings.IMAP_EMAIL,
        imap_password=settings.IMAP_PASSWORD,
        sender_filter=settings.MEAL_EMAIL_SENDER,
        subject_filter=settings.MEAL_EMAIL_SUBJECT,
        reprocess=reprocess,
        since_date=since_date,
    )

    if not emails:
        # Return diagnostic info when no emails found
        return {
            "status": "ok",
            "emails_found": 0,
            "debug": {
                "imap_host": settings.IMAP_HOST,
                "imap_email": settings.IMAP_EMAIL,
                "subject_filter": settings.MEAL_EMAIL_SUBJECT,
                "sender_filter": settings.MEAL_EMAIL_SENDER,
                "reprocess": reprocess,
                "since_date": since_date.isoformat() if since_date else None,
            },
        }

    total_created = 0
    total_updated = 0
    processed = []

    for em in emails:
        try:
            async with AsyncSessionLocal() as db:
                result = await _upsert_daily_meals(
                    db, em["date"], em["rows"], source="email_imap",
                )
                total_created += result["created"]
                total_updated += result["updated"]
                processed.append({
                    "subject": em["subject"],
                    "date": em["date"].isoformat(),
                    "rows": len(em["rows"]),
                    **result,
                })
        except Exception as e:
            logger.error(f"Error upserting meal data from email: {e}")
            processed.append({
                "subject": em.get("subject", "?"),
                "error": str(e),
            })

    logger.info(
        f"Meal email poll complete: {len(emails)} emails, "
        f"{total_created} created, {total_updated} updated"
    )

    return {
        "status": "ok",
        "emails_found": len(emails),
        "total_created": total_created,
        "total_updated": total_updated,
        "processed": processed,
    }


# ──────────────────────────────────────────────────────
#  Background scheduler
# ──────────────────────────────────────────────────────

async def start_meal_email_scheduler():
    """Background scheduler that polls for meal emails once daily at a configured hour."""
    from zoneinfo import ZoneInfo

    settings = get_settings()

    if not settings.IMAP_HOST:
        logger.info("IMAP not configured — meal email poller disabled")
        return

    target_hour = settings.MEAL_POLL_HOUR  # e.g. 5 = 5:00 AM Israel time
    tz = ZoneInfo("Asia/Jerusalem")

    logger.info(
        f"Meal email poller started: checking {settings.IMAP_EMAIL} "
        f"daily at {target_hour:02d}:00 Israel time"
    )

    # Initial delay to let the app finish startup
    await asyncio.sleep(30)

    # Run once on startup (catches any missed emails from downtime)
    try:
        result = await poll_meal_emails()
        if result.get("emails_found", 0) > 0:
            logger.info(f"Startup meal poll result: {result}")
    except Exception as e:
        logger.error(f"Startup meal poll error: {e}")

    while True:
        # Calculate seconds until next target hour
        now = datetime.now(tz)
        next_run = now.replace(hour=target_hour, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        wait_seconds = (next_run - now).total_seconds()

        logger.info(f"Next meal poll at {next_run.isoformat()} ({wait_seconds / 3600:.1f}h from now)")
        await asyncio.sleep(wait_seconds)

        try:
            result = await poll_meal_emails()
            if result.get("emails_found", 0) > 0:
                logger.info(f"Meal poll result: {result}")
        except Exception as e:
            logger.error(f"Meal email scheduler error: {e}")
