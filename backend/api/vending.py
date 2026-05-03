"""
Vending machines (מ.א אוטומטים) endpoints.

Two input sources per monthly invoice:
- PDF invoice  → creates Proforma + ProformaItems (8 categories @ unit prices)
- Excel detail → creates VendingTransaction rows (per-day / per-product / per-qty)

Dashboard reads totals from Proforma (budget vs actual).
/analytics page reads drill-down detail from VendingTransaction.
"""
from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import select, func, delete
from sqlalchemy.ext.asyncio import AsyncSession

from backend.api.auth import get_current_user
from backend.database import get_db
from backend.models.proforma import Proforma, ProformaItem
from backend.models.supplier import Supplier
from backend.models.user import User
from backend.models.vending_transaction import VendingTransaction
from backend.utils.db_compat import year_equals, month_equals

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/vending", tags=["vending"])

VENDING_SUPPLIER_NAME = "מ.א אוטומטים"


# ── Helpers ─────────────────────────────────────────────────────────────

def _reverse_hebrew_if_rtl(s: str) -> str:
    """pdfplumber often returns Hebrew reversed — reverse when detected."""
    if not s:
        return s
    # Heuristic: if the first Hebrew char is a final-form letter, text is reversed
    hebrew = [c for c in s if '\u0590' <= c <= '\u05FF']
    if hebrew and hebrew[0] in "ךםןףץ":
        return s[::-1]
    return s


async def _get_vending_supplier(db: AsyncSession) -> Supplier:
    result = await db.execute(select(Supplier).where(Supplier.name == VENDING_SUPPLIER_NAME))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(
            status_code=400,
            detail=f"Supplier '{VENDING_SUPPLIER_NAME}' not found — it should be auto-seeded at startup"
        )
    return supplier


def _parse_invoice_pdf(raw: bytes) -> dict:
    """Parse vending invoice PDF → {invoice_number, invoice_date, total_amount, items: [...]}.

    Returns categories with unit_price + quantity + total.
    """
    import pdfplumber

    result = {
        "invoice_number": None,
        "invoice_date": None,
        "total_amount": 0.0,
        "items": [],  # [{category, quantity, unit_price, total_price}]
    }

    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)

        # Invoice number: IN followed by digits
        m = re.search(r"IN\d{6,}", full_text)
        if m:
            result["invoice_number"] = m.group(0)

        # Invoice date: dd/mm/yy
        for line in full_text.split("\n"):
            if "תינובשח ךיראת" in line or "חשבונית תאריך" in line or "תאריך חשבונית" in line:
                m2 = re.search(r"(\d{2})/(\d{2})/(\d{2,4})", line)
                if m2:
                    dd, mm, yy = m2.groups()
                    year = int(yy) if len(yy) == 4 else 2000 + int(yy)
                    try:
                        result["invoice_date"] = date(year, int(mm), int(dd))
                        break
                    except ValueError:
                        pass
        if not result["invoice_date"]:
            m3 = re.search(r"(\d{2})/(\d{2})/(\d{2})", full_text)
            if m3:
                dd, mm, yy = m3.groups()
                try:
                    result["invoice_date"] = date(2000 + int(yy), int(mm), int(dd))
                except ValueError:
                    pass

        # Extract line items from tables.
        # Confirmed table layout (consistent for both 2025 and 2026 invoices):
        #   col 0: total          e.g. "5,144.88"
        #   col 1: unit_price     e.g. 'ח"ש 5.82'   (NIS prefix + price)
        #   col 2: quantity       e.g. "'חי 884.00" (יח' prefix + qty)
        #   col 3: product_desc   e.g. '1.25 תג תירק-םינטק םיכירכ' or 'הניטב'ג'
        #          (reversed Hebrew; may be prefixed with a "<MM.YY> " or numeric filler)
        #   col 4: SKU            e.g. "000" (placeholder, ignored)
        #   col 5: line_no
        # The previous version mistakenly read qty from col 3 — it was finding the
        # "1.25" filler instead of the real quantity, and using col 4 ("000") as the
        # product name. This block fixes column positions so all formats work.
        total_pre_vat = 0.0
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                for row in table:
                    if not row or len(row) < 4:
                        continue
                    cells = [str(c or "").strip() for c in row]
                    try:
                        total = float(cells[0].replace(",", ""))
                    except (ValueError, IndexError):
                        continue
                    unit_m = re.search(r"(-?[\d,]+\.?\d*)", cells[1] if len(cells) > 1 else "")
                    if not unit_m:
                        continue
                    try:
                        unit_price = float(unit_m.group(1).replace(",", ""))
                    except ValueError:
                        continue
                    qty_m = re.search(r"(-?[\d,]+\.?\d*)", cells[2] if len(cells) > 2 else "")
                    if not qty_m:
                        continue
                    try:
                        quantity = float(qty_m.group(1).replace(",", ""))
                    except ValueError:
                        continue
                    if quantity == 0:
                        continue
                    # Product description: strip any leading "<number> " filler
                    # (the 2025 invoices prepend "1.25 " or "MM.YY " before the product name)
                    desc_raw = cells[3] if len(cells) > 3 else ""
                    desc_raw = re.sub(r"^[\d.,]+\s+", "", desc_raw).strip()
                    desc = _reverse_hebrew_if_rtl(desc_raw).strip()
                    if not desc:
                        continue
                    result["items"].append({
                        "category": desc,
                        "quantity": quantity,
                        "unit_price": unit_price,
                        "total_price": total,
                    })
                    total_pre_vat += total

        # Use PRE-VAT total (after discount) — invoice text "החנה ירחא ריחמ" (reversed: "ריחמ ירחא החנה")
        # Example: "59,860.17 החנה ירחא ריחמ" → stored as 59,860.17 (not the post-VAT 70,635)
        m_pre_vat = re.search(r"([\d,]+\.\d+)\s+החנה\s+ירחא\s+ריחמ", full_text) or \
                    re.search(r"([\d,]+\.\d+)\s+ללוכ\s+ריחמ", full_text)
        if m_pre_vat:
            try:
                result["total_amount"] = float(m_pre_vat.group(1).replace(",", ""))
            except ValueError:
                result["total_amount"] = total_pre_vat
        else:
            result["total_amount"] = total_pre_vat

    return result


def _detect_xlsx_format(ws) -> str:
    """Probe the header row to decide which parser to use.

    Returns:
        '2025'  — 7-col format with SKU at col B and monthly aggregate at cols F+G
                  (header pattern: ת.התעודה|מק"ט|תאור מוצר|כמות|<blank>|שם המוצר|כמות חודשית שסופקה)
        '2026'  — 6-col format with per-day data at cols A+B+C
                  (header pattern: ת.התעודה|תאור מוצר|כמות|<blank>|שם המוצר|כמות חודשית שסופקה)
    """
    h_b = str(ws.cell(1, 2).value or "").strip()
    # 2025 has a SKU column at B labelled מק"ט. 2026's col B is the product description.
    if "מק" in h_b and "ט" in h_b:
        return "2025"
    return "2026"


def _parse_consumption_xlsx_2025(
    raw: bytes,
    invoice_month: Optional[date] = None,
) -> tuple[list[dict], dict]:
    """Parse 2025-format DataSheet XLSX → list of transactions.

    The 2025 file is a monthly aggregate: each unique product appears once with
    its total monthly quantity in col G. The per-machine columns (C+D) repeat
    the same products with low refill qtys, which we ignore.

    Header layout:
      col A: ת. התעודה (date)        — typically all = 1st of invoice month
      col B: מק"ט (SKU)
      col C: תאור מוצר (product description per-machine perspective)
      col D: כמות (per-machine refill qty)
      col E: (blank)
      col F: שם המוצר (canonical product name)         ← used for product
      col G: כמות חודשית שסופקה (monthly total qty)    ← used for quantity

    Returns (transactions, diagnostics) where diagnostics has counts of rows
    seen / kept / skipped + sample skipped reason.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Default tx_date = first day of invoice month (or first row's date, or today)
    base_date = invoice_month or date.today().replace(day=1)

    # Patterns we treat as summary/total rows, not real products
    TOTAL_ROW_TOKENS = ('סה"כ', 'סך הכל', 'סהכ', 'total', 'TOTAL')

    seen: dict[str, dict] = {}  # product_name → {qty}
    rows_total = 0
    rows_kept = 0
    rows_skipped = 0
    skipped_samples: list[str] = []
    for r in range(2, ws.max_row + 1):
        product = ws.cell(r, 6).value
        qty = ws.cell(r, 7).value
        if not product or qty is None:
            continue
        rows_total += 1
        name = str(product).strip()
        if any(tok in name for tok in TOTAL_ROW_TOKENS):
            rows_skipped += 1
            if len(skipped_samples) < 3:
                skipped_samples.append(f"row {r}: '{name}' looks like a total/summary row")
            continue
        if not isinstance(qty, (int, float)):
            rows_skipped += 1
            if len(skipped_samples) < 3:
                skipped_samples.append(f"row {r}: qty='{qty}' not numeric")
            continue
        if float(qty) <= 0:
            rows_skipped += 1
            if len(skipped_samples) < 3:
                skipped_samples.append(f"row {r}: qty={qty} <= 0")
            continue
        if name in seen:
            # Duplicate product within file — keep first occurrence (already aggregated)
            rows_skipped += 1
            continue
        seen[name] = {"quantity": float(qty)}
        rows_kept += 1

    transactions = [
        {
            "tx_date": base_date,
            "product_name": name,
            "quantity": data["quantity"],
            "machine_id": None,
        }
        for name, data in seen.items()
    ]
    diagnostics = {
        "format": "2025",
        "rows_total": rows_total,
        "rows_kept": rows_kept,
        "rows_skipped": rows_skipped,
        "skipped_samples": skipped_samples,
    }
    return transactions, diagnostics


def _parse_consumption_xlsx(
    raw: bytes,
    invoice_month: Optional[date] = None,
) -> list[dict]:
    """Parse DataSheet XLSX → list of transactions.

    Returns [{tx_date, product_name, quantity, machine_id (or None)}].
    Cols A=date, B=product, C=qty. Optional machine_id in future KG files.

    Date normalization:
        The source file often stores dates in ambiguous dd/mm vs mm/dd form.
        For days 1-12 the parser reads "02/01/26" as Jan 2 instead of Feb 1.
        If invoice_month is provided, each cell is normalized:
          - cell.month == invoice_month → kept as-is (day > 12 case, unambiguous)
          - cell.day == invoice_month   → SWAP day/month (ambiguous low-day case)
          - neither                     → kept as-is (out-of-month row, unusual)
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb[wb.sheetnames[0]]

    transactions = []
    # Detect optional machine_id column by looking at header row for "מכונה" / "machine"
    machine_col = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and isinstance(v, str) and ("מכונה" in v or "machine" in v.lower() or "מס' מכונה" in v):
            machine_col = c
            break

    target_month = invoice_month.month if invoice_month else None
    target_year = invoice_month.year if invoice_month else None

    def _parse_cell_date(val) -> Optional[date]:
        """Convert a date cell (datetime, date, or string) into a date. Returns None if unparseable."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            s = val.strip()
            if not s:
                return None
            # Try common formats — the Excel mixes datetime (days 1-12) with
            # strings like "2/15/2026" (mm/dd/yyyy US) for days 13-28
            for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%y", "%d/%m/%y", "%d.%m.%Y", "%d.%m.%y"):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
        return None

    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        tx_date = _parse_cell_date(raw)
        if tx_date is None:
            continue

        # Normalize ambiguous date if we have an invoice month
        if target_month is not None:
            if tx_date.month == target_month:
                pass  # already correct (day was > 12 so unambiguous, OR string date was already correct)
            elif tx_date.day == target_month and 1 <= tx_date.month <= 31:
                # dd/mm swap case: cell=(yr, day=tx.month, month=tx.day) should be (yr, tx.month, tx.day)
                try:
                    tx_date = date(target_year or tx_date.year, target_month, tx_date.month)
                except ValueError:
                    continue
            else:
                # Date doesn't match invoice month even after swap — skip to avoid spill-over
                continue

        product = ws.cell(r, 2).value
        qty = ws.cell(r, 3).value
        if not product or not isinstance(qty, (int, float)):
            continue
        machine = ws.cell(r, machine_col).value if machine_col else None
        transactions.append({
            "tx_date": tx_date,
            "product_name": str(product).strip(),
            "quantity": float(qty),
            "machine_id": str(machine).strip() if machine else None,
        })
    return transactions


def _classify_transaction(product_name: str, pdf_items: list[dict]) -> tuple[Optional[str], Optional[float]]:
    """Match a transaction product name to a PDF invoice category → returns (category, unit_price)."""
    if not pdf_items:
        return (None, None)
    name = product_name or ""
    for it in pdf_items:
        cat = it.get("category") or ""
        # Split category into keywords and check if any keyword is in the transaction
        kws = [w for w in re.split(r"\s+", cat) if len(w) >= 3]
        for kw in kws:
            if kw in name:
                return (cat, float(it.get("unit_price") or 0))
    return (None, None)


# ── Endpoints ───────────────────────────────────────────────────────────

@router.post("/upload")
async def upload_vending(
    site_id: int = Form(...),
    shift: str = Form("all"),  # 'all' | 'day' | 'evening'
    invoice_month: Optional[str] = Form(None),  # 'YYYY-MM' — used to normalize ambiguous Excel dates
    invoice_pdf: UploadFile = File(None),
    consumption_xlsx: UploadFile = File(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload vending invoice PDF + detailed consumption Excel together.

    - PDF creates/updates a Proforma with the 8 category items (prices)
    - Excel creates VendingTransaction rows (drill-down detail)
    - Transactions are linked to the Proforma and priced via category match
    """
    if not invoice_pdf and not consumption_xlsx:
        raise HTTPException(400, "Provide at least one file (PDF invoice or Excel consumption)")
    if shift not in ("all", "day", "evening"):
        raise HTTPException(400, "shift must be 'all', 'day', or 'evening'")

    # Parse invoice_month hint (YYYY-MM) used to normalize ambiguous Excel dates
    invoice_month_hint: Optional[date] = None
    if invoice_month:
        try:
            parts = invoice_month.split("-")
            invoice_month_hint = date(int(parts[0]), int(parts[1]), 1)
        except (ValueError, IndexError):
            raise HTTPException(400, "invoice_month must be in YYYY-MM format (e.g. 2026-02)")

    supplier = await _get_vending_supplier(db)

    proforma: Optional[Proforma] = None
    pdf_items: list[dict] = []
    invoice_date: Optional[date] = None

    # 1. Parse PDF → Proforma + ProformaItems
    if invoice_pdf and invoice_pdf.filename:
        raw_pdf = await invoice_pdf.read()
        try:
            parsed = _parse_invoice_pdf(raw_pdf)
        except Exception as e:
            logger.exception("Failed to parse vending invoice PDF")
            raise HTTPException(500, f"Failed to parse PDF: {e}")

        invoice_date = parsed.get("invoice_date") or date.today()
        invoice_number = parsed.get("invoice_number")
        pdf_items = parsed.get("items", [])
        total_amount = parsed.get("total_amount", 0.0) or sum(it["total_price"] for it in pdf_items)

        # Find existing proforma by (supplier, site, shift, invoice_date) or invoice_number
        exist_q = select(Proforma).where(
            Proforma.supplier_id == supplier.id,
            Proforma.site_id == site_id,
        )
        if invoice_number:
            exist_q = exist_q.where(Proforma.proforma_number == invoice_number)
        else:
            exist_q = exist_q.where(Proforma.invoice_date == invoice_date)
        existing = (await db.execute(exist_q)).scalar_one_or_none()

        if existing:
            proforma = existing
            proforma.total_amount = total_amount
            proforma.shift = shift
            proforma.invoice_date = invoice_date
            # Wipe old items
            await db.execute(delete(ProformaItem).where(ProformaItem.proforma_id == proforma.id))
        else:
            proforma = Proforma(
                supplier_id=supplier.id,
                site_id=site_id,
                shift=shift,
                invoice_date=invoice_date,
                total_amount=total_amount,
                currency="ILS",
                status="validated",
                proforma_number=invoice_number,
                notes=f"Vending invoice {shift} · uploaded by {current_user.email}",
            )
            db.add(proforma)
            await db.flush()

        for it in pdf_items:
            db.add(ProformaItem(
                proforma_id=proforma.id,
                product_name=it["category"],
                quantity=it["quantity"],
                unit=f"unit @ ₪{it['unit_price']}",
                unit_price=it["unit_price"],
                total_price=it["total_price"],
                flagged=False,
            ))
        await db.commit()
        await db.refresh(proforma)

    # 2. Parse Excel → VendingTransactions
    tx_count = 0
    tx_priced = 0
    excel_diagnostics: dict = {}
    if consumption_xlsx and consumption_xlsx.filename:
        raw_xlsx = await consumption_xlsx.read()
        # Fall back to the PDF invoice date as the month hint if user didn't supply one
        month_hint = invoice_month_hint or (invoice_date.replace(day=1) if invoice_date else None)
        # Detect format by probing the header row, then dispatch
        try:
            import openpyxl
            wb_probe = openpyxl.load_workbook(io.BytesIO(raw_xlsx), data_only=True, read_only=True)
            ws_probe = wb_probe[wb_probe.sheetnames[0]]
            xlsx_format = _detect_xlsx_format(ws_probe)
            wb_probe.close()
        except Exception as e:
            logger.exception("Failed to probe Excel format")
            raise HTTPException(500, f"Failed to read Excel: {e}")

        try:
            if xlsx_format == "2025":
                txs, excel_diagnostics = _parse_consumption_xlsx_2025(
                    raw_xlsx, invoice_month=month_hint
                )
            else:
                txs = _parse_consumption_xlsx(raw_xlsx, invoice_month=month_hint)
                excel_diagnostics = {"format": "2026", "rows_kept": len(txs)}
        except Exception as e:
            logger.exception("Failed to parse vending Excel")
            raise HTTPException(500, f"Failed to parse Excel: {e}")

        # Delete old transactions for affected months (site+shift scope)
        if txs:
            months = {(t["tx_date"].year, t["tx_date"].month) for t in txs}
            for y, m in months:
                await db.execute(
                    delete(VendingTransaction).where(
                        VendingTransaction.site_id == site_id,
                        VendingTransaction.shift == shift,
                        year_equals(VendingTransaction.tx_date, y),
                        month_equals(VendingTransaction.tx_date, m),
                    )
                )

        # Price lookup: if no PDF in this upload, pull ProformaItems from any
        # proforma for same site/month so transactions get priced from existing invoices
        # (e.g., KG Excel uploaded alone, using existing KG day + KG evening invoices).
        price_items = list(pdf_items)
        if not price_items and txs:
            months = {(t["tx_date"].year, t["tx_date"].month) for t in txs}
            for y, m in months:
                existing_items = (await db.execute(
                    select(ProformaItem.product_name, ProformaItem.unit_price)
                    .join(Proforma, Proforma.id == ProformaItem.proforma_id)
                    .where(
                        Proforma.supplier_id == supplier.id,
                        Proforma.site_id == site_id,
                        year_equals(Proforma.invoice_date, y),
                        month_equals(Proforma.invoice_date, m),
                    )
                )).all()
                for pname, uprice in existing_items:
                    price_items.append({"category": pname, "unit_price": float(uprice or 0), "quantity": 0, "total_price": 0})

        for t in txs:
            category, unit_price = _classify_transaction(t["product_name"], price_items)
            total_price = (unit_price * t["quantity"]) if unit_price else None
            if unit_price is not None:
                tx_priced += 1
            db.add(VendingTransaction(
                proforma_id=proforma.id if proforma else None,
                site_id=site_id,
                shift=shift,
                tx_date=t["tx_date"],
                product_name=t["product_name"],
                category=category,
                quantity=t["quantity"],
                unit_price=unit_price,
                total_price=total_price,
                machine_id=t.get("machine_id"),
            ))
            tx_count += 1
        await db.commit()

    return {
        "proforma_id": proforma.id if proforma else None,
        "invoice_date": invoice_date.isoformat() if invoice_date else None,
        "site_id": site_id,
        "shift": shift,
        "invoice_items": len(pdf_items),
        "transactions_saved": tx_count,
        "transactions_priced": tx_priced,
        "total_amount": proforma.total_amount if proforma else None,
        "excel_diagnostics": excel_diagnostics,
    }


@router.post("/sync-budgets")
async def sync_budgets(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Force-create the KG day/night budget rows and deactivate the legacy
    KG shift='all' row. Idempotent — safe to run anytime.

    KG day = ₪59,214/mo, KG night = ₪62,531/mo, NZ unchanged.
    """
    from backend.models.supplier_budget import SupplierBudget
    from sqlalchemy import update as sql_update

    supplier = await _get_vending_supplier(db)

    # Deactivate legacy KG shift='all' rows
    await db.execute(
        sql_update(SupplierBudget)
        .where(
            SupplierBudget.supplier_id == supplier.id,
            SupplierBudget.site_id == 2,
            SupplierBudget.shift == "all",
        )
        .values(is_active=False)
    )

    kg_amounts = {"day": 59214, "evening": 62531}
    created = []
    for yr in [2025, 2026]:
        for shift_kind, monthly in kg_amounts.items():
            existing = (await db.execute(
                select(SupplierBudget).where(
                    SupplierBudget.supplier_id == supplier.id,
                    SupplierBudget.site_id == 2,
                    SupplierBudget.year == yr,
                    SupplierBudget.shift == shift_kind,
                )
            )).scalar_one_or_none()
            if existing:
                # Make sure it's active and has the right amount
                existing.is_active = True
                if not existing.yearly_amount or existing.yearly_amount == 0:
                    existing.yearly_amount = monthly * 12
                    for m in ("jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"):
                        setattr(existing, m, monthly)
                continue
            db.add(SupplierBudget(
                supplier_id=supplier.id,
                site_id=2,
                year=yr,
                shift=shift_kind,
                yearly_amount=monthly * 12,
                jan=monthly, feb=monthly, mar=monthly,
                apr=monthly, may=monthly, jun=monthly,
                jul=monthly, aug=monthly, sep=monthly,
                oct=monthly, nov=monthly, dec=monthly,
                is_active=True,
            ))
            created.append({"year": yr, "shift": shift_kind, "yearly_amount": monthly * 12})
    await db.commit()
    return {"created": created, "kg_day_monthly": 59214, "kg_evening_monthly": 62531}


@router.post("/strip-vat")
async def strip_vat_from_existing(
    rate: float = 0.18,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """One-shot: divide vending Proforma.total_amount by (1+rate) for invoices
    uploaded before the pre-VAT change. Default rate 18% (Israeli VAT).

    Run once after deploy if you don't want to re-upload everything.
    """
    supplier = await _get_vending_supplier(db)
    proformas = (await db.execute(
        select(Proforma).where(Proforma.supplier_id == supplier.id)
    )).scalars().all()
    updated = 0
    for p in proformas:
        if p.total_amount and p.total_amount > 0:
            p.total_amount = round(p.total_amount / (1 + rate), 2)
            updated += 1
    await db.commit()
    return {"updated": updated, "rate": rate}


@router.post("/cleanup")
async def cleanup_bad_dates(
    site_id: int,
    year: int,
    month: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete all vending transactions for a given site whose tx_date falls
    OUTSIDE the stated (year, month). Used to recover from ambiguous-date
    imports where days 1-12 were mis-bucketed into other months.

    After cleanup, re-upload the Excel with the invoice month specified.
    """
    from calendar import monthrange
    last_day = monthrange(year, month)[1]
    start = date(year, month, 1)
    end = date(year, month, last_day)

    # Find tx for this site that are OUT of the target month
    to_delete_q = select(VendingTransaction).where(
        VendingTransaction.site_id == site_id,
        (VendingTransaction.tx_date < start) | (VendingTransaction.tx_date > end),
    )
    to_delete = (await db.execute(to_delete_q)).scalars().all()
    deleted = len(to_delete)
    for t in to_delete:
        await db.delete(t)
    await db.commit()
    return {"deleted": deleted, "site_id": site_id, "kept_month": f"{year}-{month:02d}"}


@router.delete("/clear")
async def clear_vending_data(
    site_id: Optional[int] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    shift: Optional[str] = None,
    include_invoices: bool = False,
    purge_all: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete vending transactions by filter.

    - site_id / year / month / shift: narrow scope (AND-combined)
    - include_invoices: also delete matching Proformas (+ ProformaItems)
    - purge_all: nuclear reset — delete every vending tx and every vending proforma
      (requires explicit purge_all=true, otherwise at least one filter is required)
    """
    if not purge_all and site_id is None and year is None and month is None and shift is None:
        raise HTTPException(400, "Refusing to clear without filters. Set purge_all=true for full reset.")

    supplier = await _get_vending_supplier(db)

    tx_q = select(VendingTransaction)
    if not purge_all:
        if site_id is not None:
            tx_q = tx_q.where(VendingTransaction.site_id == site_id)
        if year is not None:
            tx_q = tx_q.where(year_equals(VendingTransaction.tx_date, year))
        if month is not None:
            tx_q = tx_q.where(month_equals(VendingTransaction.tx_date, month))
        if shift is not None:
            tx_q = tx_q.where(VendingTransaction.shift == shift)

    txs = (await db.execute(tx_q)).scalars().all()
    tx_count = len(txs)
    for t in txs:
        await db.delete(t)

    inv_count = 0
    if include_invoices or purge_all:
        p_q = select(Proforma).where(Proforma.supplier_id == supplier.id)
        if not purge_all:
            if site_id is not None:
                p_q = p_q.where(Proforma.site_id == site_id)
            if year is not None:
                p_q = p_q.where(year_equals(Proforma.invoice_date, year))
            if month is not None:
                p_q = p_q.where(month_equals(Proforma.invoice_date, month))
            if shift is not None:
                p_q = p_q.where(Proforma.shift == shift)

        proformas = (await db.execute(p_q)).scalars().all()
        for p in proformas:
            await db.execute(delete(ProformaItem).where(ProformaItem.proforma_id == p.id))
            await db.delete(p)
            inv_count += 1

    await db.commit()
    return {
        "transactions_deleted": tx_count,
        "invoices_deleted": inv_count,
        "purge_all": purge_all,
        "filters": {"site_id": site_id, "year": year, "month": month, "shift": shift, "include_invoices": include_invoices},
    }


@router.post("/reprice")
async def reprice_transactions(
    year: Optional[int] = None,
    site_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Back-fill unit_price/total_price on VendingTransactions by matching them
    against ProformaItems from the same site+month. Use after uploading PDFs
    that came in after the Excel."""
    supplier = await _get_vending_supplier(db)
    target_year = year or datetime.now().year

    tx_q = select(VendingTransaction).where(year_equals(VendingTransaction.tx_date, target_year))
    if site_id:
        tx_q = tx_q.where(VendingTransaction.site_id == site_id)
    txs = (await db.execute(tx_q)).scalars().all()

    # Group by (site, year, month)
    by_month: dict[tuple, list[ProformaItem]] = {}
    updated = 0
    for t in txs:
        if not t.tx_date:
            continue
        key = (t.site_id, t.tx_date.year, t.tx_date.month)
        if key not in by_month:
            rows = (await db.execute(
                select(ProformaItem.product_name, ProformaItem.unit_price)
                .join(Proforma, Proforma.id == ProformaItem.proforma_id)
                .where(
                    Proforma.supplier_id == supplier.id,
                    Proforma.site_id == key[0],
                    year_equals(Proforma.invoice_date, key[1]),
                    month_equals(Proforma.invoice_date, key[2]),
                )
            )).all()
            by_month[key] = [{"category": r[0], "unit_price": float(r[1] or 0)} for r in rows]

        category, unit_price = _classify_transaction(t.product_name, by_month[key])
        if unit_price is not None:
            t.category = category
            t.unit_price = unit_price
            t.total_price = float(unit_price) * float(t.quantity or 0)
            updated += 1

    await db.commit()
    return {"updated": updated, "scanned": len(txs), "year": target_year, "site_id": site_id}


@router.get("/analytics")
async def vending_analytics(
    year: Optional[int] = None,
    month: Optional[int] = None,
    site_id: Optional[int] = None,
    shift: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Top products, monthly trend, per-machine usage for the Analytics tab."""
    target_year = year or datetime.now().year

    q = select(
        VendingTransaction.tx_date,
        VendingTransaction.product_name,
        VendingTransaction.category,
        VendingTransaction.quantity,
        VendingTransaction.total_price,
        VendingTransaction.machine_id,
        VendingTransaction.site_id,
        VendingTransaction.shift,
    ).where(year_equals(VendingTransaction.tx_date, target_year))
    if month:
        q = q.where(month_equals(VendingTransaction.tx_date, month))
    if site_id:
        q = q.where(VendingTransaction.site_id == site_id)
    if shift and shift != "all":
        q = q.where(VendingTransaction.shift == shift)

    rows = list((await db.execute(q)).all())

    # Totals
    total_qty = sum(float(r.quantity or 0) for r in rows)
    total_cost = sum(float(r.total_price or 0) for r in rows if r.total_price is not None)

    # Top products (by qty)
    by_product: dict[str, dict] = {}
    for r in rows:
        p = r.product_name or "?"
        if p not in by_product:
            by_product[p] = {"product_name": p, "category": r.category, "qty": 0.0, "cost": 0.0}
        by_product[p]["qty"] += float(r.quantity or 0)
        by_product[p]["cost"] += float(r.total_price or 0) if r.total_price else 0

    top_products = sorted(by_product.values(), key=lambda x: x["qty"], reverse=True)[:25]

    # Monthly trend — also broken down by site for stacked-by-site chart
    SITE_NAMES = {1: "Nes Ziona", 2: "Kiryat Gat"}
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly: dict[int, dict] = {m: {"month": m, "month_name": month_names[m - 1], "qty": 0, "cost": 0} for m in range(1, 13)}
    sites_seen: set = set()
    for r in rows:
        if not r.tx_date:
            continue
        m = r.tx_date.month
        monthly[m]["qty"] += float(r.quantity or 0)
        monthly[m]["cost"] += float(r.total_price or 0) if r.total_price else 0
        # Per-site breakdown for stacked chart
        site_label = SITE_NAMES.get(r.site_id, f"Site {r.site_id}")
        sites_seen.add(site_label)
        site_qty_key = f"qty_{site_label}"
        site_cost_key = f"cost_{site_label}"
        monthly[m][site_qty_key] = monthly[m].get(site_qty_key, 0) + float(r.quantity or 0)
        monthly[m][site_cost_key] = monthly[m].get(site_cost_key, 0) + (float(r.total_price or 0) if r.total_price else 0)

    # Per machine (only if machines present)
    by_machine: dict[str, dict] = {}
    for r in rows:
        if not r.machine_id:
            continue
        mid = r.machine_id
        if mid not in by_machine:
            by_machine[mid] = {"machine_id": mid, "qty": 0, "cost": 0}
        by_machine[mid]["qty"] += float(r.quantity or 0)
        by_machine[mid]["cost"] += float(r.total_price or 0) if r.total_price else 0
    top_machines = sorted(by_machine.values(), key=lambda x: x["qty"], reverse=True)

    return {
        "year": target_year,
        "month": month,
        "site_id": site_id,
        "shift": shift,
        "total_qty": round(total_qty),
        "total_cost": round(total_cost, 2),
        "top_products": [
            {"product_name": p["product_name"], "category": p["category"],
             "qty": round(p["qty"]), "cost": round(p["cost"], 2)}
            for p in top_products
        ],
        "monthly": [
            {**{"month": v["month"], "month_name": v["month_name"],
                "qty": round(v["qty"]), "cost": round(v["cost"], 2)},
             **{k: round(val) for k, val in v.items() if k.startswith("qty_") or k.startswith("cost_")}}
            for v in monthly.values()
        ],
        "sites_present": sorted(list(sites_seen)),
        "machines": [
            {"machine_id": m["machine_id"],
             "qty": round(m["qty"]), "cost": round(m["cost"], 2)}
            for m in top_machines
        ],
    }


@router.get("/product-trend")
async def product_trend(
    product_name: str,
    year: Optional[int] = None,
    site_id: Optional[int] = None,
    shift: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Daily/monthly trend for one product — used in the Analytics drill-down."""
    target_year = year or datetime.now().year
    q = select(
        VendingTransaction.tx_date,
        VendingTransaction.quantity,
        VendingTransaction.total_price,
    ).where(
        year_equals(VendingTransaction.tx_date, target_year),
        VendingTransaction.product_name == product_name,
    )
    if site_id:
        q = q.where(VendingTransaction.site_id == site_id)
    if shift and shift != "all":
        q = q.where(VendingTransaction.shift == shift)

    rows = list((await db.execute(q)).all())
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly: dict[int, dict] = {m: {"month": m, "month_name": month_names[m - 1], "qty": 0, "cost": 0} for m in range(1, 13)}
    for r in rows:
        if not r.tx_date:
            continue
        m = r.tx_date.month
        monthly[m]["qty"] += float(r.quantity or 0)
        monthly[m]["cost"] += float(r.total_price or 0) if r.total_price else 0

    return {
        "product_name": product_name,
        "year": target_year,
        "monthly": [
            {"month": v["month"], "month_name": v["month_name"],
             "qty": round(v["qty"]), "cost": round(v["cost"], 2)}
            for v in monthly.values()
        ],
    }
