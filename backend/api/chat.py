"""
Chat interface API — AI assistant with tool-use for full data access.

Claude can call data-query tools to look up specific information on demand,
then synthesize the results into a comprehensive answer.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from pydantic import BaseModel
from datetime import date, datetime, timedelta
from typing import Optional
import json
import logging
import io

from backend.database import get_db
from backend.models.user import User
from backend.models.proforma import Proforma, ProformaItem
from backend.models.supplier_budget import SupplierBudget
from backend.models.supplier import Supplier
from backend.models.site import Site
from backend.models.product import Product
from backend.models.price_list import PriceList, PriceListItem
from backend.models.violation import Violation
from backend.models.meeting import Meeting
from backend.models.todo import TodoItem
from backend.models.daily_meal_count import DailyMealCount
from backend.models.operations import Anomaly
from backend.models.project import Project, ProjectTask
from backend.api.auth import get_current_user
from backend.services.claude_service import claude_service
from backend.utils.db_compat import extract_month

logger = logging.getLogger(__name__)
router = APIRouter()

# ---------------------------------------------------------------------------
# System prompt
# ---------------------------------------------------------------------------
CHAT_SYSTEM_PROMPT = """You are an AI assistant for Catering Services at HP Israel.
You help Ziv manage catering operations across Nes Ziona (NZ) and Kiryat Gat (KG) sites.

You have tools to query live data from the system database. Use them to answer questions
with exact numbers. Always query the data — never guess or say you don't have access.

You can help with:
- Budget analysis: spending vs budget per supplier, per site, per month
- Product-level detail: what was ordered, quantities, prices, by supplier/site/month
- Meal tracking: daily meal counts by type and site
- Violation management: open violations, severity, patterns, fines
- Meeting preparation and upcoming schedule
- Project tracking: tasks, status, deadlines
- Anomaly detection and operational alerts
- Price list comparison across suppliers
- Forecasting and trend analysis based on historical data

When presenting data:
- Use ₪ for currency amounts
- Format numbers with commas (1,234)
- Use tables when comparing multiple items
- Be concise and actionable
- Respond in the same language as the user (Hebrew or English)

When the user asks for a chart, include a JSON block tagged with ```chart that the frontend
can render. Use this format:
```chart
{"type": "bar|line|pie", "title": "...", "data": [{"name": "...", "value": 123}], "xKey": "name", "yKey": "value"}
```"""

# ---------------------------------------------------------------------------
# Tool definitions for Claude
# ---------------------------------------------------------------------------
CHAT_TOOLS = [
    {
        "name": "query_spending",
        "description": "Query actual spending from proformas (invoices). Can filter by supplier, site, date range, and get per-product detail. Use this for questions about how much was spent, what was ordered, quantities, etc.",
        "input_schema": {
            "type": "object",
            "properties": {
                "supplier_name": {"type": "string", "description": "Filter by supplier name (partial match, case insensitive). E.g. 'FoodHouse', 'שף'"},
                "site_name": {"type": "string", "description": "Filter by site: 'Nes Ziona' or 'NZ' or 'נס ציונה' for Nes Ziona, 'Kiryat Gat' or 'KG' or 'קרית גת' for Kiryat Gat"},
                "month": {"type": "integer", "description": "Month number (1-12)"},
                "year": {"type": "integer", "description": "Year (default: current year)"},
                "include_items": {"type": "boolean", "description": "If true, include product-level line items (quantities, prices). Default false for summary, true for product questions."},
                "product_search": {"type": "string", "description": "Search for specific products by name (partial match). E.g. 'פירות', 'fruit', 'עוף'"},
            },
            "required": [],
        },
    },
    {
        "name": "query_budgets",
        "description": "Query budget allocations for suppliers by year. Shows planned budget vs actual spending.",
        "input_schema": {
            "type": "object",
            "properties": {
                "year": {"type": "integer", "description": "Budget year (default: current year)"},
                "supplier_name": {"type": "string", "description": "Filter by supplier name (partial match)"},
            },
            "required": [],
        },
    },
    {
        "name": "query_meals",
        "description": "Query daily meal counts. Shows meals served by type (Meat/Dairy/Main Only) and site.",
        "input_schema": {
            "type": "object",
            "properties": {
                "site_name": {"type": "string", "description": "Filter by site name"},
                "days": {"type": "integer", "description": "Number of days to look back (default: 30)"},
                "month": {"type": "integer", "description": "Specific month (1-12)"},
                "year": {"type": "integer", "description": "Year"},
            },
            "required": [],
        },
    },
    {
        "name": "query_violations",
        "description": "Query violations/complaints with status, severity, fines, and patterns.",
        "input_schema": {
            "type": "object",
            "properties": {
                "status": {"type": "string", "description": "Filter by status: open, investigating, resolved"},
                "site_name": {"type": "string", "description": "Filter by site"},
                "limit": {"type": "integer", "description": "Max results (default: 10)"},
            },
            "required": [],
        },
    },
    {
        "name": "query_meetings",
        "description": "Query upcoming or past meetings.",
        "input_schema": {
            "type": "object",
            "properties": {
                "upcoming": {"type": "boolean", "description": "True for upcoming, false for past (default: true)"},
                "limit": {"type": "integer", "description": "Max results (default: 5)"},
            },
            "required": [],
        },
    },
    {
        "name": "query_price_lists",
        "description": "Query price lists and compare product prices across suppliers.",
        "input_schema": {
            "type": "object",
            "properties": {
                "supplier_name": {"type": "string", "description": "Filter by supplier"},
                "product_search": {"type": "string", "description": "Search products by name"},
            },
            "required": [],
        },
    },
    {
        "name": "query_projects",
        "description": "Query projects and their tasks with status and deadlines.",
        "input_schema": {
            "type": "object",
            "properties": {
                "project_name": {"type": "string", "description": "Filter by project name (partial match)"},
                "include_tasks": {"type": "boolean", "description": "Include task details (default: true)"},
            },
            "required": [],
        },
    },
    {
        "name": "query_summary",
        "description": "Get a high-level operational summary: total suppliers, open violations, upcoming meetings, budget status, meal averages. Use for general status questions.",
        "input_schema": {
            "type": "object",
            "properties": {},
            "required": [],
        },
    },
]


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------
class ChatMessage(BaseModel):
    message: str


class ChatResponse(BaseModel):
    response: str
    suggestions: list[str] = []


# ---------------------------------------------------------------------------
# Helper: resolve site name to ID
# ---------------------------------------------------------------------------
def _match_site(name: str, sites: dict) -> Optional[int]:
    """Resolve a site name/alias to site ID."""
    if not name:
        return None
    name_lower = name.lower().strip()
    aliases = {
        "nz": "nes ziona", "nes ziona": "nes ziona", "נס ציונה": "nes ziona",
        "kg": "kiryat gat", "kiryat gat": "kiryat gat", "קרית גת": "kiryat gat",
    }
    normalized = aliases.get(name_lower, name_lower)
    for sid, sname in sites.items():
        if normalized in sname.lower() or sname.lower() in normalized:
            return sid
    return None


# ---------------------------------------------------------------------------
# Tool execution handlers
# ---------------------------------------------------------------------------
async def _exec_query_spending(db: AsyncSession, params: dict, suppliers: dict, sites: dict) -> str:
    """Execute spending query with optional product-level detail."""
    year = params.get("year", date.today().year)
    month = params.get("month")
    include_items = params.get("include_items", False)
    product_search = params.get("product_search")
    supplier_name = params.get("supplier_name")
    site_name = params.get("site_name")

    # If product search requested, force include_items
    if product_search:
        include_items = True

    # Build filters
    filters = [Proforma.invoice_date.isnot(None)]

    if month:
        start = date(year, month, 1)
        end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        filters.append(Proforma.invoice_date >= start)
        filters.append(Proforma.invoice_date < end)
    else:
        filters.append(Proforma.invoice_date >= date(year, 1, 1))
        filters.append(Proforma.invoice_date < date(year + 1, 1, 1))

    if supplier_name:
        matching_ids = [sid for sid, sn in suppliers.items() if supplier_name.lower() in sn.lower()]
        if matching_ids:
            filters.append(Proforma.supplier_id.in_(matching_ids))
        else:
            return f"No supplier found matching '{supplier_name}'"

    site_id = _match_site(site_name, sites) if site_name else None
    if site_name and not site_id:
        return f"No site found matching '{site_name}'"
    if site_id:
        filters.append(Proforma.site_id == site_id)

    # Summary query
    result = await db.execute(
        select(
            Proforma.supplier_id,
            Proforma.site_id,
            func.count(Proforma.id).label("count"),
            func.sum(Proforma.total_amount).label("total"),
        )
        .where(and_(*filters))
        .group_by(Proforma.supplier_id, Proforma.site_id)
    )
    rows = result.all()

    if not rows:
        period = f"{year}/{month}" if month else str(year)
        return f"No spending data found for the specified filters (period: {period})"

    lines = []
    grand_total = 0
    for row in rows:
        s_name = suppliers.get(row.supplier_id, f"Supplier #{row.supplier_id}")
        st_name = sites.get(row.site_id, f"Site #{row.site_id}")
        total = float(row.total or 0)
        grand_total += total
        lines.append(f"  {s_name} @ {st_name}: ₪{total:,.0f} ({row.count} invoices)")

    period_label = f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][month-1]} {year}" if month else str(year)
    header = f"Spending summary ({period_label}):\n  TOTAL: ₪{grand_total:,.0f}\n"
    result_text = header + "\n".join(lines)

    # Product-level detail
    if include_items:
        # Get proforma IDs matching filters
        proforma_result = await db.execute(
            select(Proforma.id, Proforma.supplier_id, Proforma.site_id)
            .where(and_(*filters))
        )
        proformas = proforma_result.all()
        proforma_ids = [p.id for p in proformas]
        proforma_map = {p.id: (suppliers.get(p.supplier_id, "?"), sites.get(p.site_id, "?")) for p in proformas}

        if proforma_ids:
            item_filters = [ProformaItem.proforma_id.in_(proforma_ids)]
            if product_search:
                item_filters.append(ProformaItem.product_name.ilike(f"%{product_search}%"))

            item_result = await db.execute(
                select(
                    ProformaItem.product_name,
                    ProformaItem.proforma_id,
                    ProformaItem.quantity,
                    ProformaItem.unit_price,
                    ProformaItem.total_price,
                    ProformaItem.unit,
                )
                .where(and_(*item_filters))
                .order_by(ProformaItem.total_price.desc())
                .limit(50)
            )
            items = item_result.all()

            if items:
                result_text += f"\n\nProduct details ({len(items)} items):"
                for item in items:
                    supplier_site = proforma_map.get(item.proforma_id, ("?", "?"))
                    qty_str = f"{item.quantity:,.1f}" if item.quantity else "?"
                    unit_str = item.unit or ""
                    price_str = f"₪{item.unit_price:,.2f}" if item.unit_price else "?"
                    total_str = f"₪{item.total_price:,.2f}" if item.total_price else "?"
                    result_text += (
                        f"\n  {item.product_name}: {qty_str} {unit_str} × {price_str} = {total_str}"
                        f" [{supplier_site[0]} @ {supplier_site[1]}]"
                    )
            elif product_search:
                result_text += f"\n\nNo products found matching '{product_search}'"

    return result_text


async def _exec_query_budgets(db: AsyncSession, params: dict, suppliers: dict) -> str:
    """Query budget allocations."""
    year = params.get("year", date.today().year)
    supplier_name = params.get("supplier_name")

    filters = [SupplierBudget.year == year, SupplierBudget.is_active == True]
    if supplier_name:
        matching_ids = [sid for sid, sn in suppliers.items() if supplier_name.lower() in sn.lower()]
        if matching_ids:
            filters.append(SupplierBudget.supplier_id.in_(matching_ids))

    result = await db.execute(select(SupplierBudget).where(and_(*filters)))
    budgets = result.scalars().all()

    if not budgets:
        return f"No budgets found for {year}"

    month_cols = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    lines = [f"Budget allocations ({year}):"]
    total_yearly = 0
    for b in budgets:
        s_name = suppliers.get(b.supplier_id, f"Supplier #{b.supplier_id}")
        yearly = b.yearly_amount or 0
        total_yearly += yearly
        monthly_parts = []
        for i, col in enumerate(month_cols):
            val = getattr(b, col) or 0
            if val > 0:
                monthly_parts.append(f"{month_names[i]}=₪{val:,.0f}")
        lines.append(f"  {s_name}: yearly=₪{yearly:,.0f} | {', '.join(monthly_parts)}")

    lines.insert(1, f"  TOTAL: ₪{total_yearly:,.0f}")
    return "\n".join(lines)


async def _exec_query_meals(db: AsyncSession, params: dict, sites: dict) -> str:
    """Query meal counts."""
    site_name = params.get("site_name")
    days = params.get("days", 30)
    month = params.get("month")
    year = params.get("year", date.today().year)

    filters = []
    if month:
        start = date(year, month, 1)
        end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        filters.append(DailyMealCount.date >= start)
        filters.append(DailyMealCount.date < end)
    else:
        cutoff = date.today() - timedelta(days=days)
        filters.append(DailyMealCount.date >= cutoff)

    site_id = _match_site(site_name, sites) if site_name else None
    if site_id:
        filters.append(DailyMealCount.site_id == site_id)

    result = await db.execute(
        select(
            DailyMealCount.site_id,
            DailyMealCount.meal_type_en,
            func.sum(DailyMealCount.quantity).label("total"),
            func.count(DailyMealCount.id).label("days"),
            func.avg(DailyMealCount.quantity).label("avg"),
        )
        .where(and_(*filters))
        .group_by(DailyMealCount.site_id, DailyMealCount.meal_type_en)
    )
    rows = result.all()

    if not rows:
        return "No meal data found for the specified filters"

    lines = ["Meal counts:"]
    for row in rows:
        st_name = sites.get(row.site_id, f"Site #{row.site_id}")
        meal_type = row.meal_type_en or "Unknown"
        total = float(row.total or 0)
        avg_val = float(row.avg or 0)
        lines.append(f"  {st_name} — {meal_type}: {total:,.0f} total ({row.days} days, avg {avg_val:,.0f}/day)")

    return "\n".join(lines)


async def _exec_query_violations(db: AsyncSession, params: dict, sites: dict) -> str:
    """Query violations."""
    status = params.get("status")
    site_name = params.get("site_name")
    limit = params.get("limit", 10)

    filters = []
    if status:
        filters.append(Violation.status == status)
    site_id = _match_site(site_name, sites) if site_name else None
    if site_id:
        filters.append(Violation.site_id == site_id)

    result = await db.execute(
        select(Violation)
        .where(and_(*filters)) if filters else select(Violation)
        .order_by(Violation.received_at.desc())
        .limit(limit)
    )
    violations = result.scalars().all()

    if not violations:
        return "No violations found"

    lines = [f"Violations ({len(violations)} results):"]
    for v in violations:
        st_name = sites.get(v.site_id, "?")
        fine = f", fine=₪{v.fine_amount:,.0f}" if v.fine_amount else ""
        text = (v.violation_text or "")[:100]
        lines.append(
            f"  [{v.severity or '?'}] {v.category or '?'} @ {st_name} — {v.status}{fine}"
            f"\n    {text}"
        )

    return "\n".join(lines)


async def _exec_query_meetings(db: AsyncSession, params: dict, sites: dict) -> str:
    """Query meetings."""
    upcoming = params.get("upcoming", True)
    limit = params.get("limit", 5)
    now = datetime.now()

    query = select(Meeting)
    if upcoming:
        query = query.where(Meeting.scheduled_at >= now).order_by(Meeting.scheduled_at.asc())
    else:
        query = query.where(Meeting.scheduled_at < now).order_by(Meeting.scheduled_at.desc())
    query = query.limit(limit)

    result = await db.execute(query)
    meetings = result.scalars().all()

    if not meetings:
        return "No meetings found"

    label = "Upcoming" if upcoming else "Past"
    lines = [f"{label} meetings ({len(meetings)}):"]
    for m in meetings:
        st_name = sites.get(m.site_id, "")
        site_str = f" @ {st_name}" if st_name else ""
        lines.append(
            f"  {m.scheduled_at.strftime('%Y-%m-%d %H:%M')} | {m.title or 'Untitled'}"
            f" ({m.meeting_type or 'general'}{site_str}, {m.duration_minutes or 60}min)"
        )

    return "\n".join(lines)


async def _exec_query_price_lists(db: AsyncSession, params: dict, suppliers: dict) -> str:
    """Query price lists."""
    supplier_name = params.get("supplier_name")
    product_search = params.get("product_search")

    query = select(PriceList)
    if supplier_name:
        matching_ids = [sid for sid, sn in suppliers.items() if supplier_name.lower() in sn.lower()]
        if matching_ids:
            query = query.where(PriceList.supplier_id.in_(matching_ids))

    result = await db.execute(query.order_by(PriceList.effective_date.desc()).limit(10))
    price_lists = result.scalars().all()

    if not price_lists:
        return "No price lists found"

    lines = [f"Price lists ({len(price_lists)}):"]
    for pl in price_lists:
        s_name = suppliers.get(pl.supplier_id, "?")
        lines.append(f"\n  {pl.name} ({s_name}) — effective: {pl.effective_date}")

        # Get items
        item_query = select(PriceListItem).where(PriceListItem.price_list_id == pl.id)
        if product_search:
            item_query = item_query.where(PriceListItem.product_name.ilike(f"%{product_search}%"))
        item_query = item_query.order_by(PriceListItem.product_name).limit(30)

        item_result = await db.execute(item_query)
        items = item_result.scalars().all()
        for item in items:
            unit = item.unit or ""
            lines.append(f"    {item.product_name}: ₪{item.price:,.2f}/{unit}" if item.price else f"    {item.product_name}: no price")

    return "\n".join(lines)


async def _exec_query_projects(db: AsyncSession, params: dict) -> str:
    """Query projects and tasks."""
    project_name = params.get("project_name")
    include_tasks = params.get("include_tasks", True)

    from sqlalchemy.orm import selectinload
    query = select(Project).options(selectinload(Project.tasks), selectinload(Project.site))
    if project_name:
        query = query.where(Project.name.ilike(f"%{project_name}%"))

    result = await db.execute(query)
    projects = result.scalars().all()

    if not projects:
        return "No projects found"

    lines = [f"Projects ({len(projects)}):"]
    for p in projects:
        site_name = p.site.name if p.site else "?"
        tasks = list(p.tasks) if p.tasks else []
        done = sum(1 for t in tasks if t.status == "done")
        lines.append(
            f"\n  {p.name} — {p.status} ({p.priority}) @ {site_name}"
            f"\n    Progress: {done}/{len(tasks)} tasks"
            f" | Target: {p.target_end_date or '?'}"
        )
        if include_tasks and tasks:
            for t in tasks:
                overdue_mark = " ⚠OVERDUE" if t.due_date and t.status != "done" and t.due_date < date.today() else ""
                lines.append(
                    f"    - [{t.status}] {t.title}"
                    f" (due: {t.due_date or '?'}, assigned: {t.assigned_to or '?'}){overdue_mark}"
                )

    return "\n".join(lines)


async def _exec_query_summary(db: AsyncSession, sites: dict, suppliers: dict) -> str:
    """Get high-level operational summary."""
    now = datetime.now()
    today = date.today()
    year = today.year

    lines = [f"Operational Summary ({today.strftime('%Y-%m-%d')}):"]

    # Suppliers
    lines.append(f"  Active suppliers: {len(suppliers)}")

    # Proformas this year
    result = await db.execute(
        select(func.count(Proforma.id), func.sum(Proforma.total_amount))
        .where(Proforma.invoice_date >= date(year, 1, 1))
    )
    row = result.one()
    lines.append(f"  Proformas this year: {row[0] or 0}, total: ₪{float(row[1] or 0):,.0f}")

    # Budget
    budget_result = await db.execute(
        select(func.sum(SupplierBudget.yearly_amount))
        .where(SupplierBudget.year == year, SupplierBudget.is_active == True)
    )
    total_budget = float(budget_result.scalar() or 0)
    lines.append(f"  Total budget {year}: ₪{total_budget:,.0f}")

    # Open violations
    viol_result = await db.execute(
        select(func.count(Violation.id)).where(Violation.status != "resolved")
    )
    lines.append(f"  Open violations: {viol_result.scalar() or 0}")

    # Upcoming meetings
    meeting_result = await db.execute(
        select(func.count(Meeting.id)).where(Meeting.scheduled_at >= now)
    )
    lines.append(f"  Upcoming meetings: {meeting_result.scalar() or 0}")

    # Meals (last 7 days avg)
    cutoff = today - timedelta(days=7)
    meal_result = await db.execute(
        select(func.sum(DailyMealCount.quantity))
        .where(DailyMealCount.date >= cutoff)
    )
    total_meals_week = float(meal_result.scalar() or 0)
    lines.append(f"  Meals last 7 days: {total_meals_week:,.0f}")

    # Projects
    proj_result = await db.execute(
        select(func.count(Project.id)).where(Project.status.in_(["planning", "active"]))
    )
    lines.append(f"  Active projects: {proj_result.scalar() or 0}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Tool dispatch
# ---------------------------------------------------------------------------
TOOL_HANDLERS = {
    "query_spending": _exec_query_spending,
    "query_budgets": _exec_query_budgets,
    "query_meals": _exec_query_meals,
    "query_violations": _exec_query_violations,
    "query_meetings": _exec_query_meetings,
    "query_price_lists": _exec_query_price_lists,
    "query_projects": _exec_query_projects,
    "query_summary": _exec_query_summary,
}


async def _execute_tool(tool_name: str, tool_input: dict, db: AsyncSession, suppliers: dict, sites: dict) -> str:
    """Execute a tool and return the result as text."""
    handler = TOOL_HANDLERS.get(tool_name)
    if not handler:
        return f"Unknown tool: {tool_name}"

    try:
        # Different handlers have different signatures
        if tool_name in ("query_spending",):
            return await handler(db, tool_input, suppliers, sites)
        elif tool_name in ("query_budgets",):
            return await handler(db, tool_input, suppliers)
        elif tool_name in ("query_meals",):
            return await handler(db, tool_input, sites)
        elif tool_name in ("query_violations", "query_meetings"):
            return await handler(db, tool_input, sites)
        elif tool_name in ("query_price_lists",):
            return await handler(db, tool_input, suppliers)
        elif tool_name in ("query_projects",):
            return await handler(db, tool_input)
        elif tool_name in ("query_summary",):
            return await handler(db, sites, suppliers)
        else:
            return await handler(db, tool_input)
    except Exception as e:
        logger.error(f"Tool execution error ({tool_name}): {e}")
        return f"Error executing {tool_name}: {str(e)}"


# ---------------------------------------------------------------------------
# Chat endpoint
# ---------------------------------------------------------------------------
MONTH_COLS = ["jan", "feb", "mar", "apr", "may", "jun",
              "jul", "aug", "sep", "oct", "nov", "dec"]


@router.post("/", response_model=ChatResponse)
@router.post("", response_model=ChatResponse, include_in_schema=False)
async def chat(
    chat_message: ChatMessage,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Send a message to the AI assistant with tool-use for full data access."""
    if not claude_service.is_available:
        raise HTTPException(
            status_code=503,
            detail="AI service is not configured. Please set the ANTHROPIC_API_KEY environment variable."
        )

    try:
        # Load reference data once
        supplier_result = await db.execute(select(Supplier).where(Supplier.is_active == True))
        suppliers = {s.id: s.name for s in supplier_result.scalars().all()}

        site_result = await db.execute(select(Site).where(Site.is_active == True))
        sites = {s.id: s.name for s in site_result.scalars().all()}

        # Start conversation with user message
        messages = [{"role": "user", "content": chat_message.message}]

        # Tool-use loop (max 5 rounds to prevent infinite loops)
        final_response = ""
        for _ in range(5):
            response = await claude_service.generate_with_tools(
                messages=messages,
                system_prompt=CHAT_SYSTEM_PROMPT,
                tools=CHAT_TOOLS,
                max_tokens=4096,
            )

            # Check if Claude wants to use tools
            if response.stop_reason == "tool_use":
                # Process all tool calls in this response
                tool_results = []
                assistant_content = []

                for block in response.content:
                    if block.type == "text":
                        assistant_content.append({"type": "text", "text": block.text})
                    elif block.type == "tool_use":
                        assistant_content.append({
                            "type": "tool_use",
                            "id": block.id,
                            "name": block.name,
                            "input": block.input,
                        })
                        # Execute the tool
                        result = await _execute_tool(block.name, block.input, db, suppliers, sites)
                        tool_results.append({
                            "type": "tool_result",
                            "tool_use_id": block.id,
                            "content": result,
                        })

                # Add assistant response + tool results to messages
                messages.append({"role": "assistant", "content": assistant_content})
                messages.append({"role": "user", "content": tool_results})
            else:
                # Claude is done — extract final text response
                for block in response.content:
                    if block.type == "text":
                        final_response += block.text
                break

        if not final_response:
            final_response = "I couldn't generate a response. Please try rephrasing your question."

        suggestions = [
            "Show spending breakdown by supplier",
            "What products were ordered this month?",
            "Check budget status",
        ]

        return ChatResponse(response=final_response, suggestions=suggestions)

    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail="AI service temporarily unavailable. Please try again later.")


# ---------------------------------------------------------------------------
# Data export endpoint — generate downloadable Excel from query
# ---------------------------------------------------------------------------
class ExportRequest(BaseModel):
    query_type: str  # spending, budgets, meals, violations, price_lists
    filters: dict = {}


@router.post("/export")
async def export_data(
    req: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export data as Excel file based on query type and filters."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    # Load reference data
    supplier_result = await db.execute(select(Supplier).where(Supplier.is_active == True))
    suppliers = {s.id: s.name for s in supplier_result.scalars().all()}
    site_result = await db.execute(select(Site).where(Site.is_active == True))
    sites = {s.id: s.name for s in site_result.scalars().all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.query_type.replace("_", " ").title()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    if req.query_type == "spending":
        # Export proforma items
        year = req.filters.get("year", date.today().year)
        month = req.filters.get("month")
        site_name = req.filters.get("site_name")
        supplier_name = req.filters.get("supplier_name")

        filters = [Proforma.invoice_date.isnot(None)]
        if month:
            start = date(year, month, 1)
            end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
            filters.append(Proforma.invoice_date >= start)
            filters.append(Proforma.invoice_date < end)
        else:
            filters.append(Proforma.invoice_date >= date(year, 1, 1))

        site_id = _match_site(site_name, sites) if site_name else None
        if site_id:
            filters.append(Proforma.site_id == site_id)

        if supplier_name:
            matching = [sid for sid, sn in suppliers.items() if supplier_name.lower() in sn.lower()]
            if matching:
                filters.append(Proforma.supplier_id.in_(matching))

        proforma_result = await db.execute(
            select(Proforma).where(and_(*filters)).order_by(Proforma.invoice_date)
        )
        proformas = proforma_result.scalars().all()
        proforma_ids = [p.id for p in proformas]
        proforma_map = {p.id: p for p in proformas}

        headers = ["Date", "Supplier", "Site", "Product", "Quantity", "Unit", "Unit Price", "Total"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        if proforma_ids:
            item_result = await db.execute(
                select(ProformaItem)
                .where(ProformaItem.proforma_id.in_(proforma_ids))
                .order_by(ProformaItem.proforma_id)
            )
            items = item_result.scalars().all()
            for row_idx, item in enumerate(items, 2):
                p = proforma_map.get(item.proforma_id)
                ws.cell(row=row_idx, column=1, value=str(p.invoice_date) if p else "")
                ws.cell(row=row_idx, column=2, value=suppliers.get(p.supplier_id, "?") if p else "")
                ws.cell(row=row_idx, column=3, value=sites.get(p.site_id, "?") if p else "")
                ws.cell(row=row_idx, column=4, value=item.product_name)
                ws.cell(row=row_idx, column=5, value=item.quantity)
                ws.cell(row=row_idx, column=6, value=item.unit or "")
                ws.cell(row=row_idx, column=7, value=item.unit_price)
                ws.cell(row=row_idx, column=8, value=item.total_price)

    elif req.query_type == "budgets":
        year = req.filters.get("year", date.today().year)
        budget_result = await db.execute(
            select(SupplierBudget).where(SupplierBudget.year == year, SupplierBudget.is_active == True)
        )
        budgets = budget_result.scalars().all()

        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        month_cols = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
        headers = ["Supplier", "Yearly Total"] + month_names
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, b in enumerate(budgets, 2):
            ws.cell(row=row_idx, column=1, value=suppliers.get(b.supplier_id, "?"))
            ws.cell(row=row_idx, column=2, value=b.yearly_amount or 0)
            for mi, col_name in enumerate(month_cols):
                ws.cell(row=row_idx, column=3 + mi, value=getattr(b, col_name) or 0)

    elif req.query_type == "meals":
        days = req.filters.get("days", 30)
        cutoff = date.today() - timedelta(days=days)
        meal_result = await db.execute(
            select(DailyMealCount)
            .where(DailyMealCount.date >= cutoff)
            .order_by(DailyMealCount.date.desc())
        )
        meals = meal_result.scalars().all()

        headers = ["Date", "Site", "Meal Type", "Quantity"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, m in enumerate(meals, 2):
            ws.cell(row=row_idx, column=1, value=str(m.date))
            ws.cell(row=row_idx, column=2, value=sites.get(m.site_id, "?"))
            ws.cell(row=row_idx, column=3, value=m.meal_type_en or m.meal_type or "?")
            ws.cell(row=row_idx, column=4, value=m.quantity)

    else:
        raise HTTPException(status_code=400, detail=f"Unsupported export type: {req.query_type}")

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{req.query_type}_export_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
