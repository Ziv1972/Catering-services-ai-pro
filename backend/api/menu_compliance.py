"""
Menu Compliance API endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional, Dict, Any
from datetime import date
from pathlib import Path
from pydantic import BaseModel
import logging
import re

from sqlalchemy.orm.attributes import flag_modified
from backend.database import get_db

logger = logging.getLogger(__name__)
from backend.models.user import User
from backend.models.menu_compliance import MenuCheck, MenuDay, CheckResult, ComplianceRule
from backend.api.auth import get_current_user

router = APIRouter()

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB
SAFE_FILENAME_RE = re.compile(r'^[\w\-. ()\[\]]+$')


def _sanitize_filename(filename: str) -> str:
    """Sanitize upload filename to prevent path traversal."""
    # Take only the basename, stripping any directory components
    safe_name = Path(filename).name.strip()
    # Remove any remaining path separators
    safe_name = safe_name.replace("/", "").replace("\\", "").replace("..", "")
    if not safe_name or not SAFE_FILENAME_RE.match(safe_name):
        safe_name = "upload.xlsx"
    return safe_name


def _normalize_special_chars(text: str) -> str:
    """Normalize special characters and whitespace for matching.

    Handles:
    - '&' with varying spaces: 'פיש & ציפס', 'פיש& ציפס', 'פיש&ציפס' → all same
    - Multiple spaces → single space
    - Various dash types → standard dash
    - Trim whitespace
    """
    import re
    # Normalize spaces around & (remove all spaces around &, then add exactly one on each side)
    text = re.sub(r'\s*&\s*', ' & ', text)
    # Normalize various dash types
    text = re.sub(r'[–—−]', '-', text)
    # Normalize spaces around dashes
    text = re.sub(r'\s*-\s*', ' - ', text)
    # Collapse multiple spaces to single
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _generate_ampersand_variants(text: str) -> list:
    """Generate variants with different & spacing for substring matching.

    For 'פיש & ציפס' generates:
    - 'פיש & ציפס'  (normalized with spaces)
    - 'פיש& ציפס'   (no space before &)
    - 'פיש &ציפס'   (no space after &)
    - 'פיש&ציפס'    (no spaces around &)
    """
    if '&' not in text:
        return [text]
    # Start with the normalized form
    normalized = _normalize_special_chars(text)
    variants = [normalized]
    # Generate spacing variants around &
    # Split on ' & ' and rejoin with different spacing
    parts = normalized.split(' & ')
    if len(parts) == 2:
        left, right = parts[0], parts[1]
        for fmt in ['{}&{}', '{}& {}', '{} &{}']:
            v = fmt.format(left, right)
            if v not in variants:
                variants.append(v)
    return variants


class ComplianceRuleResponse(BaseModel):
    id: int
    name: str
    rule_type: str
    description: Optional[str]
    category: Optional[str]
    parameters: Optional[dict]
    priority: int
    is_active: bool

    class Config:
        from_attributes = True


class ComplianceRuleCreate(BaseModel):
    name: str
    rule_type: str = "mandatory"
    description: Optional[str] = None
    category: Optional[str] = None
    parameters: Optional[dict] = None
    priority: int = 1


class ComplianceRuleUpdate(BaseModel):
    name: Optional[str] = None
    rule_type: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    parameters: Optional[dict] = None
    priority: Optional[int] = None
    is_active: Optional[bool] = None


class CheckResultResponse(BaseModel):
    id: int
    rule_name: str
    rule_category: Optional[str]
    passed: bool
    severity: str
    finding_text: Optional[str]
    evidence: Optional[dict]
    reviewed: bool
    review_status: Optional[str]
    review_notes: Optional[str]

    class Config:
        from_attributes = True


class MenuCheckResponse(BaseModel):
    id: int
    site_id: int
    month: str
    year: int
    total_findings: int
    critical_findings: int
    warnings: int
    passed_rules: int
    dishes_above: int = 0
    dishes_under: int = 0
    dishes_even: int = 0
    checked_at: date
    file_path: Optional[str]

    class Config:
        from_attributes = True


class MenuCheckDetailResponse(MenuCheckResponse):
    site_name: Optional[str] = None


@router.get("/checks", response_model=List[MenuCheckDetailResponse])
async def list_checks(
    site_id: Optional[int] = None,
    year: Optional[int] = None,
    limit: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List menu compliance checks"""
    query = select(MenuCheck).options(selectinload(MenuCheck.site))

    if site_id:
        query = query.where(MenuCheck.site_id == site_id)
    if year:
        query = query.where(MenuCheck.year == year)

    query = query.order_by(MenuCheck.checked_at.desc())

    if limit:
        query = query.limit(limit)

    result = await db.execute(query)
    checks = result.scalars().all()

    return [
        MenuCheckDetailResponse(
            id=c.id,
            site_id=c.site_id,
            month=c.month,
            year=c.year,
            total_findings=c.total_findings,
            critical_findings=c.critical_findings,
            warnings=c.warnings,
            passed_rules=c.passed_rules,
            dishes_above=c.dishes_above or 0,
            dishes_under=c.dishes_under or 0,
            dishes_even=c.dishes_even or 0,
            checked_at=c.checked_at,
            file_path=c.file_path,
            site_name=c.site.name if c.site else None,
        )
        for c in checks
    ]


@router.get("/checks/{check_id}")
async def get_check(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a single menu check with site info"""
    result = await db.execute(
        select(MenuCheck)
        .options(selectinload(MenuCheck.site))
        .where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()

    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    return {
        "id": check.id,
        "site_id": check.site_id,
        "site_name": check.site.name if check.site else None,
        "month": check.month,
        "year": check.year,
        "total_findings": check.total_findings,
        "critical_findings": check.critical_findings,
        "warnings": check.warnings,
        "passed_rules": check.passed_rules,
        "dishes_above": check.dishes_above or 0,
        "dishes_under": check.dishes_under or 0,
        "dishes_even": check.dishes_even or 0,
        "checked_at": check.checked_at.isoformat(),
        "file_path": check.file_path,
    }


@router.get("/checks/{check_id}/results", response_model=List[CheckResultResponse])
async def get_check_results(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get results for a menu check"""
    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Menu check not found")

    results = await db.execute(
        select(CheckResult)
        .where(CheckResult.menu_check_id == check_id)
        .order_by(CheckResult.passed.asc(), CheckResult.severity.asc())
    )

    return results.scalars().all()


@router.get("/stats")
async def get_compliance_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get overall compliance statistics"""
    result = await db.execute(
        select(
            func.count(MenuCheck.id).label("total_checks"),
            func.sum(MenuCheck.critical_findings).label("total_critical"),
            func.sum(MenuCheck.warnings).label("total_warnings"),
            func.sum(MenuCheck.passed_rules).label("total_passed"),
            func.sum(MenuCheck.total_findings).label("total_findings"),
            func.sum(MenuCheck.dishes_above).label("total_above"),
            func.sum(MenuCheck.dishes_under).label("total_under"),
            func.sum(MenuCheck.dishes_even).label("total_even"),
        )
    )
    row = result.one()

    return {
        "total_checks": row.total_checks or 0,
        "total_critical": row.total_critical or 0,
        "total_warnings": row.total_warnings or 0,
        "total_passed": row.total_passed or 0,
        "total_findings": row.total_findings or 0,
        "total_above": row.total_above or 0,
        "total_under": row.total_under or 0,
        "total_even": row.total_even or 0,
    }


@router.post("/upload-menu")
async def upload_menu(
    site_id: int = Form(...),
    month: str = Form(...),
    year: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload a menu file and run compliance analysis"""
    import os
    from backend.services.menu_analysis_service import run_compliance_check

    upload_dir = "uploads/menus"
    os.makedirs(upload_dir, exist_ok=True)

    safe_name = _sanitize_filename(file.filename)
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    file_path = f"{upload_dir}/{site_id}_{year}_{month}_{safe_name}"
    # Verify resolved path stays within upload_dir
    resolved = Path(file_path).resolve()
    if not str(resolved).startswith(str(Path(upload_dir).resolve())):
        raise HTTPException(status_code=400, detail="Invalid filename")

    with open(file_path, "wb") as f:
        f.write(content)

    check = MenuCheck(
        site_id=site_id,
        month=month,
        year=year,
        file_path=file_path,
        checked_at=date.today(),
        total_findings=0,
        critical_findings=0,
        warnings=0,
        passed_rules=0,
    )
    db.add(check)
    await db.commit()
    await db.refresh(check)

    # Run compliance analysis
    try:
        analysis = await run_compliance_check(check.id, db)
        return {
            "id": check.id,
            "message": "Menu uploaded and compliance check completed.",
            "file_path": file_path,
            "analysis": analysis,
        }
    except Exception as e:
        return {
            "id": check.id,
            "message": f"Menu uploaded. Compliance check failed: {str(e)}",
            "file_path": file_path,
        }


@router.post("/checks/{check_id}/run")
async def rerun_check(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Re-run compliance analysis on an existing menu check"""
    from backend.services.menu_analysis_service import run_compliance_check

    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    # Clear old results (days are handled by run_compliance_check)
    from sqlalchemy import delete
    await db.execute(
        delete(CheckResult).where(CheckResult.menu_check_id == check_id)
    )
    await db.commit()

    analysis = await run_compliance_check(check_id, db)
    return {
        "id": check_id,
        "message": "Compliance check re-run completed.",
        "analysis": analysis,
    }


@router.post("/checks/{check_id}/reupload")
async def reupload_menu_file(
    check_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Re-upload a menu file for an existing check, then re-parse and re-run."""
    import os
    from sqlalchemy import delete as sql_delete
    from backend.services.menu_analysis_service import run_compliance_check

    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    # Save new file
    upload_dir = "uploads/menus"
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = _sanitize_filename(file.filename)
    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large (max 10 MB)")

    file_path = f"{upload_dir}/{check.site_id}_{check.year}_{check.month}_{safe_name}"
    resolved = Path(file_path).resolve()
    if not str(resolved).startswith(str(Path(upload_dir).resolve())):
        raise HTTPException(status_code=400, detail="Invalid filename")

    with open(file_path, "wb") as f:
        f.write(content)

    # Update check with new file path
    check.file_path = file_path
    check.checked_at = date.today()

    # Clear old results AND old days (force fresh parse)
    await db.execute(
        sql_delete(CheckResult).where(CheckResult.menu_check_id == check_id)
    )
    await db.execute(
        sql_delete(MenuDay).where(MenuDay.menu_check_id == check_id)
    )
    await db.commit()

    # Run fresh compliance analysis (will parse the new file)
    try:
        analysis = await run_compliance_check(check_id, db)
        return {
            "id": check_id,
            "message": "File re-uploaded and compliance check completed.",
            "file_path": file_path,
            "analysis": analysis,
        }
    except Exception as e:
        return {
            "id": check_id,
            "message": f"File re-uploaded. Compliance check failed: {str(e)}",
            "file_path": file_path,
        }


@router.delete("/checks/{check_id}")
async def delete_check(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a menu compliance check and all its results and days."""
    from sqlalchemy import delete as sql_delete

    result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    # Delete related data
    await db.execute(
        sql_delete(CheckResult).where(CheckResult.menu_check_id == check_id)
    )
    await db.execute(
        sql_delete(MenuDay).where(MenuDay.menu_check_id == check_id)
    )
    await db.delete(check)
    await db.commit()

    return {"message": "Check deleted successfully"}


@router.get("/checks/{check_id}/menu-items")
async def get_menu_items(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all parsed menu items for a check, grouped by day."""
    result = await db.execute(
        select(MenuDay)
        .where(MenuDay.menu_check_id == check_id)
        .order_by(MenuDay.date)
    )
    days = result.scalars().all()

    if not days:
        raise HTTPException(status_code=404, detail="No parsed menu days found for this check")

    return [
        {
            "id": day.id,
            "date": day.date,
            "day_of_week": day.day_of_week,
            "week_number": day.week_number,
            "items": day.menu_items or {},
        }
        for day in days
    ]


@router.get("/checks/{check_id}/search-items")
async def search_menu_items(
    check_id: int,
    keyword: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Search for a keyword across parsed menu items AND raw file.

    Two-pass search:
    1. Parsed data (MenuDay items) — structured matches
    2. Raw file search (like Excel Ctrl+F) — catches items AI parsing missed
    """
    import os
    from backend.services.menu_analysis_service import (
        _normalize_hebrew, _strip_hebrew_prefixes
    )

    # Get the check to access file_path
    check_result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = check_result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    result = await db.execute(
        select(MenuDay)
        .where(MenuDay.menu_check_id == check_id)
        .order_by(MenuDay.date)
    )
    days = result.scalars().all()

    keyword_lower = keyword.lower().strip()
    # Normalize special chars (& spacing, dashes, etc.)
    keyword_lower = _normalize_special_chars(keyword_lower)

    # Also strip Hebrew prefixes from the keyword itself for broader search
    keyword_stems = _strip_hebrew_prefixes(keyword_lower)
    variants = []
    for stem in keyword_stems:
        for v in _normalize_hebrew(stem):
            # Add the base variant
            if v not in variants:
                variants.append(v)
            # Add ampersand spacing variants
            for av in _generate_ampersand_variants(v):
                if av not in variants:
                    variants.append(av)

    # --- Pass 1: Search parsed MenuDay data ---
    matches = []
    if days:
        for day in days:
            items = day.menu_items or {}
            for category, item_list in items.items():
                if not isinstance(item_list, list):
                    continue
                for item in item_list:
                    item_str = str(item).lower()
                    match_type = _classify_match(keyword_lower, variants, item_str)
                    if match_type:
                        matches.append({
                            "date": day.date,
                            "day_of_week": day.day_of_week,
                            "category": category,
                            "item": item,
                            "match_type": match_type,
                            "source": "parsed",
                        })

    # --- Pass 2: Search raw file (like Ctrl+F) ---
    raw_file_matches = []
    if check.file_path and os.path.exists(check.file_path):
        raw_file_matches = _search_raw_file(
            check.file_path, keyword_lower, variants
        )

    # Merge: add raw file matches that weren't already found in parsed data
    parsed_items_lower = {m["item"].lower() for m in matches}
    for rfm in raw_file_matches:
        if rfm["item"].lower() not in parsed_items_lower:
            matches.append(rfm)

    # Sort: exact first, then contains, then prefix, then raw_file, then fuzzy
    type_order = {
        "exact": 0, "contains": 1, "prefix": 2,
        "raw_file": 3, "fuzzy": 4,
    }
    matches.sort(key=lambda m: (type_order.get(m["match_type"], 9), m.get("date", "")))

    # Get unique items for summary
    unique_items = {}
    for m in matches:
        key = m["item"]
        if key not in unique_items:
            unique_items[key] = {
                "item": m["item"],
                "match_type": m["match_type"],
                "source": m.get("source", "parsed"),
                "days": [],
            }
        day_val = m.get("date", "")
        if day_val and day_val not in unique_items[key]["days"]:
            unique_items[key]["days"].append(day_val)

    return {
        "keyword": keyword,
        "variants_searched": variants,
        "total_matches": len(matches),
        "unique_items": list(unique_items.values()),
        "matches_by_day": matches,
        "raw_file_searched": bool(raw_file_matches) or (
            check.file_path and os.path.exists(check.file_path)
        ),
    }


def _search_raw_file(
    file_path: str, keyword: str, variants: list[str]
) -> list[dict]:
    """Search the actual uploaded file for keyword (like Excel Ctrl+F).

    Reads the raw file and finds every cell/line containing the keyword.
    Tries to associate matches with dates from column headers.
    """
    import re
    matches = []

    try:
        if file_path.endswith((".xlsx", ".xls")):
            matches = _search_excel_file(file_path, keyword, variants)
        elif file_path.endswith(".csv"):
            matches = _search_csv_file(file_path, keyword, variants)
        else:
            matches = _search_text_file(file_path, keyword, variants)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Raw file search error: {e}")

    return matches


def _search_excel_file(
    file_path: str, keyword: str, variants: list[str]
) -> list[dict]:
    """Search Excel file cell by cell, tracking column dates."""
    try:
        import openpyxl
    except ImportError:
        return []

    matches = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            continue

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # Detect date row — look for dates in first 5 rows
        col_dates: dict[int, str] = {}
        import re
        from datetime import date as date_type
        for row in all_rows[:5]:
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                # Try date patterns: "3/19/2026", "19.3.2026", "2026-03-19"
                date_match = re.search(
                    r'(\d{1,2})[./](\d{1,2})[./](\d{2,4})', cell_str
                )
                if date_match:
                    parts = [int(x) for x in date_match.groups()]
                    # Try M/D/Y and D/M/Y
                    for m, d, y in [(parts[0], parts[1], parts[2]),
                                    (parts[1], parts[0], parts[2])]:
                        if y < 100:
                            y += 2000
                        try:
                            dt = date_type(y, m, d)
                            col_dates[col_idx] = dt.isoformat()
                            break
                        except (ValueError, TypeError):
                            continue
                # Also try ISO format
                iso_match = re.match(r'(\d{4})-(\d{2})-(\d{2})', cell_str)
                if iso_match and col_idx not in col_dates:
                    col_dates[col_idx] = cell_str[:10]

        # Search all cells for keyword
        for row in all_rows:
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                cell_lower = cell_str.lower()
                # Also check normalized form for & and spacing differences
                cell_normalized = _normalize_special_chars(cell_lower)

                # Check if any variant is a substring of this cell
                found = False
                for var in variants:
                    if var in cell_lower or var in cell_normalized:
                        found = True
                        break

                if found and len(cell_str) >= 3:
                    # Find the date for this column
                    matched_date = col_dates.get(col_idx, "")
                    matches.append({
                        "date": matched_date,
                        "day_of_week": "",
                        "category": f"Sheet: {sheet_name}",
                        "item": cell_str,
                        "match_type": "raw_file",
                        "source": "raw_file",
                    })

    wb.close()
    return matches


def _search_csv_file(
    file_path: str, keyword: str, variants: list[str]
) -> list[dict]:
    """Search CSV file line by line."""
    import csv
    matches = []

    for enc in ("utf-8-sig", "cp1255", "latin-1"):
        try:
            with open(file_path, "r", encoding=enc) as f:
                reader = csv.reader(f)
                for row_idx, row in enumerate(reader):
                    for cell in row:
                        cell_str = cell.strip()
                        cell_lower = cell_str.lower()
                        cell_normalized = _normalize_special_chars(cell_lower)
                        for var in variants:
                            if (var in cell_lower or var in cell_normalized) and len(cell_str) >= 3:
                                matches.append({
                                    "date": "",
                                    "day_of_week": "",
                                    "category": f"Row {row_idx + 1}",
                                    "item": cell_str,
                                    "match_type": "raw_file",
                                    "source": "raw_file",
                                })
                                break
            break
        except (UnicodeDecodeError, UnicodeError):
            continue

    return matches


def _search_text_file(
    file_path: str, keyword: str, variants: list[str]
) -> list[dict]:
    """Search plain text file line by line."""
    matches = []
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            for line_num, line in enumerate(f, 1):
                line_str = line.strip()
                line_lower = line_str.lower()
                line_normalized = _normalize_special_chars(line_lower)
                for var in variants:
                    if (var in line_lower or var in line_normalized) and len(line_str) >= 3:
                        matches.append({
                            "date": "",
                            "day_of_week": "",
                            "category": f"Line {line_num}",
                            "item": line_str,
                            "match_type": "raw_file",
                            "source": "raw_file",
                        })
                        break
    except Exception:
        pass
    return matches


def _classify_match(keyword: str, variants: list[str], item_text: str) -> str | None:
    """Classify how well a keyword matches an item.

    Returns match type or None if no match:
    - "exact": keyword equals item or item word
    - "contains": keyword is a substring of item
    - "prefix": keyword matches after stripping Hebrew prefix letters
    """
    from backend.services.menu_analysis_service import _strip_hebrew_prefixes

    # Also compare using normalized forms (handles & spacing, dashes, etc.)
    item_normalized = _normalize_special_chars(item_text)

    # Exact match (whole item or any word)
    words = item_text.split()
    words_normalized = item_normalized.split()
    for var in variants:
        if var == item_text or var == item_normalized:
            return "exact"
        if var in words or var in words_normalized:
            return "exact"

    # Contains (substring)
    for var in variants:
        if var in item_text or var in item_normalized:
            return "contains"

    # Prefix-stripped match — only check var==stem or var-in-stem.
    # Never check stem-in-var (causes false positives: "צל" in "צלי כתף")
    all_words = set(words + words_normalized)
    for word in all_words:
        stems = _strip_hebrew_prefixes(word)
        for stem in stems:
            for var in variants:
                if var == stem:
                    return "prefix"
                if var in stem and len(var) >= 3:
                    return "prefix"

    # No fuzzy matching — too many false positives with Hebrew
    # Raw file search handles the cases fuzzy was meant to catch
    return None


# --- Match Approval ---


class ApproveMatchesRequest(BaseModel):
    approved_items: List[Dict[str, Any]]


@router.put("/checks/{check_id}/results/{result_id}/approve-matches")
async def approve_matches(
    check_id: int,
    result_id: int,
    data: ApproveMatchesRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Approve/reject search matches and update the CheckResult actual count.

    The user reviews search results (parsed + raw file) and approves
    the items they confirm are valid matches. This updates the actual
    count to reflect the user's review.
    """
    try:
        result = await db.execute(
            select(CheckResult).where(
                CheckResult.id == result_id,
                CheckResult.menu_check_id == check_id
            )
        )
        check_result = result.scalar_one_or_none()
        if not check_result:
            raise HTTPException(status_code=404, detail="Check result not found")

        evidence = dict(check_result.evidence or {})

        # Count unique dates across all approved items
        approved_days: set[str] = set()
        for item in data.approved_items:
            days_list = item.get("days", [])
            if isinstance(days_list, list):
                for day in days_list:
                    if day and isinstance(day, str) and len(day) >= 8:
                        approved_days.add(day)

        new_actual = len(approved_days)

        # Fallback: if no dates detected but items were approved, count items
        if new_actual == 0 and len(data.approved_items) > 0:
            new_actual = len(data.approved_items)

        expected = evidence.get("expected_count", 0)
        is_max = evidence.get("is_max_rule", False)

        # Re-derive comparison and passed
        if new_actual > expected:
            new_comparison = "above"
        elif new_actual < expected:
            new_comparison = "under"
        else:
            new_comparison = "even"

        if is_max:
            new_passed = new_actual <= expected
        else:
            new_passed = new_actual >= expected

        # Build updated evidence (new dict for SQLAlchemy change detection)
        sorted_days = sorted(approved_days) if approved_days else []
        updated_evidence = dict(evidence)
        updated_evidence["actual_count"] = new_actual
        updated_evidence["comparison"] = new_comparison
        updated_evidence["found_on_days"] = sorted_days
        updated_evidence["user_approved_items"] = [
            {"item": it.get("item", ""), "days": it.get("days", []), "source": it.get("source", "")}
            for it in data.approved_items
        ]
        updated_evidence["user_reviewed"] = True

        item_keyword = evidence.get("item_searched", evidence.get("category_keyword", ""))
        freq_key = "monthly_freq" if "monthly_freq" in evidence else "weekly_freq"
        is_monthly = "monthly_freq" in evidence

        new_finding = (
            f"'{item_keyword}': Expected "
            f"{'≤' if is_max else '≥'}{expected}"
            f"{'/' + ('month' if is_monthly else 'week') if freq_key in evidence else ''}, "
            f"Actual {new_actual} (user-reviewed)"
        )

        # Update the result columns
        check_result.evidence = updated_evidence
        flag_modified(check_result, "evidence")
        check_result.passed = new_passed
        check_result.finding_text = new_finding
        check_result.severity = "info" if new_passed else "critical"
        check_result.reviewed = True
        check_result.review_status = "approved"

        await db.commit()
        await db.refresh(check_result)

        logger.info(
            f"Approve matches: result={result_id}, "
            f"approved={len(data.approved_items)} items, "
            f"days={sorted_days}, actual={new_actual}, expected={expected}"
        )

        # Recalculate parent MenuCheck summary counts
        all_results = await db.execute(
            select(CheckResult).where(CheckResult.menu_check_id == check_id)
        )
        all_cr = all_results.scalars().all()

        above_count = sum(1 for cr in all_cr if (cr.evidence or {}).get("comparison") == "above")
        under_count = sum(1 for cr in all_cr if (cr.evidence or {}).get("comparison") == "under")
        even_count = sum(1 for cr in all_cr if (cr.evidence or {}).get("comparison") == "even")
        critical_count = sum(1 for cr in all_cr if cr.severity == "critical")
        warning_count = sum(1 for cr in all_cr if cr.severity == "warning")
        passed_count = sum(1 for cr in all_cr if cr.passed)

        check_obj = await db.execute(
            select(MenuCheck).where(MenuCheck.id == check_id)
        )
        menu_check = check_obj.scalar_one_or_none()
        if menu_check:
            menu_check.dishes_above = above_count
            menu_check.dishes_under = under_count
            menu_check.dishes_even = even_count
            menu_check.critical_findings = critical_count
            menu_check.warnings = warning_count
            menu_check.passed_rules = passed_count
            menu_check.total_findings = critical_count + warning_count
            await db.commit()

        return {
            "result_id": result_id,
            "old_actual": evidence.get("actual_count", 0),
            "new_actual": new_actual,
            "expected": expected,
            "passed": new_passed,
            "comparison": new_comparison,
            "approved_days": sorted_days,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error approving matches: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to save: {str(e)}")


# --- Compliance Rules CRUD ---


@router.get("/rules", response_model=List[ComplianceRuleResponse])
async def list_rules(
    active_only: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List all compliance rules"""
    query = select(ComplianceRule).order_by(ComplianceRule.priority, ComplianceRule.name)
    if active_only:
        query = query.where(ComplianceRule.is_active == True)
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/rules", response_model=ComplianceRuleResponse)
async def create_rule(
    data: ComplianceRuleCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new compliance rule"""
    rule = ComplianceRule(**data.model_dump(exclude_none=True), is_active=True)
    db.add(rule)
    await db.commit()
    await db.refresh(rule)
    return rule


@router.put("/rules/{rule_id}", response_model=ComplianceRuleResponse)
async def update_rule(
    rule_id: int,
    data: ComplianceRuleUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update a compliance rule"""
    result = await db.execute(
        select(ComplianceRule).where(ComplianceRule.id == rule_id)
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")

    updates = data.model_dump(exclude_none=True)
    for key, value in updates.items():
        if key == "parameters" and isinstance(value, dict):
            # Merge new parameters into existing ones (don't replace)
            existing = rule.parameters or {}
            setattr(rule, key, {**existing, **value})
        else:
            setattr(rule, key, value)

    await db.commit()
    await db.refresh(rule)
    return rule


@router.delete("/rules/{rule_id}")
async def delete_rule(
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Deactivate a compliance rule"""
    result = await db.execute(
        select(ComplianceRule).where(ComplianceRule.id == rule_id)
    )
    rule = result.scalar_one_or_none()
    if not rule:
        raise HTTPException(status_code=404, detail="Rule not found")

    rule.is_active = False
    await db.commit()
    return {"message": "Rule deactivated"}


# ---------------------------------------------------------------------------
# AI-powered compliance check
# ---------------------------------------------------------------------------

@router.post("/checks/{check_id}/ai-check")
async def run_ai_check(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Run AI-powered compliance check using Claude.

    Sends the full parsed menu + contract rules to Claude for intelligent
    matching. Returns results in the manual-check spreadsheet format and
    generates a downloadable Excel file with the check sheet added.
    """
    from backend.services.menu_analysis_service import run_ai_compliance_check

    check_result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = check_result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    try:
        ai_results = await run_ai_compliance_check(check_id, db)
    except RuntimeError as e:
        raise HTTPException(status_code=503, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return {
        "message": f"AI check complete: {len(ai_results)} items checked",
        "total_items": len(ai_results),
        "shortages": sum(1 for r in ai_results if r.get("shortage", 0) > 0),
        "surplus": sum(1 for r in ai_results if r.get("shortage", 0) < 0),
        "ok": sum(1 for r in ai_results if r.get("shortage", 0) == 0),
        "results": ai_results,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

@router.get("/checks/{check_id}/export-excel")
async def export_compliance_excel(
    check_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export compliance check results as Excel with summary sheet."""
    import io
    import os
    from copy import copy
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse

    check_result = await db.execute(
        select(MenuCheck).where(MenuCheck.id == check_id)
    )
    check = check_result.scalar_one_or_none()
    if not check:
        raise HTTPException(status_code=404, detail="Menu check not found")

    results_q = await db.execute(
        select(CheckResult)
        .where(CheckResult.menu_check_id == check_id)
        .order_by(CheckResult.id)
    )
    results = results_q.scalars().all()

    # Try to open original menu file, otherwise create new workbook
    wb = None
    if check.file_path and os.path.exists(check.file_path):
        try:
            wb = load_workbook(check.file_path)
        except Exception:
            wb = None
    if wb is None:
        wb = Workbook()
        wb.active.title = "Menu"

    # Remove existing report sheet if present
    report_name = "חוסרים"
    if report_name in wb.sheetnames:
        del wb[report_name]
    ws = wb.create_sheet(report_name, 0)

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")
    red_font = Font(color="9C0006", bold=True)
    green_font = Font(color="006100", bold=True)
    blue_font = Font(color="003399", bold=True)
    group_fill = PatternFill("solid", fgColor="D9E2F3")
    group_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers — matching manual check format
    headers = ["קבוצה", "סוג", "תדירות מינימלית", "תקן", "בפועל", "חוסר", "פריטים שנמצאו בתפריט", "הערות"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 65
    ws.column_dimensions["H"].width = 25

    # Data rows — group by category
    row_num = 2
    current_group = None
    for r in results:
        evidence = r.evidence or {}
        group = r.rule_category or ""
        dish_name = r.rule_name or ""
        frequency_text = evidence.get("frequency_text", "")
        expected = evidence.get("expected_count")
        actual = evidence.get("actual_count")
        found_days = evidence.get("found_on_days") or []
        matched_items = evidence.get("matched_items") or []
        notes = evidence.get("notes", "")

        if expected is None:
            continue

        deficit = (expected - actual) if actual is not None else expected

        # Group header row
        if group != current_group:
            current_group = group
            group_cell = ws.cell(row=row_num, column=1, value=group)
            group_cell.font = group_font
            group_cell.fill = group_fill
            group_cell.border = thin_border
            # Only fill first cell with group name, leave rest empty but styled
            ws.cell(row=row_num, column=1, value=group)

        ws.cell(row=row_num, column=1, value=group).border = thin_border
        ws.cell(row=row_num, column=2, value=dish_name).border = thin_border
        ws.cell(row=row_num, column=3, value=frequency_text).border = thin_border

        exp_cell = ws.cell(row=row_num, column=4, value=expected)
        exp_cell.border = thin_border
        exp_cell.alignment = Alignment(horizontal="center")

        act_cell = ws.cell(row=row_num, column=5, value=actual if actual is not None else 0)
        act_cell.border = thin_border
        act_cell.alignment = Alignment(horizontal="center")

        def_cell = ws.cell(row=row_num, column=6, value=deficit)
        def_cell.border = thin_border
        def_cell.alignment = Alignment(horizontal="center")

        if deficit > 0:
            def_cell.fill = red_fill
            def_cell.font = red_font
        elif deficit < 0:
            def_cell.fill = blue_fill
            def_cell.font = blue_font
        else:
            def_cell.fill = green_fill
            def_cell.font = green_font

        # Show matched menu items (what Claude found)
        items_text = ", ".join(matched_items) if matched_items else ""
        items_cell = ws.cell(row=row_num, column=7, value=items_text)
        items_cell.border = thin_border
        items_cell.alignment = Alignment(wrap_text=True)

        notes_cell = ws.cell(row=row_num, column=8, value=notes)
        notes_cell.border = thin_border
        notes_cell.alignment = Alignment(wrap_text=True)

        row_num += 1

    # RTL sheet direction
    ws.sheet_view.rightToLeft = True

    # Write to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"compliance_check_{check.month}_{check.year}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
