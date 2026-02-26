"""
Universal Attachments API — upload files to any entity, with optional AI processing.
Supported entity_types: project, task, meeting, maintenance, todo, budget, complaint
"""
import os
import json
import uuid
import logging
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, field_validator

from backend.database import get_db
from backend.models.user import User
from backend.models.attachment import Attachment
from backend.api.auth import get_current_user

logger = logging.getLogger(__name__)

UPLOAD_DIR = "uploads/attachments"
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB

ALLOWED_EXTENSIONS = {
    ".txt", ".csv", ".json", ".md", ".log", ".xml", ".html",
    ".pdf", ".xlsx", ".xls", ".docx", ".doc",
    ".jpg", ".jpeg", ".png", ".gif", ".webp",
}

router = APIRouter()


# ─── Schemas ───

class AttachmentResponse(BaseModel):
    id: int
    entity_type: str
    entity_id: int
    filename: str
    original_filename: str
    file_size: Optional[int] = None
    content_type: Optional[str] = None
    ai_summary: Optional[str] = None
    ai_extracted_data: Optional[str] = None
    processing_status: Optional[str] = None
    uploaded_by: Optional[int] = None
    created_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class ProcessRequest(BaseModel):
    mode: str = "summarize"

    @field_validator("mode")
    @classmethod
    def validate_mode(cls, v: str) -> str:
        if v not in ("summarize", "extract", "both"):
            raise ValueError("mode must be 'summarize', 'extract', or 'both'")
        return v


# ─── File content extraction helpers ───

def _read_text_file(file_path: str) -> str:
    """Read plain text / CSV / JSON files."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read(50_000)
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            return f.read(50_000)


def _read_pdf_file(file_path: str) -> str:
    """Extract text from PDF using pdfplumber."""
    try:
        import pdfplumber
        text_parts = []
        with pdfplumber.open(file_path) as pdf:
            for page in pdf.pages[:20]:
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        return "\n\n".join(text_parts)[:50_000]
    except ImportError:
        return "[PDF reading not available — pdfplumber not installed]"
    except Exception as e:
        return f"[PDF read error: {e}]"


def _read_excel_file(file_path: str) -> str:
    """Extract text from Excel using openpyxl."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames[:3]:
            ws = wb[sheet_name]
            lines.append(f"=== Sheet: {sheet_name} ===")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append(" | ".join(cells))
                row_count += 1
                if row_count > 200:
                    lines.append("... (truncated)")
                    break
        wb.close()
        return "\n".join(lines)[:50_000]
    except ImportError:
        return "[Excel reading not available — openpyxl not installed]"
    except Exception as e:
        return f"[Excel read error: {e}]"


def extract_file_content(file_path: str, content_type: Optional[str], filename: str) -> Optional[str]:
    """Extract readable text from a file based on its type."""
    lower_name = filename.lower()

    if lower_name.endswith((".txt", ".csv", ".json", ".md", ".log", ".xml", ".html")):
        return _read_text_file(file_path)
    if lower_name.endswith(".pdf") or (content_type and "pdf" in content_type):
        return _read_pdf_file(file_path)
    if lower_name.endswith((".xlsx", ".xls")) or (content_type and "spreadsheet" in str(content_type)):
        return _read_excel_file(file_path)
    if content_type and content_type.startswith("text/"):
        return _read_text_file(file_path)

    return None


def _validate_file_path(file_path: str) -> None:
    """Prevent path traversal — ensure file is within UPLOAD_DIR."""
    real_path = os.path.realpath(file_path)
    upload_base = os.path.realpath(UPLOAD_DIR)
    if not real_path.startswith(upload_base):
        raise HTTPException(403, "Access denied")


# ─── Endpoints ───

@router.post("/upload", response_model=AttachmentResponse)
async def upload_attachment(
    entity_type: str = Form(...),
    entity_id: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload a file attachment to any entity."""
    allowed_types = {"project", "task", "meeting", "maintenance", "todo", "budget", "complaint"}
    if entity_type not in allowed_types:
        raise HTTPException(400, f"Invalid entity_type. Allowed: {', '.join(sorted(allowed_types))}")

    # Validate file extension
    ext = os.path.splitext(file.filename or "file")[1].lower()
    if ext and ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"File type '{ext}' not allowed.")

    # Read and validate file size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, f"File too large. Maximum size: {MAX_FILE_SIZE // (1024 * 1024)}MB")

    # Create upload directory
    entity_dir = f"{UPLOAD_DIR}/{entity_type}/{entity_id}"
    os.makedirs(entity_dir, exist_ok=True)

    # Generate unique filename
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = f"{entity_dir}/{unique_name}"

    # Save file to disk
    with open(file_path, "wb") as f:
        f.write(content)

    logger.info(f"User {current_user.id} uploaded '{file.filename}' for {entity_type}:{entity_id} ({len(content)} bytes)")

    attachment = Attachment(
        entity_type=entity_type,
        entity_id=entity_id,
        filename=unique_name,
        original_filename=file.filename or "file",
        file_path=file_path,
        file_size=len(content),
        content_type=file.content_type,
        uploaded_by=current_user.id,
    )
    db.add(attachment)
    await db.commit()
    await db.refresh(attachment)

    return attachment


@router.get("/", response_model=List[AttachmentResponse])
async def list_attachments(
    entity_type: str = Query(...),
    entity_id: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List all attachments for a specific entity."""
    result = await db.execute(
        select(Attachment)
        .where(Attachment.entity_type == entity_type, Attachment.entity_id == entity_id)
        .order_by(Attachment.created_at.desc())
    )
    return result.scalars().all()


@router.get("/{attachment_id}/download")
async def download_attachment(
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Download an attachment file."""
    result = await db.execute(
        select(Attachment).where(Attachment.id == attachment_id)
    )
    att = result.scalar_one_or_none()
    if not att:
        raise HTTPException(404, "Attachment not found")

    if not os.path.exists(att.file_path):
        raise HTTPException(404, "File not found on disk")

    _validate_file_path(att.file_path)

    return FileResponse(
        path=att.file_path,
        filename=att.original_filename,
        media_type=att.content_type or "application/octet-stream",
    )


@router.delete("/{attachment_id}")
async def delete_attachment(
    attachment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete an attachment."""
    result = await db.execute(
        select(Attachment).where(Attachment.id == attachment_id)
    )
    att = result.scalar_one_or_none()
    if not att:
        raise HTTPException(404, "Attachment not found")

    try:
        if os.path.exists(att.file_path):
            os.remove(att.file_path)
    except FileNotFoundError:
        logger.warning(f"File already deleted: {att.file_path}")
    except Exception as e:
        logger.error(f"Failed to delete file {att.file_path}: {e}")

    logger.info(f"User {current_user.id} deleted attachment {attachment_id} ({att.original_filename})")

    await db.delete(att)
    await db.commit()
    return {"success": True, "message": "Attachment deleted"}


@router.post("/{attachment_id}/process", response_model=AttachmentResponse)
async def process_attachment(
    attachment_id: int,
    body: ProcessRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    AI-process an attachment.
    Body: { "mode": "summarize" | "extract" | "both" }
    """
    result = await db.execute(
        select(Attachment).where(Attachment.id == attachment_id)
    )
    att = result.scalar_one_or_none()
    if not att:
        raise HTTPException(404, "Attachment not found")

    # Read file content
    file_text = extract_file_content(att.file_path, att.content_type, att.original_filename)
    if not file_text or not file_text.strip():
        raise HTTPException(
            400,
            f"Cannot read content from '{att.original_filename}'. "
            "AI processing supports: text, CSV, JSON, PDF, and Excel files."
        )

    att.processing_status = "processing"
    await db.commit()

    try:
        from backend.services.claude_service import claude_service

        truncated = file_text[:30_000]
        file_context = f"File: {att.original_filename} ({att.content_type or 'unknown'})\n\nContent:\n{truncated}"

        if body.mode in ("summarize", "both"):
            summary_prompt = (
                f"Analyze this document and provide a clear, structured summary in the same language as the document. "
                f"Include key points, important numbers/dates, and any action items.\n\n{file_context}"
            )
            att.ai_summary = await claude_service.generate_response(
                prompt=summary_prompt,
                system_prompt="You are a document analysis assistant for a catering operations manager. Provide concise, actionable summaries.",
                max_tokens=2000,
            )

        if body.mode in ("extract", "both"):
            extract_prompt = (
                f"Extract structured data from this document. Return a JSON object with relevant fields. "
                f"Identify: dates, amounts, names, quantities, categories, action items, and any key-value pairs.\n\n{file_context}"
            )
            extracted = await claude_service.generate_structured_response(
                prompt=extract_prompt,
                system_prompt="You are a data extraction assistant. Extract structured data as clean JSON.",
                response_format={"dates": [], "amounts": [], "items": [], "key_data": {}},
            )
            att.ai_extracted_data = json.dumps(extracted, ensure_ascii=False)

        att.processing_status = "done"
    except Exception as e:
        logger.error(f"AI processing failed for attachment {attachment_id}: {e}")
        att.processing_status = "error"
        if not att.ai_summary:
            att.ai_summary = f"Processing error: {str(e)[:200]}"

    await db.commit()
    await db.refresh(att)
    return att
