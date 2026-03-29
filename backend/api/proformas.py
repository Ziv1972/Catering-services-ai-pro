"""
Proformas (invoices) API endpoints
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload
from typing import List, Optional
from datetime import date, timedelta
from pydantic import BaseModel
import csv
import io
import json
import logging

from backend.database import get_db
from backend.models.user import User
from backend.models.proforma import Proforma, ProformaItem
from backend.models.supplier import Supplier
from backend.models.price_list import PriceList, PriceListItem
from backend.models.product import Product
from backend.api.auth import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter()


class ProformaItemResponse(BaseModel):
    id: int
    product_name: str
    quantity: float
    unit: Optional[str]
    unit_price: float
    total_price: float
    price_variance: Optional[float]
    flagged: bool

    class Config:
        from_attributes = True


class ProformaResponse(BaseModel):
    id: int
    supplier_id: int
    supplier_name: Optional[str] = None
    site_id: Optional[int]
    proforma_number: Optional[str]
    invoice_date: date
    delivery_date: Optional[date]
    total_amount: float
    currency: str
    status: str
    notes: Optional[str]

    class Config:
        from_attributes = True


@router.get("/", response_model=List[ProformaResponse])
async def list_proformas(
    months: int = Query(6, ge=1, le=24),
    supplier_id: Optional[int] = None,
    site_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """List proformas with filters"""
    cutoff = date.today() - timedelta(days=months * 30)

    query = (
        select(Proforma)
        .options(selectinload(Proforma.supplier), selectinload(Proforma.items))
        .where(Proforma.invoice_date >= cutoff)
    )

    if supplier_id:
        query = query.where(Proforma.supplier_id == supplier_id)
    if site_id:
        query = query.where(Proforma.site_id == site_id)

    query = query.order_by(Proforma.invoice_date.desc())

    result = await db.execute(query)
    proformas = result.scalars().all()

    response = []
    for p in proformas:
        flagged_count = sum(1 for item in (p.items or []) if item.flagged)
        item_count = len(p.items or [])
        resp = {
            "id": p.id,
            "supplier_id": p.supplier_id,
            "supplier_name": p.supplier.name if p.supplier else None,
            "site_id": p.site_id,
            "proforma_number": p.proforma_number,
            "invoice_date": p.invoice_date.isoformat(),
            "delivery_date": p.delivery_date.isoformat() if p.delivery_date else None,
            "total_amount": p.total_amount,
            "currency": p.currency,
            "status": p.status,
            "notes": p.notes,
            "item_count": item_count,
            "flagged_count": flagged_count,
        }
        response.append(resp)

    return response


# ── File Upload (XLSX + CSV) ──────────────────────────────────────────


def _parse_proforma_csv(raw: bytes) -> list[dict]:
    """Parse CSV bytes into row dicts. Tries multiple encodings."""
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1255", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        raise ValueError("Cannot decode file — try UTF-8 or cp1255 encoding")

    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict] = []
    for row in reader:
        cleaned = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        rows.append(cleaned)
    return rows


async def _parse_proforma_pdf(raw: bytes) -> list[dict]:
    """Parse PDF bytes into row dicts.

    Strategy 1: Extract tables with pdfplumber, validate headers are recognizable.
    Strategy 2: Extract full text and use Claude AI to parse items.
    Hebrew PDFs often have reversed RTL text in pdfplumber — if headers
    aren't recognizable after reversal attempt, we skip to AI.
    """
    import pdfplumber

    rows: list[dict] = []
    all_text_parts: list[str] = []

    # Known column keywords for validation
    _all_keywords = [
        "product", "item", "name", "description", "מוצר", "שם", "פריט", "תיאור",
        "quantity", "qty", "כמות", "price", "cost", "מחיר", "עלות", "תעריף",
        "unit", "uom", "יחידה", "יח",
    ]

    def _fix_reversed_hebrew(text: str) -> str:
        """Reverse Hebrew text that pdfplumber extracted backwards (RTL issue)."""
        if not text:
            return text
        hebrew_chars = sum(1 for c in text if "\u0590" <= c <= "\u05FF")
        if hebrew_chars > len(text) * 0.3:
            return text[::-1]
        return text

    def _headers_recognizable(headers: list[str]) -> bool:
        """Check if at least one header matches a known keyword."""
        for h in headers:
            hl = h.lower().strip()
            if any(k in hl for k in _all_keywords):
                return True
        return False

    # --- Strategy 1: Table extraction ---
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            all_text_parts.append(page.extract_text() or "")
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue
                # Try headers as-is, then reversed
                raw_headers = [str(h).strip().lower() if h else f"col_{j}" for j, h in enumerate(table[0])]
                fixed_headers = [_fix_reversed_hebrew(h) for h in raw_headers]
                headers = fixed_headers if _headers_recognizable(fixed_headers) else raw_headers

                if not _headers_recognizable(headers):
                    logger.info("PDF table headers not recognizable: %s — skipping to AI", headers[:5])
                    continue

                for data_row in table[1:]:
                    if all(c is None or str(c).strip() == "" for c in data_row):
                        continue
                    cleaned = {}
                    for j, cell in enumerate(data_row):
                        if j < len(headers):
                            val = str(cell).strip() if cell is not None else ""
                            cleaned[headers[j]] = _fix_reversed_hebrew(val)
                    rows.append(cleaned)

    if rows:
        # Validate pdfplumber rows have usable column headers before returning
        test_headers = list(rows[0].keys())
        name_col, qty_col, _, price_col = _detect_proforma_columns(test_headers)
        if name_col and (qty_col or price_col):
            return rows
        # Headers are mangled — discard and fall through to AI extraction
        logger.info("pdfplumber rows have unusable headers %s — falling through to AI", test_headers[:3])
        rows = []

    # --- Strategy 2: AI extraction from raw text ---
    full_text = "\n".join(all_text_parts).strip()
    if not full_text:
        raise ValueError("PDF has no extractable text (might be a scanned image)")

    from backend.services.claude_service import claude_service

    prompt = f"""Extract ALL line items from this Hebrew/English invoice/proforma text.
Return a JSON array of objects. Each object must have these keys:
- "מוצר" (product name)
- "כמות" (quantity, number)
- "מחיר" (unit price, number)
- "יחידה" (unit, e.g. ק"ג, יח', ליטר — if not clear use "יח'")

Skip subtotals, VAT, grand totals, headers, and empty rows.
Return ONLY the JSON array, no markdown, no explanation.

Invoice text:
{full_text}"""

    try:
        ai_response = await claude_service.generate_response(
            prompt=prompt,
            system_prompt="You are a precise invoice data extractor. Return only valid JSON arrays.",
        )
        text_result = ai_response.strip()
        # Strip markdown fences if present
        if text_result.startswith("```"):
            text_result = text_result.split("\n", 1)[1] if "\n" in text_result else text_result[3:]
            if text_result.endswith("```"):
                text_result = text_result[:-3]
            text_result = text_result.strip()

        parsed = json.loads(text_result)
        if isinstance(parsed, list):
            return [{k.strip().lower(): str(v).strip() for k, v in item.items()} for item in parsed]
    except Exception as e:
        logger.warning("AI PDF extraction failed: %s — falling back to raw text", e)

    raise ValueError(
        "Could not extract table data from PDF. "
        "Try converting to Excel first, or ensure the PDF contains text (not scanned images)."
    )


def _parse_proforma_xlsx(raw: bytes) -> list[dict]:
    """Parse XLSX bytes into row dicts using openpyxl.

    Handles FoodHouse-style proformas with:
    - Multiple title/header rows before data
    - Multiple invoice sections (חשבונית 1, חשבונית 2, ...)
    - Header rows containing כמות/מחיר keywords
    - Subtotal/VAT rows mixed in (סה"כ, מע"מ, סך הכל)

    Strategy: scan all rows, auto-detect header rows by keyword match,
    then collect data rows from each section.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)

    # Prefer the "חשבוניות" (invoices) sheet if it exists — FoodHouse proformas
    # have multiple sheets but the invoice data is always on חשבוניות
    target_sheet_names = ["חשבוניות", "invoices", "sheet1"]
    ws = None
    for name in target_sheet_names:
        if name in wb.sheetnames:
            ws = wb[name]
            break
    if ws is None:
        # Fallback: try sheets with כמות/מחיר headers, else use active
        for sname in wb.sheetnames:
            candidate = wb[sname]
            for row in candidate.iter_rows(values_only=True, max_row=30):
                cells_text = " ".join(str(c) for c in row if c is not None)
                if "כמות" in cells_text and "מחיר" in cells_text:
                    ws = candidate
                    break
            if ws is not None:
                break
    if ws is None:
        ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no usable sheet")

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("Excel file is empty")

    # --- Strategy 1: Try simple header-in-row-0 format first ---
    first_non_empty = None
    for i, row in enumerate(all_rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        non_empty = [c for c in cells if c]
        if len(non_empty) >= 2:
            first_non_empty = i
            break

    if first_non_empty is not None:
        test_headers = [str(c).strip().lower() if c is not None else "" for c in all_rows[first_non_empty]]
        name_kw = ["product", "item", "name", "description", "מוצר", "שם", "פריט", "תיאור"]
        qty_kw = ["כמות", "quantity", "qty"]
        price_kw = ["מחיר", "price", "cost", "עלות", "תעריף"]
        has_name = any(any(k in h for k in name_kw) for h in test_headers)
        has_qty = any(any(k in h for k in qty_kw) for h in test_headers)
        has_price = any(any(k in h for k in price_kw) for h in test_headers)

        # Simple format: first non-empty row has recognizable column headers
        if has_name and (has_qty or has_price):
            headers = [str(h).strip().lower() if h is not None else f"col_{j}" for j, h in enumerate(all_rows[first_non_empty])]
            rows: list[dict] = []
            for row in all_rows[first_non_empty + 1:]:
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                cleaned = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        cleaned[headers[j]] = str(cell).strip() if cell is not None else ""
                rows.append(cleaned)
            return rows

    # --- Strategy 2: FoodHouse multi-section format ---
    # Scan for rows containing כמות + מחיר as section headers
    skip_keywords = {"סה\"כ", "סהכ", "מע\"מ", "מעמ", "סך הכל", "סך הכול", "total", "vat", "subtotal"}
    section_header_indices: list[int] = []
    name_col_idx = -1
    qty_col_idx = -1
    price_col_idx = -1
    total_col_idx = -1

    for i, row in enumerate(all_rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        cells_lower = [c.lower() for c in cells]
        # Look for a row that has כמות and מחיר
        if any("כמות" in c for c in cells_lower) and any("מחיר" in c for c in cells_lower):
            section_header_indices.append(i)
            # Detect column positions from the FIRST header row found
            if name_col_idx == -1:
                for j, c in enumerate(cells_lower):
                    if "כמות" in c:
                        qty_col_idx = j
                    elif "מחיר" in c:
                        price_col_idx = j
                    elif 'סה"כ' in c or "סהכ" in c or "total" in c:
                        total_col_idx = j
                # Name column = the column just before כמות (where the description is)
                if qty_col_idx > 0:
                    name_col_idx = qty_col_idx - 1

    if not section_header_indices:
        # Fallback: treat first row as header
        headers = [str(h).strip().lower() if h is not None else f"col_{j}" for j, h in enumerate(all_rows[0])]
        rows = []
        for row in all_rows[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            cleaned = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    cleaned[headers[j]] = str(cell).strip() if cell is not None else ""
            rows.append(cleaned)
        return rows

    # Collect data rows from ALL sections
    result: list[dict] = []
    for sec_idx, header_row_idx in enumerate(section_header_indices):
        # Data rows start after the header row
        next_boundary = section_header_indices[sec_idx + 1] if sec_idx + 1 < len(section_header_indices) else len(all_rows)

        for i in range(header_row_idx + 1, next_boundary):
            row = all_rows[i]
            cells = [str(c).strip() if c is not None else "" for c in row]

            # Skip empty rows
            if all(c == "" for c in cells):
                continue

            # Skip subtotal/VAT/total rows
            row_text = " ".join(cells).lower()
            if any(sk in row_text for sk in skip_keywords):
                continue

            # Skip section title rows like "חשבונית 2"
            if any("חשבונית" in c for c in cells):
                continue

            # Extract product name
            product_name = cells[name_col_idx] if 0 <= name_col_idx < len(cells) else ""
            if not product_name:
                continue

            qty_str = cells[qty_col_idx] if 0 <= qty_col_idx < len(cells) else ""
            price_str = cells[price_col_idx] if 0 <= price_col_idx < len(cells) else ""
            total_str = cells[total_col_idx] if 0 <= total_col_idx < len(cells) else ""

            result.append({
                "מוצר": product_name,
                "כמות": qty_str,
                "מחיר": price_str,
                "סה\"כ": total_str,
            })

    return result


def _detect_proforma_columns(headers: list[str]) -> tuple[str | None, str | None, str | None, str | None]:
    """Auto-detect product_name, quantity, unit, unit_price columns from headers."""
    name_col = None
    qty_col = None
    unit_col = None
    price_col = None

    name_keywords = ["product", "item", "name", "description", "מוצר", "שם", "פריט", "תיאור"]
    qty_keywords = ["quantity", "qty", "amount", "count", "כמות"]
    unit_keywords = ["unit", "uom", "measure", "יחידה", "יח"]
    price_keywords = ["price", "cost", "rate", "מחיר", "עלות", "תעריף", "סכום"]
    total_keywords = ["total", "סה\"כ", "סהכ", "סך"]

    for h in headers:
        hl = h.lower().strip()
        if not name_col and any(k in hl for k in name_keywords):
            name_col = h
        elif not qty_col and any(k in hl for k in qty_keywords):
            qty_col = h
        elif not price_col and any(k in hl for k in price_keywords):
            price_col = h
        elif not unit_col and any(k in hl for k in unit_keywords):
            unit_col = h

    return name_col, qty_col, unit_col, price_col


async def _extract_and_save_meal_breakdown(
    raw: bytes, proforma_id: int, supplier_id: int, site_id: int, invoice_date: date, db: AsyncSession,
) -> bool:
    """Extract meal data from ריכוז הכנסות or ריכוז ארוחות sheet and save to MealBreakdown."""
    import openpyxl
    from backend.models.meal_breakdown import MealBreakdown

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

    # MUST use ריכוז הכנסות (has monthly qty in col D). ריכוז ארוחות has daily monetary values.
    ws = None
    qty_col = 4  # D column = כמות in ריכוז הכנסות
    for sheet_name in wb.sheetnames:
        if "ריכוז הכנסות" in sheet_name:
            ws = wb[sheet_name]
            break
    if ws is None:
        return False

    def get_qty(row: int) -> float:
        """Read quantity from column D (כמות) in ריכוז הכנסות."""
        v = ws.cell(row, 4).value
        return float(v) if isinstance(v, (int, float)) else 0.0

    # Find working days — search for "ימי עבודה" label, value is in the cell to the right or below
    working_days = 0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(r, c).value
            if cell_val and isinstance(cell_val, str) and "ימי עבודה" in cell_val:
                # Check cell below for the number
                below = ws.cell(r + 1, c).value
                if isinstance(below, (int, float)):
                    working_days = int(below)
                    break
                # Check cell to the right
                right = ws.cell(r, c + 1).value
                if isinstance(right, (int, float)):
                    working_days = int(right)
                    break
        if working_days:
            break

    month_start = date(invoice_date.year, invoice_date.month, 1)

    # Upsert: delete old breakdown for this proforma, then insert new
    existing = await db.execute(
        select(MealBreakdown).where(MealBreakdown.proforma_id == proforma_id)
    )
    old = existing.scalar_one_or_none()
    if old:
        await db.delete(old)

    vals = {
        "hp_meat": get_qty(5), "scitex_meat": get_qty(6),
        "evening_hp": get_qty(7), "evening_contractors": get_qty(8),
        "hp_dairy": get_qty(9), "scitex_dairy": get_qty(10),
        "supplement": get_qty(11),
        "contractors_meat": get_qty(14), "contractors_dairy": get_qty(15),
    }
    # Log row labels for debugging
    for r in [5, 6, 7, 8, 9, 10, 11, 14, 15]:
        label = ws.cell(r, 2).value or ws.cell(r, 1).value or ""
        logger.info("  ריכוז הכנסות row %d: '%s' = %s", r, str(label)[:50], get_qty(r))
    logger.info("  Working days = %d", working_days)

    breakdown = MealBreakdown(
        proforma_id=proforma_id,
        site_id=site_id,
        invoice_month=month_start,
        working_days=working_days,
        **vals,
    )
    db.add(breakdown)
    await db.commit()
    logger.info("Saved meal breakdown for proforma %d (site %d, month %s): %s", proforma_id, site_id, month_start, vals)
    return True


@router.post("/upload")
async def upload_proforma(
    file: UploadFile = File(...),
    supplier_id: int = Form(...),
    site_id: int = Form(None),
    invoice_date: str = Form(None),
    proforma_number: str = Form(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Upload an XLSX, CSV, or PDF file to create a proforma with line items."""
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv", "tsv", "txt", "pdf"):
        raise HTTPException(
            status_code=400,
            detail="Supported formats: Excel (.xlsx), CSV (.csv), or PDF (.pdf)",
        )

    # Verify supplier exists
    result = await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    raw = await file.read()
    try:
        if ext == "pdf":
            rows = await _parse_proforma_pdf(raw)
        elif ext in ("xlsx", "xls"):
            rows = _parse_proforma_xlsx(raw)
        else:
            rows = _parse_proforma_csv(raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty or has no data rows")

    # Detect columns
    headers = list(rows[0].keys())
    name_col, qty_col, unit_col, price_col = _detect_proforma_columns(headers)

    if not name_col:
        raise HTTPException(
            status_code=400,
            detail=f"Could not detect product name column. Headers found: {headers}",
        )
    if not price_col and not qty_col:
        raise HTTPException(
            status_code=400,
            detail=f"Could not detect quantity or price columns. Headers found: {headers}",
        )

    # Parse invoice date
    inv_date = date.today()
    if invoice_date:
        try:
            inv_date = date.fromisoformat(invoice_date)
        except ValueError:
            pass

    # Create proforma
    proforma = Proforma(
        supplier_id=supplier_id,
        site_id=site_id,
        proforma_number=proforma_number or None,
        invoice_date=inv_date,
        total_amount=0,
        currency="ILS",
        status="pending",
        notes=f"Uploaded from {file.filename}",
    )
    db.add(proforma)
    await db.flush()

    # Parse items
    items_created = 0
    skipped = 0
    total_amount = 0.0

    for row in rows:
        product_name = row.get(name_col, "").strip()
        if not product_name:
            skipped += 1
            continue

        # Parse quantity
        quantity = 1.0
        if qty_col and row.get(qty_col):
            try:
                quantity = float(row[qty_col].replace(",", ""))
            except (ValueError, AttributeError):
                quantity = 1.0

        # Parse unit price
        unit_price = 0.0
        if price_col and row.get(price_col):
            try:
                unit_price = float(row[price_col].replace(",", "").replace("₪", "").strip())
            except (ValueError, AttributeError):
                unit_price = 0.0

        # Parse unit
        unit = "unit"
        if unit_col and row.get(unit_col):
            unit = row[unit_col].strip() or "unit"

        total_price = quantity * unit_price
        total_amount += total_price

        item = ProformaItem(
            proforma_id=proforma.id,
            product_name=product_name,
            quantity=quantity,
            unit=unit,
            unit_price=unit_price,
            total_price=total_price,
            flagged=False,
        )
        db.add(item)
        items_created += 1

    proforma.total_amount = total_amount
    await db.commit()
    await db.refresh(proforma)

    # Auto-extract meal breakdown from FoodHouse proformas (ריכוז הכנסות sheet)
    meal_extracted = False
    if ext in ("xlsx", "xls"):
        try:
            meal_extracted = await _extract_and_save_meal_breakdown(
                raw, proforma.id, proforma.supplier_id, proforma.site_id or 1, inv_date, db
            )
        except Exception as e:
            logger.warning("Meal breakdown extraction skipped: %s", e)

    return {
        "id": proforma.id,
        "supplier_id": proforma.supplier_id,
        "supplier_name": supplier.name,
        "invoice_date": proforma.invoice_date.isoformat(),
        "total_amount": round(proforma.total_amount, 2),
        "items_created": items_created,
        "skipped": skipped,
        "status": "pending",
        "message": f"Created proforma with {items_created} items from {file.filename}",
        "meal_breakdown_extracted": meal_extracted,
    }


@router.get("/duplicates")
async def find_duplicate_proformas(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Find potential duplicate proformas (same supplier + site + invoice_date + similar total)."""
    result = await db.execute(
        select(Proforma)
        .options(selectinload(Proforma.supplier), selectinload(Proforma.items))
        .order_by(Proforma.supplier_id, Proforma.site_id, Proforma.invoice_date)
    )
    proformas = result.scalars().all()

    # Group by supplier_id + site_id + invoice_date
    groups: dict = {}
    for p in proformas:
        key = (p.supplier_id, p.site_id, p.invoice_date.isoformat() if p.invoice_date else "none")
        groups.setdefault(key, []).append(p)

    duplicates = []
    for key, group in groups.items():
        if len(group) > 1:
            duplicates.append({
                "supplier_name": group[0].supplier.name if group[0].supplier else "Unknown",
                "site_id": group[0].site_id,
                "invoice_date": key[2],
                "count": len(group),
                "proformas": [
                    {
                        "id": p.id,
                        "proforma_number": p.proforma_number,
                        "total_amount": p.total_amount,
                        "item_count": len(p.items) if p.items else 0,
                        "status": p.status,
                    }
                    for p in group
                ],
            })

    return {"duplicates": duplicates, "total_duplicate_groups": len(duplicates)}


@router.get("/generate-meal-summary")
async def generate_meal_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate ריכוז מספרי ארוחות Excel from MealBreakdown records (extracted from ריכוז הכנסות sheet)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    from backend.models.meal_breakdown import MealBreakdown

    result = await db.execute(select(MealBreakdown).order_by(MealBreakdown.invoice_month.asc()))
    breakdowns = result.scalars().all()
    if not breakdowns:
        raise HTTPException(
            status_code=404,
            detail="No meal breakdown data. Upload FoodHouse proforma Excels first (click Upload Excel with the original .xlsx files).",
        )

    by_key: dict[tuple[date, int], MealBreakdown] = {}
    months_set: set[date] = set()
    for b in breakdowns:
        months_set.add(b.invoice_month)
        by_key[(b.invoice_month, b.site_id)] = b
    sorted_months = sorted(months_set)

    def gv(b: MealBreakdown | None, field: str) -> int:
        if b is None:
            return 0
        return int(getattr(b, field, 0) or 0)

    def make_rows(sid: int, label: str):
        return [
            ("ימי עבודה", lambda m, s=sid: gv(by_key.get((m,s)), "working_days")),
            (f'סה"כ צהרים {label}', lambda m, s=sid: gv(by_key.get((m,s)),"hp_meat")+gv(by_key.get((m,s)),"scitex_meat")),
            (f'סה"כ חלבי {label}', lambda m, s=sid: gv(by_key.get((m,s)),"hp_dairy")+gv(by_key.get((m,s)),"scitex_dairy")),
            ('סה"כ קבלנים בשרי' if sid==1 else 'סה"כ קבלנים צהרים בשרי', lambda m, s=sid: gv(by_key.get((m,s)),"contractors_meat")),
            ('סה"כ קבלנים חלבי' if sid==1 else 'סה"כ קבלנים צהרים חלבי', lambda m, s=sid: gv(by_key.get((m,s)),"contractors_dairy")),
            (f'סה"כ ערב {label}', lambda m, s=sid: gv(by_key.get((m,s)),"evening_hp")+gv(by_key.get((m,s)),"evening_contractors")),
            ('תוספת למנה עיקרית(HP+קבלנים)', lambda m, s=sid: gv(by_key.get((m,s)),"supplement")),
        ]

    nz_rows = make_rows(1, "נס ציונה")
    kg_rows = make_rows(2, "קרית גת")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ריכוז ארוחות"
    ws.sheet_view.rightToLeft = True
    hf = Font(bold=True, size=11)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    sfill = PatternFill("solid", fgColor="E2EFDA")
    brd = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    ws.cell(1, 1, "ריכוז מספרי ארוחות").font = Font(bold=True, size=14)
    ws.cell(3, 1, "אתר").font = hf
    ws.cell(3, 2, "פרטים").font = hf
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    mhe = {1:"ינואר",2:"פברואר",3:"מרץ",4:"אפריל",5:"מאי",6:"יוני",7:"יולי",8:"אוגוסט",9:"ספטמבר",10:"אוקטובר",11:"נובמבר",12:"דצמבר"}
    for ci, m in enumerate(sorted_months):
        c = ci + 3
        cell = ws.cell(3, c, f"{mhe.get(m.month,str(m.month))} {m.year}")
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal="center"); cell.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 14

    r = 4
    ws.cell(r, 1, "נס ציונה").font = Font(bold=True)
    ws.cell(r, 1).fill = sfill
    for ri, (lbl, fn) in enumerate(nz_rows):
        ws.cell(r+ri, 2, lbl).border = brd
        for ci, m in enumerate(sorted_months):
            cell = ws.cell(r+ri, ci+3, fn(m)); cell.border = brd; cell.number_format = "#,##0"

    r2 = r + len(nz_rows) + 1
    ws.cell(r2, 1, "קרית גת").font = Font(bold=True)
    ws.cell(r2, 1).fill = sfill
    for ri, (lbl, fn) in enumerate(kg_rows):
        ws.cell(r2+ri, 2, lbl).border = brd
        for ci, m in enumerate(sorted_months):
            cell = ws.cell(r2+ri, ci+3, fn(m)); cell.border = brd; cell.number_format = "#,##0"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="meal_summary.xlsx"'},
    )


@router.get("/{proforma_id}")
async def get_proforma(
    proforma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get a single proforma with items"""
    result = await db.execute(
        select(Proforma)
        .options(
            selectinload(Proforma.supplier),
            selectinload(Proforma.site),
            selectinload(Proforma.items),
        )
        .where(Proforma.id == proforma_id)
    )
    proforma = result.scalar_one_or_none()

    if not proforma:
        raise HTTPException(status_code=404, detail="Proforma not found")

    return {
        "id": proforma.id,
        "supplier_name": proforma.supplier.name if proforma.supplier else None,
        "site_name": proforma.site.name if proforma.site else None,
        "proforma_number": proforma.proforma_number,
        "invoice_date": proforma.invoice_date.isoformat(),
        "delivery_date": proforma.delivery_date.isoformat() if proforma.delivery_date else None,
        "total_amount": proforma.total_amount,
        "currency": proforma.currency,
        "status": proforma.status,
        "notes": proforma.notes,
        "items": [
            {
                "id": item.id,
                "product_name": item.product_name,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_price": item.unit_price,
                "total_price": item.total_price,
                "price_variance": item.price_variance,
                "flagged": item.flagged,
            }
            for item in proforma.items
        ],
    }


@router.get("/{proforma_id}/items", response_model=List[ProformaItemResponse])
async def get_proforma_items(
    proforma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get items for a proforma"""
    result = await db.execute(
        select(ProformaItem).where(ProformaItem.proforma_id == proforma_id)
    )
    return result.scalars().all()


class ProformaItemCreate(BaseModel):
    product_name: str
    quantity: float
    unit: Optional[str] = None
    unit_price: float


class ProformaCreate(BaseModel):
    supplier_id: int
    site_id: Optional[int] = None
    proforma_number: Optional[str] = None
    invoice_date: date
    delivery_date: Optional[date] = None
    currency: str = "ILS"
    status: str = "pending"
    notes: Optional[str] = None
    items: List[ProformaItemCreate] = []


@router.post("/", response_model=ProformaResponse)
async def create_proforma(
    data: ProformaCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a new proforma with items"""
    result = await db.execute(
        select(Supplier).where(Supplier.id == data.supplier_id)
    )
    supplier = result.scalar_one_or_none()
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")

    total_amount = sum(item.quantity * item.unit_price for item in data.items)

    proforma = Proforma(
        supplier_id=data.supplier_id,
        site_id=data.site_id,
        proforma_number=data.proforma_number,
        invoice_date=data.invoice_date,
        delivery_date=data.delivery_date,
        total_amount=total_amount,
        currency=data.currency,
        status=data.status,
        notes=data.notes,
    )
    db.add(proforma)
    await db.flush()

    for item_data in data.items:
        item = ProformaItem(
            proforma_id=proforma.id,
            product_name=item_data.product_name,
            quantity=item_data.quantity,
            unit=item_data.unit,
            unit_price=item_data.unit_price,
            total_price=item_data.quantity * item_data.unit_price,
            flagged=False,
        )
        db.add(item)

    await db.commit()
    await db.refresh(proforma)

    return ProformaResponse(
        id=proforma.id,
        supplier_id=proforma.supplier_id,
        supplier_name=supplier.name,
        site_id=proforma.site_id,
        proforma_number=proforma.proforma_number,
        invoice_date=proforma.invoice_date,
        delivery_date=proforma.delivery_date,
        total_amount=proforma.total_amount,
        currency=proforma.currency,
        status=proforma.status,
        notes=proforma.notes,
    )


@router.get("/vendor-spending/summary")
async def get_vendor_spending(
    months: int = Query(12, ge=1, le=24),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get vendor spending summary with monthly breakdown"""
    cutoff = date.today() - timedelta(days=months * 30)

    # Get all proformas with supplier info
    result = await db.execute(
        select(Proforma)
        .options(selectinload(Proforma.supplier))
        .where(Proforma.invoice_date >= cutoff)
        .order_by(Proforma.invoice_date.asc())
    )
    proformas = result.scalars().all()

    # Group by supplier, then by month
    vendor_data: dict = {}
    monthly_totals: dict = {}

    for p in proformas:
        supplier_name = p.supplier.name if p.supplier else "Unknown"
        month_key = p.invoice_date.strftime("%Y-%m")

        if supplier_name not in vendor_data:
            vendor_data[supplier_name] = {}
        vendor_data[supplier_name][month_key] = (
            vendor_data[supplier_name].get(month_key, 0) + p.total_amount
        )

        monthly_totals[month_key] = monthly_totals.get(month_key, 0) + p.total_amount

    # Build monthly series with moving average
    sorted_months = sorted(monthly_totals.keys())
    monthly_series = []
    running_values: list[float] = []

    for month in sorted_months:
        total = monthly_totals[month]
        running_values.append(total)

        # 3-month moving average
        window = running_values[-3:] if len(running_values) >= 3 else running_values
        ma = sum(window) / len(window)

        monthly_series.append({
            "month": month,
            "total": round(total, 2),
            "moving_avg": round(ma, 2),
        })

    # Per-vendor totals
    vendor_totals = []
    for name, months_data in vendor_data.items():
        total = sum(months_data.values())
        vendor_totals.append({
            "supplier": name,
            "total": round(total, 2),
            "months": {k: round(v, 2) for k, v in sorted(months_data.items())},
        })

    vendor_totals.sort(key=lambda x: x["total"], reverse=True)

    return {
        "monthly_series": monthly_series,
        "vendor_totals": vendor_totals,
        "grand_total": round(sum(monthly_totals.values()), 2),
    }


@router.get("/vendor-analysis/{supplier_id}")
async def get_vendor_analysis(
    supplier_id: int,
    months: int = Query(12, ge=1, le=24),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get detailed analysis of a vendor's proformas: spending timeline, top products, price trends."""
    cutoff = date.today() - timedelta(days=months * 30)

    # Get proformas for this supplier
    result = await db.execute(
        select(Proforma)
        .options(selectinload(Proforma.items), selectinload(Proforma.site))
        .where(Proforma.supplier_id == supplier_id, Proforma.invoice_date >= cutoff)
        .order_by(Proforma.invoice_date.asc())
    )
    proformas = result.scalars().all()

    # Monthly spending timeline
    monthly_spending: dict[str, float] = {}
    for p in proformas:
        month_key = p.invoice_date.strftime("%Y-%m")
        monthly_spending[month_key] = monthly_spending.get(month_key, 0) + (p.total_amount or 0)

    spending_timeline = [
        {"month": m, "total": round(v, 2)}
        for m, v in sorted(monthly_spending.items())
    ]

    # Aggregate products across all proformas
    product_agg: dict[str, dict] = {}
    product_monthly: dict[str, dict[str, list[float]]] = {}

    for p in proformas:
        month_key = p.invoice_date.strftime("%Y-%m")
        for item in (p.items or []):
            name = (item.product_name or "").strip()
            if not name:
                continue
            if name not in product_agg:
                product_agg[name] = {"qty": 0, "total": 0, "count": 0, "prices": []}
            product_agg[name]["qty"] += float(item.quantity or 0)
            product_agg[name]["total"] += float(item.total_price or 0)
            product_agg[name]["count"] += 1
            if item.unit_price:
                product_agg[name]["prices"].append(float(item.unit_price))

            # For price trends
            if name not in product_monthly:
                product_monthly[name] = {}
            if month_key not in product_monthly[name]:
                product_monthly[name][month_key] = []
            if item.unit_price:
                product_monthly[name][month_key].append(float(item.unit_price))

    # Top products by spend
    top_products = sorted(
        [
            {
                "name": name,
                "total_spend": round(agg["total"], 2),
                "total_qty": round(agg["qty"], 2),
                "avg_price": round(sum(agg["prices"]) / len(agg["prices"]), 2) if agg["prices"] else 0,
                "invoice_count": agg["count"],
            }
            for name, agg in product_agg.items()
        ],
        key=lambda x: x["total_spend"],
        reverse=True,
    )[:30]

    # Price trends for top 10 products
    top_names = [p["name"] for p in top_products[:10]]
    all_months = sorted(set(m for pm in product_monthly.values() for m in pm))
    price_trends = []
    for name in top_names:
        monthly_prices = product_monthly.get(name, {})
        trend = {
            "name": name,
            "data": [
                {"month": m, "price": round(sum(monthly_prices[m]) / len(monthly_prices[m]), 2) if m in monthly_prices else None}
                for m in all_months
            ],
        }
        price_trends.append(trend)

    return {
        "proforma_count": len(proformas),
        "total_spend": round(sum(p.total_amount or 0 for p in proformas), 2),
        "spending_timeline": spending_timeline,
        "top_products": top_products,
        "price_trends": price_trends,
    }


def _normalize_product_name(name: str) -> str:
    """Normalize product name for matching (lowercase, stripped)."""
    return name.strip().lower()


@router.post("/{proforma_id}/compare-prices")
async def compare_proforma_prices(
    proforma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Compare proforma items against the supplier's latest price list.
    Updates price_variance and flagged fields on each ProformaItem.
    Returns the comparison results.
    """
    # Load the proforma with items
    result = await db.execute(
        select(Proforma)
        .options(selectinload(Proforma.items), selectinload(Proforma.supplier))
        .where(Proforma.id == proforma_id)
    )
    proforma = result.scalar_one_or_none()
    if not proforma:
        raise HTTPException(status_code=404, detail="Proforma not found")

    # Find the latest price list for this supplier (effective_date <= proforma date)
    pl_result = await db.execute(
        select(PriceList)
        .options(selectinload(PriceList.items).selectinload(PriceListItem.product))
        .where(
            PriceList.supplier_id == proforma.supplier_id,
            PriceList.effective_date <= proforma.invoice_date,
        )
        .order_by(PriceList.effective_date.desc())
        .limit(1)
    )
    price_list = pl_result.scalar_one_or_none()

    # If no price list before invoice date, try the latest one for this supplier
    if not price_list:
        pl_result = await db.execute(
            select(PriceList)
            .options(selectinload(PriceList.items).selectinload(PriceListItem.product))
            .where(PriceList.supplier_id == proforma.supplier_id)
            .order_by(PriceList.effective_date.desc())
            .limit(1)
        )
        price_list = pl_result.scalar_one_or_none()

    if not price_list:
        return {
            "proforma_id": proforma_id,
            "price_list_id": None,
            "price_list_date": None,
            "message": "No price list found for this supplier",
            "items": [
                {
                    "id": item.id,
                    "product_name": item.product_name,
                    "quantity": item.quantity,
                    "unit": item.unit,
                    "unit_price": item.unit_price,
                    "total_price": item.total_price,
                    "expected_price": None,
                    "price_variance": None,
                    "flagged": False,
                    "match_status": "no_price_list",
                }
                for item in proforma.items
            ],
            "summary": {"total_items": len(proforma.items), "matched": 0, "flagged": 0, "unmatched": len(proforma.items)},
        }

    # Build lookup: product name (normalized) → expected price
    price_lookup: dict[str, float] = {}
    for pl_item in price_list.items:
        if pl_item.product:
            price_lookup[_normalize_product_name(pl_item.product.name)] = pl_item.price
            if pl_item.product.hebrew_name:
                price_lookup[_normalize_product_name(pl_item.product.hebrew_name)] = pl_item.price

    # Compare each proforma item
    comparison_items = []
    matched_count = 0
    flagged_count = 0
    variance_threshold = 5.0  # flag if > 5% variance

    for item in proforma.items:
        normalized_name = _normalize_product_name(item.product_name)
        expected_price = price_lookup.get(normalized_name)

        if expected_price is not None:
            matched_count += 1
            variance_pct = round(
                ((item.unit_price - expected_price) / expected_price) * 100, 1
            ) if expected_price > 0 else 0.0

            is_flagged = abs(variance_pct) > variance_threshold

            # Update the item in DB
            item.price_variance = variance_pct
            item.flagged = is_flagged

            if is_flagged:
                flagged_count += 1

            comparison_items.append({
                "id": item.id,
                "product_name": item.product_name,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_price": round(item.unit_price, 2),
                "total_price": round(item.total_price, 2),
                "expected_price": round(expected_price, 2),
                "price_variance": variance_pct,
                "flagged": is_flagged,
                "match_status": "matched",
            })
        else:
            # No match found in price list
            item.price_variance = None
            item.flagged = False
            comparison_items.append({
                "id": item.id,
                "product_name": item.product_name,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_price": round(item.unit_price, 2),
                "total_price": round(item.total_price, 2),
                "expected_price": None,
                "price_variance": None,
                "flagged": False,
                "match_status": "unmatched",
            })

    await db.commit()

    # Sort: flagged first, then by absolute variance
    comparison_items.sort(
        key=lambda x: (
            0 if x["flagged"] else 1,
            -(abs(x["price_variance"]) if x["price_variance"] is not None else 0),
        )
    )

    return {
        "proforma_id": proforma_id,
        "price_list_id": price_list.id,
        "price_list_date": price_list.effective_date.isoformat(),
        "supplier_name": proforma.supplier.name if proforma.supplier else None,
        "items": comparison_items,
        "summary": {
            "total_items": len(comparison_items),
            "matched": matched_count,
            "flagged": flagged_count,
            "unmatched": len(comparison_items) - matched_count,
        },
    }


@router.delete("/{proforma_id}")
async def delete_proforma(
    proforma_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a proforma and all its line items."""
    result = await db.execute(select(Proforma).where(Proforma.id == proforma_id))
    proforma = result.scalar_one_or_none()
    if not proforma:
        raise HTTPException(status_code=404, detail="Proforma not found")

    # Delete meal breakdowns, items, then the proforma itself
    from sqlalchemy import delete as sql_delete
    from backend.models.meal_breakdown import MealBreakdown
    await db.execute(sql_delete(MealBreakdown).where(MealBreakdown.proforma_id == proforma_id))
    await db.execute(sql_delete(ProformaItem).where(ProformaItem.proforma_id == proforma_id))
    await db.execute(sql_delete(Proforma).where(Proforma.id == proforma_id))
    await db.commit()

    return {"message": f"Proforma #{proforma_id} deleted", "id": proforma_id}


class BulkDeleteRequest(BaseModel):
    ids: List[int]


@router.post("/bulk-delete")
async def bulk_delete_proformas(
    request: BulkDeleteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete multiple proformas at once (for cleaning up duplicates)."""
    from sqlalchemy import delete as sql_delete
    from backend.models.meal_breakdown import MealBreakdown

    deleted = []
    not_found = []
    for pid in request.ids:
        result = await db.execute(select(Proforma).where(Proforma.id == pid))
        if not result.scalar_one_or_none():
            not_found.append(pid)
            continue
        await db.execute(sql_delete(MealBreakdown).where(MealBreakdown.proforma_id == pid))
        await db.execute(sql_delete(ProformaItem).where(ProformaItem.proforma_id == pid))
        await db.execute(sql_delete(Proforma).where(Proforma.id == pid))
        deleted.append(pid)

    await db.commit()
    return {"deleted": deleted, "not_found": not_found}
