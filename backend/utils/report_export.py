"""
Excel export for the Report Generator.

Builds a single .xlsx with two sheets:
  1. "Report"  — the aggregated rows, RTL-aware, header styled, totals footer.
                 When `month` is in group_by and there's exactly one metric, the
                 sheet is pivoted: months become columns with row subtotals and
                 a grand-total row.
  2. "Filters" — the report config + filters (audit trail)
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.schemas.report import ReportConfig, ReportResponse


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


# Pretty labels for metric / column keys (English-friendly defaults)
COLUMN_LABELS = {
    "qty": "Quantity",
    "total": "Total",
    "unit_price": "Unit Price",
    "effective_price": "Effective ₪/unit",
    "budget": "Budget",
    "actual": "Actual",
    "variance": "Variance",
    "supplier": "Supplier",
    "site": "Site",
    "category": "Category",
    "product": "Product",
    "month": "Month",
    "shift": "Shift",
    "meal_type": "Meal Type",
    "family": "Family",
    "severity": "Severity",
}


def _label(key: str) -> str:
    return COLUMN_LABELS.get(key, key.replace("_", " ").title())


MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _should_pivot(config: ReportConfig) -> bool:
    """Pivot when `month` is a group-by dimension and there's exactly one
    visible metric. Multi-metric reports would need a column-per-metric-per-
    month layout; flat fallback keeps that case readable."""
    if "month" not in config.group_by:
        return False
    if len(config.metrics) != 1:
        return False
    return True


def _build_pivot(report: ReportResponse, config: ReportConfig) -> dict:
    """Reshape rows into a pivot layout with months as columns.

    Returns a dict with:
      row_dims      — list of non-month group_by keys (in order)
      months        — list of months present, chronological
      rows          — list of (key_tuple, {month: value}, row_total)
      subtotals     — {outer_value: {month: value, "total": value}} when 2+ row_dims
      grand_totals  — {month: value, "total": value}
      metric_name   — the single metric being pivoted
    """
    metric_name = config.metrics[0].name
    row_dims = [g for g in config.group_by if g != "month"]

    cells: dict[tuple, dict[str, float]] = {}
    months_present: set[str] = set()
    for r in report.rows:
        m = r.get("month", "—")
        months_present.add(m)
        key = tuple(r.get(d, "—") for d in row_dims)
        cells.setdefault(key, {})[m] = float(r.get(metric_name, 0) or 0)

    months = [m for m in MONTH_ORDER if m in months_present]
    months += sorted(m for m in months_present if m not in MONTH_ORDER)

    sorted_keys = sorted(cells.keys(), key=lambda k: tuple(str(v) for v in k))
    rows = []
    for key in sorted_keys:
        row_values = {m: cells[key].get(m, 0.0) for m in months}
        row_total = sum(row_values.values())
        rows.append((key, row_values, row_total))

    subtotals: dict = {}
    if len(row_dims) >= 2:
        for key, row_values, row_total in rows:
            outer = key[0]
            sub = subtotals.setdefault(outer, {**{m: 0.0 for m in months}, "total": 0.0})
            for m in months:
                sub[m] += row_values[m]
            sub["total"] += row_total

    grand_totals: dict = {**{m: 0.0 for m in months}, "total": 0.0}
    for _, row_values, row_total in rows:
        for m in months:
            grand_totals[m] += row_values[m]
        grand_totals["total"] += row_total

    return {
        "row_dims": row_dims,
        "months": months,
        "rows": rows,
        "subtotals": subtotals,
        "grand_totals": grand_totals,
        "metric_name": metric_name,
    }


def _style_metric_cell(cell, *, bold: bool = False, filled: bool = False) -> None:
    cell.number_format = "#,##0.##"
    cell.alignment = Alignment(horizontal="right")
    cell.border = BORDER
    if bold:
        cell.font = TOTAL_FONT
    if filled:
        cell.fill = TOTAL_FILL


def _write_pivot_sheet(ws, layout: dict, title: str) -> int:
    """Render the pivot layout. Returns the last row used."""
    row_dims = layout["row_dims"]
    months = layout["months"]
    rows = layout["rows"]
    subtotals = layout["subtotals"]
    grand_totals = layout["grand_totals"]
    metric_name = layout["metric_name"]

    total_cols = max(1, len(row_dims) + len(months) + 1)

    # Title (merged)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Header row
    header_row = 3
    col = 1
    for d in row_dims:
        cell = ws.cell(row=header_row, column=col, value=_label(d))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        col += 1
    for m in months:
        cell = ws.cell(row=header_row, column=col, value=m)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        col += 1
    cell = ws.cell(row=header_row, column=col, value=f"Total {_label(metric_name)}")
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER

    def write_subtotal(row_num: int, outer_value, sub: dict) -> None:
        c = 1
        first = ws.cell(row=row_num, column=c, value=f"{outer_value} Total")
        first.font = TOTAL_FONT
        first.fill = TOTAL_FILL
        first.border = BORDER
        c += 1
        for _ in range(len(row_dims) - 1):
            blank = ws.cell(row=row_num, column=c)
            blank.fill = TOTAL_FILL
            blank.border = BORDER
            c += 1
        for m in months:
            _style_metric_cell(ws.cell(row=row_num, column=c, value=sub[m]),
                               bold=True, filled=True)
            c += 1
        _style_metric_cell(ws.cell(row=row_num, column=c, value=sub["total"]),
                           bold=True, filled=True)

    # Data rows + subtotals on outer-dim change
    current_row = header_row + 1
    prev_outer = None
    for key, row_values, row_total in rows:
        outer = key[0] if row_dims else None
        if len(row_dims) >= 2 and prev_outer is not None and outer != prev_outer:
            write_subtotal(current_row, prev_outer, subtotals[prev_outer])
            current_row += 1

        col = 1
        for v in key:
            cell = ws.cell(row=current_row, column=col, value=v)
            cell.border = BORDER
            col += 1
        for m in months:
            _style_metric_cell(ws.cell(row=current_row, column=col, value=row_values[m]))
            col += 1
        _style_metric_cell(ws.cell(row=current_row, column=col, value=row_total),
                           bold=True)

        current_row += 1
        prev_outer = outer

    # Final subtotal for the last outer group
    if len(row_dims) >= 2 and prev_outer is not None:
        write_subtotal(current_row, prev_outer, subtotals[prev_outer])
        current_row += 1

    # Grand total row
    col = 1
    gt_label = ws.cell(row=current_row, column=col, value="Grand Total")
    gt_label.font = TOTAL_FONT
    gt_label.fill = TOTAL_FILL
    gt_label.border = BORDER
    col += 1
    for _ in range(max(0, len(row_dims) - 1)):
        blank = ws.cell(row=current_row, column=col)
        blank.fill = TOTAL_FILL
        blank.border = BORDER
        col += 1
    for m in months:
        _style_metric_cell(ws.cell(row=current_row, column=col, value=grand_totals[m]),
                           bold=True, filled=True)
        col += 1
    _style_metric_cell(ws.cell(row=current_row, column=col, value=grand_totals["total"]),
                       bold=True, filled=True)

    return current_row


def _write_flat_sheet(ws, report: ReportResponse, config: ReportConfig, title: str) -> int:
    """Original flat (row-per-record) layout. Returns last row used."""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    if report.columns:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report.columns))

    header_row = 3
    for ci, col in enumerate(report.columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=_label(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for ri, row in enumerate(report.rows, start=header_row + 1):
        for ci, col in enumerate(report.columns, 1):
            value = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.border = BORDER
            if isinstance(value, (int, float)):
                cell.number_format = "#,##0.##"
                cell.alignment = Alignment(horizontal="right")

    total_row = header_row + 1 + len(report.rows)
    if report.totals and report.columns:
        group_by_set = set(config.group_by)
        first_metric_idx = next(
            (i + 1 for i, c in enumerate(report.columns) if c not in group_by_set),
            len(report.columns) + 1,
        )
        label_cell = ws.cell(row=total_row, column=max(1, first_metric_idx - 1), value="Total")
        label_cell.font = TOTAL_FONT
        label_cell.fill = TOTAL_FILL
        for ci, col in enumerate(report.columns, 1):
            if col in report.totals:
                cell = ws.cell(row=total_row, column=ci, value=report.totals[col])
                _style_metric_cell(cell, bold=True, filled=True)

    return total_row


def build_xlsx(report: ReportResponse, config: ReportConfig, title: str) -> bytes:
    """Render a ReportResponse as .xlsx bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Report data ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True

    pivot = _should_pivot(config)
    if pivot:
        layout = _build_pivot(report, config)
        last_row = _write_pivot_sheet(ws, layout, title)
        col_count = len(layout["row_dims"]) + len(layout["months"]) + 1
        ws.freeze_panes = ws.cell(row=4, column=len(layout["row_dims"]) + 1)
    else:
        last_row = _write_flat_sheet(ws, report, config, title)
        col_count = len(report.columns)
        ws.freeze_panes = ws.cell(row=4, column=1)

    # Auto-size columns
    for ci in range(1, max(1, col_count) + 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, last_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 36)

    # ── Sheet 2: Filters / config audit ──────────────────────────────
    ws2 = wb.create_sheet("Filters")
    ws2.sheet_view.rightToLeft = True
    ws2.cell(row=1, column=1, value="Report Configuration").font = Font(bold=True, size=12)

    f = config.filters
    info_pairs = [
        ("Generated", date.today().isoformat()),
        ("Data Source", config.data_source),
        ("Year", f.year if f.year else "All"),
        ("From Month", f.from_month),
        ("To Month", f.to_month),
        ("Site ID", f.site_id if f.site_id else "All"),
        ("Supplier ID", f.supplier_id if f.supplier_id else "All"),
        ("Shift", f.shift or "—"),
        ("Category", f.category or "—"),
        ("Family", f.family or "—"),
        ("Product Search", f.product_name_like or "—"),
        ("Group By", ", ".join(config.group_by) or "—"),
        ("Metrics", ", ".join(f"{m.name}({m.agg})" for m in config.metrics) or "default"),
        ("Pivot (months as columns)", "Yes" if pivot else "No"),
        ("Chart Type", config.chart_type),
        ("Row Count", report.row_count),
    ]
    for ri, (k, v) in enumerate(info_pairs, start=3):
        ws2.cell(row=ri, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=ri, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
